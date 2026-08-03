from __future__ import annotations

from io import BytesIO
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from catalog_backend.db import normalize_product_data
from catalog_backend.fields import EXCEL_HEADERS, PRODUCT_FIELDS


def normalize_header(header: str | None) -> str:
    if not header:
        return ""
    compact = re.sub(r"\s+", "", str(header))
    return compact.replace("（", "(").replace("）", ")")


HEADER_KEY_LOOKUP = {
    normalize_header(field.excel_header): field.key
    for field in PRODUCT_FIELDS
}


def parse_workbook(file_obj) -> list[dict]:
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    raw_headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    normalized_headers = [normalize_header(header) for header in raw_headers]
    expected_headers = [normalize_header(header) for header in EXCEL_HEADERS]
    missing_headers = [header for header in expected_headers if header not in normalized_headers]
    if missing_headers:
        raise ValueError("模板表头不匹配，请确认使用参考模板的第一行表头。")

    products = []
    for row_index in range(2, worksheet.max_row + 1):
        row_payload = {}
        has_content = False
        for column_index, normalized in enumerate(normalized_headers, start=1):
            field_key = HEADER_KEY_LOOKUP.get(normalized)
            if not field_key:
                continue
            value = worksheet.cell(row_index, column_index).value
            if value not in (None, ""):
                has_content = True
            row_payload[field_key] = value
        if not has_content:
            continue
        products.append(normalize_product_data(row_payload))
    return products


def workbook_bytes(products: list[dict], visible_fields) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "商品资料"
    headers = [field.excel_header for field in visible_fields]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="5B4B3A")
    for product in products:
        row = [product.get(field.key) for field in visible_fields]
        worksheet.append(row)
    worksheet.freeze_panes = "A2"
    for index, field in enumerate(visible_fields, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = max(12, len(field.label) + 4)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

