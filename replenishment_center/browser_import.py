from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


REPORT_HEADERS = {
    "date": ("日期", "统计日期"),
    "barcode": ("条码", "条形码", "SKU"),
    "product_id": ("商品ID", "V_SKU"),
    "style": ("款号",),
    "outer_sku": ("货号",),
    "size": ("尺码名称", "尺码"),
    "brand": ("品牌名称", "品牌"),
    "name": ("商品名称",),
    "quantity": ("销售量", "销售数量", "销量"),
    "stock": ("在售库存", "可售库存", "库存"),
}

MASTER_HEADERS = {
    "external_spu": ("V_SPU", "商品ID"),
    "external_sku": ("V_SKU",),
    "style": ("款号",),
    "outer_sku": ("货号",),
    "barcode": ("条形码", "条码"),
    "name": ("商品名称",),
    "category": ("商品类目",),
    "brand": ("品牌名称", "品牌"),
    "standard_size": ("标准尺码",),
    "custom_size": ("自定义尺码",),
    "standard_color": ("标准颜色", "颜色"),
    "custom_color": ("自定义颜色",),
}


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _integer(value) -> int:
    text = _text(value).replace(",", "")
    if not text:
        return 0
    try:
        return max(0, int(round(float(text))))
    except ValueError as exc:
        raise ValueError(f"报表中存在无法识别的数字：{text}") from exc


def _date_text(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    for format_string in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], format_string).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"报表中存在无法识别的日期：{text}")


def _brand_matches(value: str, expected: str) -> bool:
    actual = "".join(_text(value).lower().split())
    target = "".join(_text(expected).lower().split())
    return bool(target and target in actual)


def _indexes(headers: list[str], aliases: dict[str, tuple[str, ...]], label: str) -> dict[str, int]:
    normalized = {_text(header).lower(): index for index, header in enumerate(headers) if _text(header)}
    result = {}
    missing = []
    for field, names in aliases.items():
        match = next((normalized[name.lower()] for name in names if name.lower() in normalized), None)
        if match is None:
            missing.append("/".join(names))
        else:
            result[field] = match
    if missing:
        raise ValueError(f"{label}缺少字段：{'、'.join(missing)}")
    return result


def _worksheet(workbook, preferred_name: str | None = None):
    sheet = workbook[preferred_name] if preferred_name and preferred_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    if sheet.calculate_dimension() == "A1:A1":
        sheet.reset_dimensions()
    return sheet


def _row_value(row: tuple, index: int):
    return row[index] if index < len(row) else None


def normalize_browser_reports(
    report_path: str | Path,
    master_path: str | Path,
    *,
    expected_brand: str = "马天奴",
    minimum_days: int = 14,
    maximum_unmatched_ratio: float = 0.05,
) -> dict:
    report_book = load_workbook(report_path, read_only=True, data_only=True)
    report_sheet = _worksheet(report_book)
    report_rows = report_sheet.iter_rows(values_only=True)
    report_headers = [_text(value) for value in next(report_rows, ())]
    report_indexes = _indexes(report_headers, REPORT_HEADERS, "魔方罗盘商品明细")

    report_skus: dict[str, dict] = {}
    sales_totals: dict[tuple[str, str], int] = defaultdict(int)
    inventory_latest: dict[str, tuple[str, int]] = {}
    report_dates: set[str] = set()
    source_row_count = 0
    for row in report_rows:
        brand = _text(_row_value(row, report_indexes["brand"]))
        if not _brand_matches(brand, expected_brand):
            continue
        barcode = _text(_row_value(row, report_indexes["barcode"]))
        if not barcode:
            continue
        sale_date = _date_text(_row_value(row, report_indexes["date"]))
        report_dates.add(sale_date)
        source_row_count += 1
        report_skus.setdefault(
            barcode,
            {
                "style_code": _text(_row_value(row, report_indexes["style"])),
                "outer_sku_id": _text(_row_value(row, report_indexes["outer_sku"])),
                "external_spu_id": _text(_row_value(row, report_indexes["product_id"])),
                "style_name": _text(_row_value(row, report_indexes["name"])),
                "report_size": _text(_row_value(row, report_indexes["size"])),
            },
        )
        quantity = _integer(_row_value(row, report_indexes["quantity"]))
        if quantity:
            sales_totals[(barcode, sale_date)] += quantity
        stock = _integer(_row_value(row, report_indexes["stock"]))
        previous = inventory_latest.get(barcode)
        if not previous or sale_date > previous[0]:
            inventory_latest[barcode] = (sale_date, stock)
        elif sale_date == previous[0]:
            inventory_latest[barcode] = (sale_date, max(previous[1], stock))
    report_book.close()

    if len(report_dates) < minimum_days:
        raise ValueError(f"销售报表只覆盖 {len(report_dates)} 个统计日期，至少需要 {minimum_days} 天。")
    if not report_skus:
        raise ValueError(f"销售报表中没有找到品牌“{expected_brand}”的条码数据。")

    master_book = load_workbook(master_path, read_only=True, data_only=True)
    master_sheet = _worksheet(master_book, "商品资料")
    master_rows = master_sheet.iter_rows(values_only=True)
    master_headers = [_text(value) for value in next(master_rows, ())]
    master_indexes = _indexes(master_headers, MASTER_HEADERS, "VIS 商品基础信息")
    master_by_barcode: dict[str, dict] = {}
    conflicting_barcodes: set[str] = set()
    for row in master_rows:
        barcode = _text(_row_value(row, master_indexes["barcode"]))
        if barcode not in report_skus:
            continue
        brand = _text(_row_value(row, master_indexes["brand"]))
        if not _brand_matches(brand, expected_brand):
            continue
        color = _text(_row_value(row, master_indexes["custom_color"])) or _text(
            _row_value(row, master_indexes["standard_color"])
        )
        size = _text(_row_value(row, master_indexes["custom_size"])) or _text(
            _row_value(row, master_indexes["standard_size"])
        )
        item = {
            "external_sku_id": barcode,
            "external_spu_id": _text(_row_value(row, master_indexes["external_spu"])),
            "outer_sku_id": _text(_row_value(row, master_indexes["outer_sku"])),
            "style_code": _text(_row_value(row, master_indexes["style"])),
            "style_name": _text(_row_value(row, master_indexes["name"])),
            "color_name": color,
            "size_name": size,
            "category": _text(_row_value(row, master_indexes["category"])),
        }
        existing = master_by_barcode.get(barcode)
        if existing and (existing["style_code"], existing["color_name"], existing["size_name"]) != (
            item["style_code"], item["color_name"], item["size_name"]
        ):
            conflicting_barcodes.add(barcode)
        elif color and size:
            master_by_barcode[barcode] = item
    master_book.close()

    missing_barcodes = sorted(set(report_skus).difference(master_by_barcode).union(conflicting_barcodes))
    unmatched_ratio = len(missing_barcodes) / max(1, len(report_skus))
    if unmatched_ratio > maximum_unmatched_ratio:
        raise ValueError(
            f"商品基础信息未匹配 {len(missing_barcodes)} / {len(report_skus)} 个条码，"
            f"比例 {unmatched_ratio:.1%} 超过允许的 {maximum_unmatched_ratio:.0%}。"
        )

    matched_barcodes = set(report_skus).intersection(master_by_barcode).difference(conflicting_barcodes)
    sku_by_key: dict[tuple[str, str, str], dict] = {}
    barcode_to_external_id: dict[str, str] = {}
    for barcode in sorted(matched_barcodes):
        master = dict(master_by_barcode[barcode])
        report = report_skus[barcode]
        master["style_code"] = master["style_code"] or report["style_code"]
        master["style_name"] = master["style_name"] or report["style_name"] or master["style_code"]
        master["outer_sku_id"] = master["outer_sku_id"] or report["outer_sku_id"]
        master["external_spu_id"] = master["external_spu_id"] or report["external_spu_id"]
        master["size_name"] = master["size_name"] or report["report_size"]
        key = (master["style_code"], master["color_name"], master["size_name"])
        if not all(key):
            continue
        existing = sku_by_key.setdefault(key, master)
        barcode_to_external_id[barcode] = existing["external_sku_id"]

    skus = list(sku_by_key.values())
    grouped_sales: dict[tuple[str, str], int] = defaultdict(int)
    for (barcode, sale_date), quantity in sales_totals.items():
        external_id = barcode_to_external_id.get(barcode)
        if external_id:
            grouped_sales[(external_id, sale_date)] += quantity

    sales = [
        {"external_sku_id": external_id, "sale_date": sale_date, "gross_units": quantity, "return_units": 0}
        for (external_id, sale_date), quantity in sorted(grouped_sales.items())
    ]
    grouped_inventory: dict[str, int] = defaultdict(int)
    for barcode, (_, stock) in inventory_latest.items():
        external_id = barcode_to_external_id.get(barcode)
        if external_id:
            grouped_inventory[external_id] += stock
    inventory = [
        {
            "external_sku_id": external_id,
            "on_hand": stock,
            "locked": 0,
            "defective": 0,
            "inbound": 0,
            "inbound_date": "",
        }
        for external_id, stock in sorted(grouped_inventory.items())
    ]
    latest_date = max(report_dates)
    return {
        "store": {"store_name": f"{expected_brand}店铺"},
        "skus": skus,
        "sales": sales,
        "inventory": inventory,
        "sales_through_date": latest_date,
        "inventory_snapshot_at": f"{latest_date} 23:59",
        "product_count": len({item["style_code"] for item in skus}),
        "order_count": 0,
        "stats": {
            "source_rows": source_row_count,
            "date_count": len(report_dates),
            "matched_barcodes": len(barcode_to_external_id),
            "sku_count": len(skus),
            "merged_barcodes": len(barcode_to_external_id) - len(skus),
            "sales_rows": len(sales),
            "inventory_rows": len(inventory),
            "unmatched_barcodes": len(missing_barcodes),
            "unmatched_ratio": unmatched_ratio,
        },
    }
