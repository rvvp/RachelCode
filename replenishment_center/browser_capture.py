from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import subprocess
import time
from datetime import datetime
from pathlib import Path
from urllib.error import URLError
from urllib.parse import urlparse
from urllib.request import urlopen

from openpyxl import load_workbook


DEFAULT_BACKEND_URL = "https://vis.vip.com/"
DEFAULT_CHROME = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
DEFAULT_NODE = "/Users/apple/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node"
DEFAULT_NODE_MODULES = "/Users/apple/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/node_modules"
REPORT_EXTENSIONS = {".xlsx", ".xls", ".csv", ".zip"}
OFFICIAL_REPORT_HOSTS = {"vis.vip.com", "compass.vip.com"}

FIELD_ALIASES = {
    "sales": {
        "date": ("日期", "统计日期", "销售日期", "下单时间", "订单创建时间", "支付时间"),
        "style": ("款号", "商品货号", "商家spu编码", "供应商款号", "外部spu"),
        "sku": ("sku", "货号", "商品条码", "商家sku编码", "条码"),
        "size": ("尺码", "尺寸"),
        "quantity": ("销售数量", "销售量", "净销售量", "销量", "商品数量", "订单数量", "数量"),
    },
    "inventory": {
        "style": ("款号", "商品货号", "商家spu编码", "供应商款号", "外部spu"),
        "sku": ("sku", "货号", "商品条码", "商家sku编码", "条码"),
        "size": ("尺码", "尺寸"),
        "stock": ("可售库存", "在售库存", "剩余库存", "库存数量", "库存"),
    },
    "master": {
        "style": ("款号",),
        "sku": ("条形码", "条码", "sku"),
        "color": ("自定义颜色", "标准颜色", "颜色"),
        "size": ("自定义尺码", "标准尺码", "尺码"),
    },
}


def is_official_report_url(url: str) -> bool:
    parsed = urlparse(str(url or "").strip())
    return parsed.scheme == "https" and (parsed.hostname or "").lower() in OFFICIAL_REPORT_HOSTS


def browser_paths(db_path: str | Path) -> dict:
    root = Path(db_path).resolve().parent
    return {
        "profile_dir": root / "vipshop_chrome_profile",
        "download_dir": root / "vipshop_browser_downloads",
        "archive_dir": root / "vipshop_raw_reports",
    }


def _debug_endpoint(port: int) -> str:
    return f"http://127.0.0.1:{port}"


def debug_endpoint_ready(port: int, timeout: float = 1.5) -> bool:
    try:
        with urlopen(f"{_debug_endpoint(port)}/json/version", timeout=timeout) as response:
            return response.status == 200
    except (URLError, TimeoutError):
        return False


def launch_dedicated_browser(db_path: str | Path, *, port: int = 9223, url: str = DEFAULT_BACKEND_URL) -> dict:
    paths = browser_paths(db_path)
    paths["profile_dir"].mkdir(parents=True, exist_ok=True)
    paths["download_dir"].mkdir(parents=True, exist_ok=True)
    chrome = os.environ.get("REPLENISH_CHROME_PATH", DEFAULT_CHROME)
    if not Path(chrome).exists():
        raise ValueError("未找到 Google Chrome，无法启动唯品专用登录会话。")
    if not debug_endpoint_ready(port):
        subprocess.Popen(
            [
                chrome,
                f"--remote-debugging-port={port}",
                "--remote-allow-origins=*",
                f"--user-data-dir={paths['profile_dir']}",
                "--no-first-run",
                "--no-default-browser-check",
                "--new-window",
                url,
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
        deadline = time.monotonic() + 12
        while time.monotonic() < deadline and not debug_endpoint_ready(port):
            time.sleep(0.25)
    if not debug_endpoint_ready(port):
        raise ValueError("专用 Chrome 已启动，但调试连接尚未就绪。")
    return run_browser_worker(db_path, port=port, action="open", url=url)


def run_browser_worker(
    db_path: str | Path,
    *,
    port: int = 9223,
    action: str = "status",
    url: str = "",
    start_date: str = "",
    end_date: str = "",
    brand: str = "",
) -> dict:
    if not debug_endpoint_ready(port):
        return {
            "ok": False,
            "sessionStatus": "browser_closed",
            "loginRequired": True,
            "current": {"url": "", "title": ""},
            "message": "唯品专用 Chrome 尚未启动。",
        }
    paths = browser_paths(db_path)
    paths["download_dir"].mkdir(parents=True, exist_ok=True)
    node = os.environ.get("REPLENISH_NODE_BIN", DEFAULT_NODE)
    node_modules = os.environ.get("REPLENISH_NODE_MODULES", DEFAULT_NODE_MODULES)
    worker = Path(__file__).resolve().parent.parent / "scripts" / "vipshop_browser_worker.js"
    if not Path(node).exists() or not Path(node_modules).exists():
        raise ValueError("缺少浏览器自动化运行时，请配置 REPLENISH_NODE_BIN 和 REPLENISH_NODE_MODULES。")
    command = {
        "endpoint": _debug_endpoint(port),
        "action": action,
        "url": url,
        "downloadDir": str(paths["download_dir"]),
        "startDate": start_date,
        "endDate": end_date,
        "brand": brand,
    }
    environment = os.environ.copy()
    environment["NODE_PATH"] = node_modules
    environment["VIPSHOP_BROWSER_COMMAND"] = json.dumps(command, ensure_ascii=False)
    result = subprocess.run(
        [node, str(worker)],
        env=environment,
        capture_output=True,
        text=True,
        timeout=360 if action == "export_product_detail" else 45,
        check=False,
    )
    if result.returncode != 0:
        raise ValueError(f"浏览器会话检查失败：{result.stderr.strip() or '未知错误'}")
    payload = json.loads(result.stdout)
    payload["message"] = "唯品后台需要重新登录。" if payload.get("loginRequired") else "唯品后台会话可用。"
    return payload


def _normalize_header(value) -> str:
    return "".join(str(value or "").strip().lower().split()).replace("（", "(").replace("）", ")")


def inspect_report(path: str | Path, kind: str) -> dict:
    report_path = Path(path)
    if report_path.suffix.lower() == ".xlsx":
        workbook = load_workbook(report_path, read_only=True, data_only=True)
        sheet = workbook["商品资料"] if kind == "master" and "商品资料" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        reset_dimensions = sheet.calculate_dimension() == "A1:A1"
        if reset_dimensions:
            sheet.reset_dimensions()
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, ())]
        row_count = sum(1 for row in rows if any(value not in (None, "") for value in row)) if reset_dimensions else max(0, sheet.max_row - 1)
        sheet_names = workbook.sheetnames
        workbook.close()
    elif report_path.suffix.lower() == ".csv":
        with report_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            reader = csv.reader(file_obj)
            headers = next(reader, [])
            row_count = sum(1 for _ in reader)
        sheet_names = ["CSV"]
    else:
        return {"supported": False, "headers": [], "row_count": 0, "sheet_names": [], "matched": {}, "missing": ["需解压或转换为 xlsx/csv"]}
    normalized = {_normalize_header(header): header for header in headers if header}
    matched = {}
    for field, aliases in FIELD_ALIASES[kind].items():
        for normalized_header, original in normalized.items():
            if any(_normalize_header(alias) in normalized_header for alias in aliases):
                matched[field] = original
                break
    missing = [field for field in FIELD_ALIASES[kind] if field not in matched]
    return {
        "supported": True,
        "headers": headers,
        "row_count": row_count,
        "sheet_names": sheet_names,
        "matched": matched,
        "missing": missing,
    }


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def archive_report(db_path: str | Path, source: str | Path, kind: str) -> Path:
    source_path = Path(source)
    paths = browser_paths(db_path)
    target_dir = paths["archive_dir"] / datetime.now().strftime("%Y%m%d")
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{kind}-{datetime.now().strftime('%H%M%S')}-{source_path.name}"
    shutil.copy2(source_path, target)
    return target


def candidate_downloads(db_path: str | Path, started_timestamp: float) -> list[Path]:
    download_dir = browser_paths(db_path)["download_dir"]
    if not download_dir.exists():
        return []
    candidates = []
    for path in download_dir.iterdir():
        if not path.is_file() or path.suffix.lower() not in REPORT_EXTENSIONS or path.name.endswith(".crdownload"):
            continue
        if path.stat().st_mtime >= started_timestamp and path.stat().st_size > 0:
            candidates.append(path)
    return sorted(candidates, key=lambda item: item.stat().st_mtime)
