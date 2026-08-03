from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from replenishment_center import db


SKU_HEADERS = ["款号", "款名", "颜色", "尺码", "品类", "供应商", "供应周期(天)", "最小起订量", "装箱倍数", "默认尺码占比", "核心尺码"]
SALES_HEADERS = ["日期", "平台", "店铺", "款号", "颜色", "尺码", "销售数量", "退货数量"]
INVENTORY_HEADERS = ["快照时间", "平台", "店铺", "款号", "颜色", "尺码", "实物库存", "锁定库存", "残次库存", "在途数量", "预计到货日期"]


def _text(value) -> str:
    return str(value or "").strip()


def _date_text(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    if not text:
        return ""
    for format_string in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], format_string).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"无法识别日期：{text}")


def _number(value, *, integer: bool = True):
    if value in (None, ""):
        return 0 if integer else 0.0
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"无法识别数字：{value}") from exc
    return int(round(number)) if integer else number


def _rows(worksheet, headers: list[str]):
    actual = [_text(cell.value) for cell in worksheet[1]]
    missing = [header for header in headers if header not in actual]
    if missing:
        raise ValueError(f"工作表“{worksheet.title}”缺少表头：{'、'.join(missing)}")
    indexes = {header: actual.index(header) for header in headers}
    for row_no, cells in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in cells):
            continue
        yield row_no, {header: cells[indexes[header]] if indexes[header] < len(cells) else None for header in headers}


def import_data_workbook(db_path: str | Path, file_obj, user_id: int | None = None) -> dict:
    workbook = load_workbook(file_obj, data_only=True)
    required_sheets = {"SKU资料", "销售数据", "库存及在途"}
    missing_sheets = required_sheets.difference(workbook.sheetnames)
    if missing_sheets:
        raise ValueError(f"导入文件缺少工作表：{'、'.join(sorted(missing_sheets))}")
    settings = db.get_settings(db_path)
    counts = {"sku": 0, "sales": 0, "inventory": 0}
    latest_sale_date = ""
    latest_snapshot = ""
    with db.get_connection(db_path) as connection:
        for row_no, row in _rows(workbook["SKU资料"], SKU_HEADERS):
            style_code = _text(row["款号"])
            color_name = _text(row["颜色"])
            size_name = _text(row["尺码"])
            if not style_code or not color_name or not size_name:
                raise ValueError(f"SKU资料第 {row_no} 行缺少款号、颜色或尺码。")
            share = _number(row["默认尺码占比"], integer=False)
            if share > 1:
                share /= 100
            connection.execute(
                """
                INSERT INTO skus(
                    store_id, style_code, style_name, color_name, size_name, category, supplier,
                    lead_time_days, moq, pack_size, default_size_share, core_size
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(store_id, style_code, color_name, size_name) DO UPDATE SET
                    style_name = excluded.style_name, category = excluded.category, supplier = excluded.supplier,
                    lead_time_days = excluded.lead_time_days, moq = excluded.moq,
                    pack_size = excluded.pack_size, default_size_share = excluded.default_size_share,
                    core_size = excluded.core_size
                """,
                (
                    settings["store_id"], style_code, _text(row["款名"]), color_name, size_name,
                    _text(row["品类"]), _text(row["供应商"]), max(0, _number(row["供应周期(天)"])),
                    max(0, _number(row["最小起订量"])), max(1, _number(row["装箱倍数"])),
                    max(0.0, share), 1 if _text(row["核心尺码"]).lower() in {"是", "1", "true", "y", "yes"} else 0,
                ),
            )
            counts["sku"] += 1

        sku_lookup = {
            (row["style_code"], row["color_name"], row["size_name"]): row["id"]
            for row in connection.execute("SELECT * FROM skus WHERE store_id = ?", (settings["store_id"],))
        }
        for row_no, row in _rows(workbook["销售数据"], SALES_HEADERS):
            key = (_text(row["款号"]), _text(row["颜色"]), _text(row["尺码"]))
            sku_id = sku_lookup.get(key)
            if not sku_id:
                raise ValueError(f"销售数据第 {row_no} 行的 SKU 未在“SKU资料”中定义：{' / '.join(key)}")
            sale_date = _date_text(row["日期"])
            gross = max(0, _number(row["销售数量"]))
            returns = max(0, _number(row["退货数量"]))
            connection.execute(
                """
                INSERT INTO sales_daily(sku_id, sale_date, gross_units, return_units, net_units)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(sku_id, sale_date) DO UPDATE SET gross_units = excluded.gross_units,
                    return_units = excluded.return_units, net_units = excluded.net_units
                """,
                (sku_id, sale_date, gross, returns, max(0, gross - returns)),
            )
            latest_sale_date = max(latest_sale_date, sale_date)
            counts["sales"] += 1

        for row_no, row in _rows(workbook["库存及在途"], INVENTORY_HEADERS):
            key = (_text(row["款号"]), _text(row["颜色"]), _text(row["尺码"]))
            sku_id = sku_lookup.get(key)
            if not sku_id:
                raise ValueError(f"库存及在途第 {row_no} 行的 SKU 未在“SKU资料”中定义：{' / '.join(key)}")
            raw_snapshot = row["快照时间"]
            if isinstance(raw_snapshot, datetime):
                snapshot_at = raw_snapshot.strftime("%Y-%m-%d %H:%M")
            else:
                snapshot_at = _text(raw_snapshot)
            if not snapshot_at:
                raise ValueError(f"库存及在途第 {row_no} 行缺少快照时间。")
            inbound_date = _date_text(row["预计到货日期"]) if row["预计到货日期"] not in (None, "") else ""
            connection.execute(
                """
                INSERT INTO inventory_current(sku_id, snapshot_at, on_hand, locked, defective, inbound, inbound_date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(sku_id) DO UPDATE SET snapshot_at = excluded.snapshot_at, on_hand = excluded.on_hand,
                    locked = excluded.locked, defective = excluded.defective, inbound = excluded.inbound,
                    inbound_date = excluded.inbound_date
                """,
                (
                    sku_id, snapshot_at, max(0, _number(row["实物库存"])), max(0, _number(row["锁定库存"])),
                    max(0, _number(row["残次库存"])), max(0, _number(row["在途数量"])), inbound_date,
                ),
            )
            latest_snapshot = max(latest_snapshot, snapshot_at)
            counts["inventory"] += 1
        cursor = connection.execute(
            """
            INSERT INTO sync_runs(store_id, source, status, sales_through_date, inventory_snapshot_at, row_count, message, created_at)
            VALUES (?, 'excel_upload', 'success', ?, ?, ?, '人工 Excel 数据导入成功。', ?)
            """,
            (settings["store_id"], latest_sale_date, latest_snapshot, sum(counts.values()), db.utc_now()),
        )
        connection.execute(
            "INSERT INTO audit_events(user_id, action, object_type, object_id, detail, created_at) VALUES (?, 'excel_import', 'sync_run', ?, ?, ?)",
            (user_id, cursor.lastrowid, str(counts), db.utc_now()),
        )
    return counts


def data_template_bytes() -> bytes:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    examples = {
        "SKU资料": [
            SKU_HEADERS,
            ["MTN260701", "轻薄针织开衫", "雾灰", "M", "针织衫", "嘉兴锦尚", 10, 80, 2, 36, "是"],
        ],
        "销售数据": [
            SALES_HEADERS,
            [date.today(), "唯品会", "马天奴店铺", "MTN260701", "雾灰", "M", 8, 1],
        ],
        "库存及在途": [
            INVENTORY_HEADERS,
            [datetime.now().replace(microsecond=0), "唯品会", "马天奴店铺", "MTN260701", "雾灰", "M", 24, 2, 0, 20, date.today()],
        ],
    }
    for sheet_name, rows in examples.items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
        _style_sheet(sheet)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def plan_workbook_bytes(plan: dict, items: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "补货明细"
    headers = [
        "计划编号", "平台", "店铺", "款号", "货号", "款名", "颜色", "尺码", "核心尺码", "命中条件", "连续有销量天数",
        "近7天销量", "近14天销量", "可售库存", "在途", "预计可售天数", "风险",
        "系统建议", "商品部确认", "调整原因", "跟单确认", "预计下单日期", "预计到货日期", "跟单状态", "跟单备注",
    ]
    sheet.append(headers)
    risk_labels = {"critical": "7天内缺货", "warning": "14天内缺货", "watch": "库存关注", "healthy": "健康", "no_sales": "暂无销量"}
    followup_labels = {"pending": "待处理", "confirmed": "已确认", "limited": "供应受限", "ordered": "已下单", "arrived": "已到货"}
    for item in items:
        condition_labels = {"condition_1": "条件1", "condition_2": "条件2"}
        selection_reason = "、".join(
            condition_labels.get(reason, reason)
            for reason in str(item.get("selection_reason") or "").split(",")
            if reason
        )
        sheet.append(
            [
                plan["plan_no"], plan["platform_name"], plan["store_name"], item["style_code"], item.get("outer_sku_id") or "", item["style_name"],
                item["color_name"], item["size_name"], "是" if item["core_size"] else "否", selection_reason,
                item.get("consecutive_sales_days") or 0, item["sales_7"], item["sales_14"],
                item["sellable"], item["inbound"], round(item["coverage_days"], 1) if item["coverage_days"] is not None else "-",
                risk_labels.get(item["risk_level"], item["risk_level"]), item["suggested_qty"], item["confirmed_qty"],
                item["adjustment_reason"], item["followup_qty"] if item["followup_qty"] is not None else "",
                item["expected_order_date"], item["expected_arrival_date"], followup_labels.get(item["followup_status"], item["followup_status"]),
                item["followup_note"],
            ]
        )
    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    summary = workbook.create_sheet("计算口径")
    summary.append(["项目", "内容"])
    summary.append(["计划编号", plan["plan_no"]])
    summary.append(["销售数据截至", plan["sales_through_date"]])
    summary.append(["库存快照", plan["inventory_snapshot_at"]])
    summary.append(["目标覆盖", f"{plan['target_days']} 天"])
    summary.append(["安全库存", f"{plan['safety_days']} 天"])
    summary.append(["货号筛选条件1", f"近7天销量 >= {plan.get('min_sales_7', 5)} 且近14天销量 >= {plan.get('min_sales_14', 10)}"])
    summary.append(["货号筛选条件2", f"最近14天内连续有销量 >= {plan.get('min_consecutive_sales_days', 3)} 天"])
    summary.append(["筛选关系", f"条件1 或 条件2；并且库存支撑 <= {plan.get('max_coverage_days', 14):g} 天"])
    summary.append(["预测日均", "近7日日均 × 60% + 近14日日均 × 40%"])
    summary.append(["尺码配比", "实际14天销量占比与默认尺码曲线平滑计算"])
    summary.append(["补货量", "目标覆盖需求 + 安全库存 - 可售库存 - 有效在途，并按装箱倍数取整"])
    _style_sheet(summary)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F5B4B")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D7DDD9"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    for index, column in enumerate(sheet.columns, start=1):
        width = min(28, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 26
