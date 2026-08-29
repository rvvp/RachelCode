from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import math
import posixpath
from pathlib import Path
import re
from typing import Callable
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill
from PIL import Image as PillowImage
from PIL import ImageOps

from catalog_backend.db import normalize_optional_date_text, normalize_product_data
from catalog_backend.fields import EXCEL_HEADERS, PRODUCT_FIELDS


CATALOG_EXPORT_HIDDEN_FIELD_KEYS = {
    "size_f",
    "size_s",
    "size_m",
    "size_l",
    "size_xl",
    "size_2xl",
    "size_3xl",
    "total_quantity",
}
EXPORT_IMAGE_MAX_SIDE = 120
EXPORT_IMAGE_MAX_PIXELS = 16_000_000


def normalize_header(header: str | None) -> str:
    if not header:
        return ""
    compact = re.sub(r"\s+", "", str(header))
    return compact.replace("（", "(").replace("）", ")")


HEADER_KEY_LOOKUP = {
    normalize_header(field.excel_header): field.key
    for field in PRODUCT_FIELDS
}
HEADER_KEY_LOOKUP.update(
    {
        normalize_header("品牌名称"): "brand_name",
        normalize_header("颜色名称"): "color_name",
        normalize_header("洗涤方式 （英文）"): "washing_method_en",
        normalize_header("洗涤方式（英文）"): "washing_method_en",
    }
)
LEGACY_COMPOSITION_HEADER = normalize_header("成分")
SUPPLIER_SETTLEMENT_HEADERS = [
    "供应商编码",
    "供应商名称",
    "开票抬头",
    "应付金额",
    "支付状态",
    "支付日期",
    "备注",
]
SUPPLIER_SETTLEMENT_HEADER_KEY_LOOKUP = {
    normalize_header("供应商编码"): "supplier_code",
    normalize_header("供应商名称"): "supplier_name",
    normalize_header("开票抬头"): "invoice_name",
    normalize_header("开票名称"): "invoice_name",
    normalize_header("应付金额"): "amount_due",
    normalize_header("支付状态"): "payment_status",
    normalize_header("支付日期"): "payment_date",
    normalize_header("备注"): "note",
}
SUPPLIER_MASTER_HEADERS = [
    "供应商编号",
    "供应商名称",
    "供应链经理",
]
SUPPLIER_MASTER_HEADER_KEY_LOOKUP = {
    normalize_header("供应商编号"): "supplier_code",
    normalize_header("供应商名称"): "supplier_name",
    normalize_header("供应链经理"): "supply_chain_manager",
}
SUPPLIER_BILL_HEADERS = [
    "供应商编号",
    "供应商名称",
    "模式",
    "供应链经理",
    "供应商款号",
    "品牌名称",
    "款色",
    "数量",
    "含税价",
    "结算金额",
]
SUPPLIER_BILL_EXPORT_HEADERS = ["所属月份", *SUPPLIER_BILL_HEADERS]
SUPPLIER_BILL_HEADER_KEY_LOOKUP = {
    normalize_header("供应商编号"): "supplier_code",
    normalize_header("供应商名称"): "supplier_name",
    normalize_header("模式"): "mode",
    normalize_header("供应链经理"): "supply_chain_manager",
    normalize_header("供应商款号"): "supplier_style_code",
    normalize_header("品牌名称"): "brand_name",
    normalize_header("款色"): "style_color",
    normalize_header("数量"): "quantity",
    normalize_header("含税价"): "tax_included_price",
    normalize_header("结算金额"): "settlement_amount",
}

BRAND_BILL_CHANNEL_ALIASES = {
    "天猫": "天猫",
    "天猫合计": "天猫",
    "唯品会": "唯品会",
    "唯品": "唯品会",
    "抖音": "抖音",
    "小程序": "小程序",
    "合计": "合计",
}
BRAND_BILL_HEADERS = [
    "年月",
    "平台",
    "店铺",
    "销售数量",
    "销售金额",
    "销售数量",
    "销售金额",
    "销量占比",
    "销额占比",
    "销售数量",
    "销售金额",
    "销量占比",
    "销额占比",
]
IMAGE_MAPPING_HEADER_KEY_LOOKUP = {
    normalize_header("款色"): "style_color",
    normalize_header("图片"): "image_filename",
    normalize_header("图片文件"): "image_filename",
    normalize_header("图片文件名"): "image_filename",
    normalize_header("图片名称"): "image_filename",
    normalize_header("图片名"): "image_filename",
    normalize_header("文件名"): "image_filename",
}
IMAGE_FORMULA_PATTERN = re.compile(
    r"DISPIMG\s*\(\s*[\"'](?P<image_id>[^\"']+)[\"']",
    re.IGNORECASE,
)
RELATIONSHIP_EMBED_ATTRIBUTE = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
EMBEDDED_IMAGE_CONTENT_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".gif": "image/gif",
}


def parse_workbook(file_obj) -> list[dict]:
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    raw_headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    normalized_headers = [normalize_header(header) for header in raw_headers]
    material_header = normalize_header("材质")
    has_material_header = material_header in normalized_headers
    recognized_headers = [
        header for header in normalized_headers
        if header in HEADER_KEY_LOOKUP or header == LEGACY_COMPOSITION_HEADER
    ]
    required_headers = [
        normalize_header("商品名称"),
        normalize_header("款号"),
    ]
    missing_required_headers = [header for header in required_headers if header not in normalized_headers]
    if missing_required_headers or not recognized_headers:
        raise ValueError("模板表头不匹配，请确认使用参考模板的第一行表头。")

    products = []
    for row_index in range(2, worksheet.max_row + 1):
        row_payload = {}
        has_content = False
        for column_index, normalized in enumerate(normalized_headers, start=1):
            field_key = HEADER_KEY_LOOKUP.get(normalized)
            if normalized == LEGACY_COMPOSITION_HEADER:
                # Keep old templates usable without reintroducing a duplicate field.
                field_key = None if has_material_header else "material"
            if not field_key:
                continue
            value = worksheet.cell(row_index, column_index).value
            if value not in (None, ""):
                has_content = True
            row_payload[field_key] = value
        if not has_content:
            continue
        products.append(normalize_product_data(row_payload))
    return products


def _parse_image_mapping_rows(worksheet, embedded_images_by_id: dict[str, dict] | None = None) -> tuple[list[dict], list[dict]]:
    raw_headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    normalized_headers = [normalize_header(header) for header in raw_headers]
    required_headers = [
        normalize_header("款色"),
        normalize_header("图片文件名"),
    ]
    if required_headers[0] not in normalized_headers:
        raise ValueError("图片映射 Excel 缺少“款色”表头。")
    if not any(IMAGE_MAPPING_HEADER_KEY_LOOKUP.get(header) == "image_filename" for header in normalized_headers):
        raise ValueError("图片映射 Excel 缺少“图片”或“图片文件名”表头。")

    rows = []
    referenced_embedded_images = []
    for row_index in range(2, worksheet.max_row + 1):
        payload = {}
        has_content = False
        for column_index, normalized in enumerate(normalized_headers, start=1):
            field_key = IMAGE_MAPPING_HEADER_KEY_LOOKUP.get(normalized)
            if not field_key:
                continue
            value = worksheet.cell(row_index, column_index).value
            if value not in (None, ""):
                has_content = True
            payload[field_key] = value
        if not has_content:
            continue
        style_color = str(payload.get("style_color") or "").strip()
        image_filename = str(payload.get("image_filename") or "").strip()
        if not style_color:
            raise ValueError(f"图片映射 Excel 第 {row_index} 行缺少款色。")
        if not image_filename:
            raise ValueError(f"图片映射 Excel 第 {row_index} 行缺少图片文件名。")
        formula_match = IMAGE_FORMULA_PATTERN.search(image_filename)
        if formula_match and embedded_images_by_id is not None:
            image_id = formula_match.group("image_id").strip()
            embedded_image = embedded_images_by_id.get(image_id)
            if not embedded_image:
                raise ValueError(f"图片映射 Excel 第 {row_index} 行的内嵌图片无法读取。")
            clean_image_filename = embedded_image["original_filename"]
            referenced_embedded_images.append(embedded_image)
        else:
            clean_image_filename = Path(image_filename.replace("\\", "/")).name.strip()
        rows.append(
            {
                "style_color": style_color,
                "image_filename": clean_image_filename,
            }
        )
    if not rows:
        raise ValueError("图片映射 Excel 里没有可导入的内容。")
    return rows, referenced_embedded_images


def parse_image_mapping_workbook(file_obj) -> list[dict]:
    workbook = load_workbook(file_obj, data_only=True)
    rows, _ = _parse_image_mapping_rows(workbook[workbook.sheetnames[0]])
    return rows


def _embedded_image_relationships(zip_file: ZipFile) -> dict[str, dict]:
    try:
        cellimages_xml = zip_file.read("xl/cellimages.xml")
        relationships_xml = zip_file.read("xl/_rels/cellimages.xml.rels")
    except KeyError:
        return {}

    relationships = {}
    relationships_root = ElementTree.fromstring(relationships_xml)
    for relationship in relationships_root:
        relationship_id = str(relationship.attrib.get("Id") or "").strip()
        target = str(relationship.attrib.get("Target") or "").strip()
        if not relationship_id or not target:
            continue
        package_path = posixpath.normpath(posixpath.join("xl", target.lstrip("/")))
        if not package_path.startswith("xl/media/"):
            continue
        relationships[relationship_id] = package_path

    images = {}
    cellimages_root = ElementTree.fromstring(cellimages_xml)
    for cell_image in cellimages_root.iter():
        if cell_image.tag.rsplit("}", 1)[-1] != "cellImage":
            continue
        image_name = ""
        relationship_id = ""
        for descendant in cell_image.iter():
            local_name = descendant.tag.rsplit("}", 1)[-1]
            if local_name == "cNvPr" and not image_name:
                image_name = str(descendant.attrib.get("name") or "").strip()
            elif local_name == "blip" and not relationship_id:
                relationship_id = str(descendant.attrib.get(RELATIONSHIP_EMBED_ATTRIBUTE) or "").strip()
        package_path = relationships.get(relationship_id)
        if not image_name or not package_path:
            continue
        extension = Path(package_path).suffix.lower()
        content_type = EMBEDDED_IMAGE_CONTENT_TYPES.get(extension)
        if not content_type:
            continue
        try:
            content = zip_file.read(package_path)
        except KeyError:
            continue
        if not content:
            continue
        images[image_name] = {
            "original_filename": Path(package_path).name,
            "extension": extension,
            "content": content,
            "content_type": content_type,
        }
    return images


def parse_image_mapping_workbook_with_embedded_images(file_obj) -> tuple[list[dict], list[dict]]:
    """Parse a mapping workbook and extract WPS/Excel DISPIMG cell images."""
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    workbook_bytes = file_obj.read()
    try:
        with ZipFile(BytesIO(workbook_bytes)) as zip_file:
            embedded_images_by_id = _embedded_image_relationships(zip_file)
    except (BadZipFile, ElementTree.ParseError) as error:
        raise ValueError("图片映射 Excel 不是有效的 xlsx 文件。") from error

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    return _parse_image_mapping_rows(workbook[workbook.sheetnames[0]], embedded_images_by_id)


def export_excel_image(image_bytes: bytes) -> ExcelImage:
    """Normalize a source image to a compact PNG or JPEG for an Excel cell."""
    with PillowImage.open(BytesIO(image_bytes)) as source:
        source.load()
        if source.width * source.height > EXPORT_IMAGE_MAX_PIXELS:
            raise ValueError("图片像素过大，无法嵌入导出文件。")
        image = ImageOps.exif_transpose(source).copy()
        has_alpha = image.mode in {"RGBA", "LA"} or (
            image.mode == "P" and "transparency" in source.info
        )

    image.thumbnail((EXPORT_IMAGE_MAX_SIDE, EXPORT_IMAGE_MAX_SIDE), PillowImage.Resampling.LANCZOS)
    output = BytesIO()
    if has_alpha:
        image.convert("RGBA").save(output, format="PNG", optimize=True)
    else:
        image.convert("RGB").save(output, format="JPEG", quality=85, optimize=True)
    output.seek(0)
    excel_image = ExcelImage(output)
    excel_image.width = image.width
    excel_image.height = image.height
    return excel_image


def workbook_bytes(
    products: list[dict],
    visible_fields,
    image_fetcher: Callable[[str], bytes] | None = None,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "商品资料"
    include_completion_flag = any(field.key == "completion_flag" for field in visible_fields)
    export_fields = [
        field for field in visible_fields
        if field.key != "completion_flag" and field.key not in CATALOG_EXPORT_HIDDEN_FIELD_KEYS
    ]
    headers = ["历时天数"]
    if include_completion_flag:
        headers.append("资料完成")
    headers.extend(field.excel_header for field in export_fields)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="5B4B3A")
    start_index = 2
    if include_completion_flag:
        start_index = 3
    image_column_index = next(
        (
            index
            for index, field in enumerate(export_fields, start=start_index)
            if field.key == "image_url"
        ),
        None,
    )
    for product in products:
        row = [product.get("elapsed_days_label", "")]
        if include_completion_flag:
            row.append(product.get("completion_flag", ""))
        row.extend(product.get(field.key) for field in export_fields)
        if image_fetcher and image_column_index:
            row[image_column_index - 1] = ""
        worksheet.append(row)
        if not image_fetcher or not image_column_index:
            continue
        image_url = str(product.get("image_url") or "").strip()
        if not image_url:
            continue
        try:
            excel_image = export_excel_image(image_fetcher(image_url))
        except Exception:
            # A failed external image must not prevent the rest of the export.
            continue
        row_index = worksheet.max_row
        image_cell = worksheet.cell(row_index, image_column_index)
        worksheet.add_image(excel_image, image_cell.coordinate)
        worksheet.row_dimensions[row_index].height = 96
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 12
    if include_completion_flag:
        worksheet.column_dimensions["B"].width = 12
    for index, field in enumerate(export_fields, start=start_index):
        minimum_width = 20 if field.key == "image_url" and image_fetcher else 12
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = max(minimum_width, len(field.label) + 4)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def normalize_excel_date(value) -> str:
    return normalize_optional_date_text(value) or ""


def normalize_payment_status(value) -> str:
    raw_value = str(value or "").strip().lower()
    if raw_value in {"已支付", "paid", "y", "yes", "1"}:
        return "paid"
    return "unpaid"


def parse_supplier_settlement_workbook(file_obj) -> list[dict]:
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    raw_headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    normalized_headers = [normalize_header(header) for header in raw_headers]
    required_headers = [
        normalize_header("供应商编码"),
        normalize_header("应付金额"),
    ]
    missing_required_headers = [header for header in required_headers if header not in normalized_headers]
    if missing_required_headers:
        raise ValueError("供应商结算模板表头不匹配，请先下载系统模板后再导入。")

    rows = []
    seen_supplier_codes = set()
    for row_index in range(2, worksheet.max_row + 1):
        payload = {}
        has_content = False
        for column_index, normalized in enumerate(normalized_headers, start=1):
            field_key = SUPPLIER_SETTLEMENT_HEADER_KEY_LOOKUP.get(normalized)
            if not field_key:
                continue
            value = worksheet.cell(row_index, column_index).value
            if value not in (None, ""):
                has_content = True
            payload[field_key] = value
        if not has_content:
            continue
        supplier_code = str(payload.get("supplier_code") or "").strip()
        if not supplier_code:
            raise ValueError(f"第 {row_index} 行缺少供应商编码。")
        if supplier_code in seen_supplier_codes:
            raise ValueError(f"导入文件中供应商编码 {supplier_code} 重复，请先合并后再导入。")
        seen_supplier_codes.add(supplier_code)
        amount_value = payload.get("amount_due")
        if amount_value in (None, ""):
            raise ValueError(f"第 {row_index} 行缺少应付金额。")
        try:
            normalized_amount = float(amount_value)
        except (TypeError, ValueError) as error:
            raise ValueError(f"第 {row_index} 行的应付金额必须是数字。") from error
        rows.append(
            {
                "supplier_code": supplier_code,
                "supplier_name": str(payload.get("supplier_name") or "").strip(),
                "invoice_name": str(payload.get("invoice_name") or "").strip(),
                "amount_due": normalized_amount,
                "payment_status": normalize_payment_status(payload.get("payment_status")),
                "payment_date": normalize_excel_date(payload.get("payment_date")),
                "note": str(payload.get("note") or "").strip(),
            }
        )
    return rows


def supplier_settlement_workbook_bytes(settlements: list[dict]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "供应商结算"
    worksheet.append(SUPPLIER_SETTLEMENT_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="355F52")
    for item in settlements:
        worksheet.append(
            [
                item.get("supplier_code", ""),
                item.get("supplier_name", ""),
                item.get("invoice_name", ""),
                float(item.get("amount_due") or 0),
                "已支付" if str(item.get("payment_status") or "") == "paid" else "待支付",
                item.get("payment_date", ""),
                item.get("note", ""),
            ]
        )
    worksheet.freeze_panes = "A2"
    column_widths = [16, 22, 20, 14, 12, 14, 24]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def supplier_settlement_template_bytes() -> bytes:
    return supplier_settlement_workbook_bytes([])


def parse_supplier_master_workbook(file_obj) -> list[dict]:
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    raw_headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    normalized_headers = [normalize_header(header) for header in raw_headers]
    required_headers = [normalize_header(header) for header in SUPPLIER_MASTER_HEADERS]
    if any(header not in normalized_headers for header in required_headers):
        raise ValueError("供应商主档模板表头不匹配，请先下载系统模板后再导入。")

    rows = []
    for row_index in range(2, worksheet.max_row + 1):
        payload = {}
        has_content = False
        for column_index, normalized_header in enumerate(normalized_headers, start=1):
            field_key = SUPPLIER_MASTER_HEADER_KEY_LOOKUP.get(normalized_header)
            if not field_key:
                continue
            value = worksheet.cell(row_index, column_index).value
            if value not in (None, ""):
                has_content = True
            payload[field_key] = str("" if value is None else value).strip()
        if not has_content:
            continue
        if not all(payload.get(key) for key in ("supplier_code", "supplier_name", "supply_chain_manager")):
            raise ValueError(f"第 {row_index} 行的供应商编号、供应商名称和供应链经理不能为空。")
        rows.append(payload)
    return rows


def supplier_master_template_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "供应商主档模板"
    worksheet.append(SUPPLIER_MASTER_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="355F52")
    worksheet.freeze_panes = "A2"
    for index, width in enumerate((16, 24, 18), start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_supplier_bill_workbook(file_obj) -> list[dict]:
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    raw_headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    normalized_headers = [normalize_header(header) for header in raw_headers]
    required_headers = [normalize_header(header) for header in SUPPLIER_BILL_HEADERS]
    missing_headers = [header for header in required_headers if header not in normalized_headers]
    if missing_headers:
        raise ValueError("账单模板表头不匹配，请先下载系统模板后再导入。")

    rows = []
    for row_index in range(2, worksheet.max_row + 1):
        payload = {}
        has_content = False
        for column_index, normalized_header in enumerate(normalized_headers, start=1):
            field_key = SUPPLIER_BILL_HEADER_KEY_LOOKUP.get(normalized_header)
            if not field_key:
                continue
            value = worksheet.cell(row_index, column_index).value
            if value not in (None, ""):
                has_content = True
            payload[field_key] = value
        if not has_content:
            continue
        supplier_code = str(payload.get("supplier_code") or "").strip()
        supplier_name = str(payload.get("supplier_name") or "").strip()
        if not supplier_code or not supplier_name:
            raise ValueError(f"第 {row_index} 行缺少供应商编号或供应商名称。")
        quantity_value = payload.get("quantity")
        try:
            quantity_number = float(quantity_value)
        except (TypeError, ValueError) as error:
            raise ValueError(f"第 {row_index} 行的数量必须是整数。") from error
        if not math.isfinite(quantity_number) or not quantity_number.is_integer():
            raise ValueError(f"第 {row_index} 行的数量必须是整数，退货可填写负整数。")
        try:
            tax_included_price = float(payload.get("tax_included_price"))
            settlement_amount = float(payload.get("settlement_amount"))
        except (TypeError, ValueError) as error:
            raise ValueError(f"第 {row_index} 行的含税价和结算金额必须是数字。") from error
        if not math.isfinite(tax_included_price) or tax_included_price < 0:
            raise ValueError(f"第 {row_index} 行的含税价必须是非负数字。")
        if not math.isfinite(settlement_amount):
            raise ValueError(f"第 {row_index} 行的结算金额必须是数字，退货可填写负数。")
        rows.append(
            {
                "source_row_no": row_index,
                "supplier_code": supplier_code,
                "supplier_name": supplier_name,
                "mode": str(payload.get("mode") or "").strip(),
                "supply_chain_manager": str(payload.get("supply_chain_manager") or "").strip(),
                "supplier_style_code": str(payload.get("supplier_style_code") or "").strip(),
                "brand_name": str(payload.get("brand_name") or "").strip(),
                "style_color": str(payload.get("style_color") or "").strip(),
                "quantity": int(quantity_number),
                "tax_included_price": tax_included_price,
                "settlement_amount": settlement_amount,
            }
        )
    return rows


def supplier_bill_workbook_bytes(lines: list[dict]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "供应商账单明细"
    worksheet.append(SUPPLIER_BILL_EXPORT_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="355F52")
    for line in lines:
        worksheet.append(
            [
                line.get("period_month", ""),
                line.get("supplier_code", ""),
                line.get("supplier_name", ""),
                line.get("mode", ""),
                line.get("supply_chain_manager", ""),
                line.get("supplier_style_code", ""),
                line.get("brand_name", ""),
                line.get("style_color", ""),
                int(line.get("quantity") or 0),
                float(line.get("tax_included_price") or 0),
                float(line.get("settlement_amount") or 0),
            ]
        )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    column_widths = [14, 16, 22, 14, 16, 18, 16, 18, 12, 14, 16]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def supplier_bill_template_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "供应商账单模板"
    worksheet.append(SUPPLIER_BILL_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="355F52")
    worksheet.freeze_panes = "A2"
    column_widths = [16, 22, 14, 16, 18, 16, 18, 12, 14, 16]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_brand_bill_workbook(file_obj) -> dict:
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    month_label = str(worksheet.cell(3, 1).value or worksheet.cell(2, 1).value or "").strip()
    summary = {
        "month_label": month_label,
        "channel_count": 0,
        "shop_count": 0,
        "total_qty": 0.0,
        "total_amount": 0.0,
        "gz_qty": 0.0,
        "gz_amount": 0.0,
        "wh_qty": 0.0,
        "wh_amount": 0.0,
        "channels": [],
        "table_rows": [],
    }

    channels: dict[str, dict] = {}
    for row_index in range(3, worksheet.max_row + 1):
        month_value = worksheet.cell(row_index, 1).value
        platform_value = worksheet.cell(row_index, 2).value
        shop_value = worksheet.cell(row_index, 3).value
        total_qty_value = worksheet.cell(row_index, 4).value
        total_amount_value = worksheet.cell(row_index, 5).value
        gz_qty_value = worksheet.cell(row_index, 6).value
        gz_amount_value = worksheet.cell(row_index, 7).value
        wh_qty_value = worksheet.cell(row_index, 8).value
        wh_amount_value = worksheet.cell(row_index, 9).value

        platform_raw = str(platform_value or "").strip()
        shop_name = str(shop_value or "").strip()
        month_row_label = str(month_value or "").strip()
        if not any(
            value not in (None, "")
            for value in (
                month_value,
                platform_value,
                shop_value,
                total_qty_value,
                total_amount_value,
                gz_qty_value,
                gz_amount_value,
                wh_qty_value,
                wh_amount_value,
            )
        ):
            continue
        summary["table_rows"].append(
            {
                "month_label": month_row_label,
                "platform_name": platform_raw,
                "shop_name": shop_name,
                "total_qty": total_qty_value,
                "total_amount": total_amount_value,
                "gz_qty": gz_qty_value,
                "gz_amount": gz_amount_value,
                "wh_qty": wh_qty_value,
                "wh_amount": wh_amount_value,
            }
        )
        channel_name = BRAND_BILL_CHANNEL_ALIASES.get(platform_raw, platform_raw)
        if not channel_name:
            continue
        row_qty_total = _numeric_value(total_qty_value)
        row_amount_total = _numeric_value(total_amount_value)
        row_qty_gz = _numeric_value(gz_qty_value)
        row_amount_gz = _numeric_value(gz_amount_value)
        row_qty_wh = _numeric_value(wh_qty_value)
        row_amount_wh = _numeric_value(wh_amount_value)

        bucket = channels.setdefault(
            channel_name,
            {
                "channel_name": channel_name,
                "shop_count": 0,
                "shops": [],
                "total_qty": 0.0,
                "total_amount": 0.0,
                "gz_qty": 0.0,
                "gz_amount": 0.0,
                "wh_qty": 0.0,
                "wh_amount": 0.0,
            },
        )

        normalized_shop = shop_name if shop_name and "合计" not in shop_name else ""
        if normalized_shop:
            bucket["shop_count"] += 1
            bucket["shops"].append(normalized_shop)
        bucket["total_qty"] += row_qty_total
        bucket["total_amount"] += row_amount_total
        bucket["gz_qty"] += row_qty_gz
        bucket["gz_amount"] += row_amount_gz
        bucket["wh_qty"] += row_qty_wh
        bucket["wh_amount"] += row_amount_wh

    if "合计" in channels:
        total_row = channels["合计"]
        summary["total_qty"] = total_row["total_qty"]
        summary["total_amount"] = total_row["total_amount"]
        summary["gz_qty"] = total_row["gz_qty"]
        summary["gz_amount"] = total_row["gz_amount"]
        summary["wh_qty"] = total_row["wh_qty"]
        summary["wh_amount"] = total_row["wh_amount"]

    filtered_channels = [value for key, value in channels.items() if key != "合计"]
    if not summary["total_qty"] and not summary["total_amount"]:
        summary["total_qty"] = sum(item["total_qty"] for item in filtered_channels)
        summary["total_amount"] = sum(item["total_amount"] for item in filtered_channels)
        summary["gz_qty"] = sum(item["gz_qty"] for item in filtered_channels)
        summary["gz_amount"] = sum(item["gz_amount"] for item in filtered_channels)
        summary["wh_qty"] = sum(item["wh_qty"] for item in filtered_channels)
        summary["wh_amount"] = sum(item["wh_amount"] for item in filtered_channels)

    summary["channels"] = filtered_channels
    summary["channel_count"] = len(filtered_channels)
    summary["shop_count"] = sum(item["shop_count"] for item in filtered_channels)
    return summary


def dashboard_rows_from_brand_bill_summary(summary: dict | None) -> list[dict]:
    if not summary:
        return []
    rows = []
    for row in summary.get("table_rows") or []:
        rows.append(
            {
                "month_label": str(row.get("month_label") or "").strip(),
                "platform_name": str(row.get("platform_name") or "").strip(),
                "shop_name": str(row.get("shop_name") or "").strip(),
                "total_qty": _numeric_value(row.get("total_qty")),
                "total_amount": _numeric_value(row.get("total_amount")),
                "gz_qty": _numeric_value(row.get("gz_qty")),
                "gz_amount": _numeric_value(row.get("gz_amount")),
                "wh_qty": _numeric_value(row.get("wh_qty")),
                "wh_amount": _numeric_value(row.get("wh_amount")),
            }
        )
    return rows


def brand_bill_dashboard_workbook_bytes(rows: list[dict]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "马天奴月销看板"
    worksheet.merge_cells("D1:E1")
    worksheet.merge_cells("F1:I1")
    worksheet.merge_cells("J1:M1")
    worksheet["D1"] = "合计"
    worksheet["F1"] = "广州仓"
    worksheet["J1"] = "武汉仓"
    for index, header in enumerate(BRAND_BILL_HEADERS, start=1):
        worksheet.cell(2, index).value = header
    for cell in worksheet[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="7F3B08")
    for row in rows:
        total_qty = _numeric_value(row.get("total_qty"))
        total_amount = _numeric_value(row.get("total_amount"))
        gz_qty = _numeric_value(row.get("gz_qty"))
        gz_amount = _numeric_value(row.get("gz_amount"))
        wh_qty = _numeric_value(row.get("wh_qty"))
        wh_amount = _numeric_value(row.get("wh_amount"))
        worksheet.append(
            [
                str(row.get("month_label") or ""),
                str(row.get("platform_name") or ""),
                str(row.get("shop_name") or ""),
                total_qty,
                total_amount,
                gz_qty,
                gz_amount,
                _ratio_value(gz_qty, total_qty),
                _ratio_value(gz_amount, total_amount),
                wh_qty,
                wh_amount,
                _ratio_value(wh_qty, total_qty),
                _ratio_value(wh_amount, total_amount),
            ]
        )
    for row in worksheet.iter_rows(min_row=3, max_col=13):
        for column_index in (4, 6, 10):
            row[column_index - 1].number_format = "#,##0"
        for column_index in (5, 7, 11):
            row[column_index - 1].number_format = "#,##0.00"
        for column_index in (8, 9, 12, 13):
            row[column_index - 1].number_format = "0.0%"
    column_widths = [16, 14, 28, 12, 14, 12, 14, 12, 12, 12, 14, 12, 12]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[worksheet.cell(2, index).column_letter].width = width
    worksheet.freeze_panes = "A3"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def brand_bill_template_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "品牌月账单"
    worksheet.merge_cells("D1:E1")
    worksheet.merge_cells("F1:I1")
    worksheet.merge_cells("J1:M1")
    worksheet["D1"] = "合计"
    worksheet["F1"] = "广州仓"
    worksheet["J1"] = "武汉仓"
    for index, header in enumerate(BRAND_BILL_HEADERS, start=1):
        worksheet.cell(2, index).value = header
    for cell in worksheet[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="7F3B08")
    column_widths = [16, 14, 28, 12, 14, 12, 14, 12, 12, 12, 14, 12, 12]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[worksheet.cell(2, index).column_letter].width = width
    worksheet.freeze_panes = "A3"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _numeric_value(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _ratio_value(value, total) -> float | None:
    normalized_total = _numeric_value(total)
    if not normalized_total:
        return None
    return _numeric_value(value) / normalized_total
