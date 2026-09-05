from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.workbook.defined_name import DefinedName


INITIAL_REVIEW_HEADERS = (
    "定价记录ID",
    "企划记录号",
    "来源商品ID",
    "来源版本",
    "年份季节",
    "款号",
    "款色",
    "商品名称",
    "供应商",
    "含税成本",
    "固定倍率",
    "供应商系数",
    "原始测算价",
    "测算上新价",
    "初审品类",
    "初审上新价",
    "渠道划分",
    "当前状态",
)
EDITABLE_HEADERS = {"初审品类", "初审上新价", "渠道划分"}
MAX_IMPORT_ROWS = 5000


def _add_defined_name(workbook: Workbook, name: str, reference: str) -> None:
    defined_name = DefinedName(name, attr_text=reference)
    if hasattr(workbook.defined_names, "add"):
        workbook.defined_names.add(defined_name)
    else:  # openpyxl 3.0 compatibility for older on-premise Python images.
        workbook.defined_names.append(defined_name)


def initial_review_workbook_bytes(
    rows: list[dict],
    category_options: list[str],
    channel_options: list[str],
    editable_statuses: set[str] | frozenset[str] | tuple[str, ...] = ("suggested", "conflict"),
    worksheet_title: str = "待初审资料",
    allow_manual_edit: bool = False,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = str(worksheet_title or "待初审资料")[:31]
    worksheet.sheet_properties.tabColor = "315447"

    # The confirmed-stage workbook is an inspection copy. It must be a normal
    # editable workbook for manual matching in WPS, while remaining excluded
    # from the import workflow by the server-side status check.
    manual_editable = bool(
        allow_manual_edit
        and rows
        and all(str(row.get("status") or "") == "confirmed" for row in rows)
    )

    header_fill = PatternFill("solid", fgColor="315447")
    readonly_fill = PatternFill("solid", fgColor="EEF1EE")
    editable_fill = PatternFill("solid", fgColor="FFF1C7")
    border = Border(bottom=Side(style="thin", color="D4DAD5"))
    for column_index, header in enumerate(INITIAL_REVIEW_HEADERS, start=1):
        cell = worksheet.cell(1, column_index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if header in EDITABLE_HEADERS and not manual_editable:
            cell.comment = Comment(
                "此列允许初审人员修改。请直接编辑黄色单元格；其余列为系统同步资料，请勿修改。",
                "商品企划中心",
            )

    status_labels = {
        "waiting": "待计算",
        "suggested": "待初审",
        "review_pending": "待复核",
        "confirmed": "复核通过，待回传",
        "published": "已回传",
        "conflict": "版本冲突",
    }
    editable_statuses = {str(status) for status in editable_statuses}
    editable_rows: list[int] = []
    for row_index, row in enumerate(rows, start=2):
        row_editable = str(row.get("status") or "") in editable_statuses
        if row_editable:
            editable_rows.append(row_index)
        values = (
            int(row["id"]),
            str(row.get("publication_id") or ""),
            int(row["source_product_id"]),
            int(row.get("source_version_no") or 1),
            str(row.get("season_year") or ""),
            str(row.get("style_code") or ""),
            str(row.get("style_color") or row.get("color_name") or ""),
            str(row.get("product_name") or ""),
            str(row.get("supplier") or ""),
            row.get("cost"),
            row.get("fixed_multiplier"),
            row.get("supplier_coefficient"),
            row.get("raw_price"),
            int(row.get("calculated_price") or 0),
            str(row.get("category") or ""),
            int(row.get("launch_price") or row.get("calculated_price") or 0),
            str(row.get("channel") or ""),
            status_labels.get(str(row.get("status") or ""), str(row.get("status") or "")),
        )
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row_index, column_index, value)
            header = INITIAL_REVIEW_HEADERS[column_index - 1]
            if manual_editable:
                cell.fill = PatternFill(fill_type=None)
            else:
                cell.fill = editable_fill if row_editable and header in EDITABLE_HEADERS else readonly_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=header in {"商品名称", "供应商"})
            cell.protection = Protection(
                locked=False
                if manual_editable
                else not row_editable or header not in EDITABLE_HEADERS
            )
        for column_index in (1, 2, 3, 4, 6, 7):
            worksheet.cell(row_index, column_index).number_format = "@"
        for column_index in (14, 16):
            worksheet.cell(row_index, column_index).number_format = "0"
        for column_index in (10, 11, 12, 13):
            worksheet.cell(row_index, column_index).number_format = "0.00"
        worksheet.row_dimensions[row_index].height = 25

    option_sheet = workbook.create_sheet("填写说明")
    option_sheet.sheet_properties.tabColor = "C8883D"
    option_sheet.append(["品类选项", "渠道选项", "填写规则"])
    max_options = max(len(category_options), len(channel_options), 1)
    for index in range(max_options):
        option_sheet.append(
            [
                category_options[index] if index < len(category_options) else "",
                channel_options[index] if index < len(channel_options) else "",
                (
                    (
                        "本工作簿为复核通过阶段的可编辑检查副本，可在 WPS/Excel 中临时调整内容用于人工匹配和验证；"
                        "修改不会回写商品企划中心，也不能通过本工作簿回传藏宝阁。发现错误请回到工作台点击条目中的“修改”，重新走初审与复核。"
                        if manual_editable
                        else (
                            "可修改字段：初审品类、初审上新价、渠道划分。黄色列可编辑；"
                            "为兼容 WPS 和 LibreOffice，初审表不启用整表保护，灰色资料列请勿修改；"
                            "导入只保存初审资料，不会自动提交复核。"
                            if editable_rows
                            else "当前工作簿为系统资料导出，仅用于查看；如需修改，请回到工作台按流程操作。"
                        )
                    )
                    if index == 0
                    else ""
                ),
            ]
        )
    option_sheet.append(
        [
            "",
            "",
            (
                "操作提示：当前导出允许直接编辑，用于人工匹配和验证；修改内容不会被导入系统。"
                if manual_editable
                else (
                    "操作提示：请直接编辑主表中的黄色单元格，不要整行粘贴，也不要修改表头和灰色资料列。"
                    if editable_rows
                    else "操作提示：本工作簿仅用于查看；发现错误请回到工作台使用“修改”按钮。"
                )
            ),
        ]
    )
    option_sheet.append(
        [
            "",
            "",
            (
                "复核通过阶段不接受 Excel 导入，只有回到工作台使用“修改”并重新完成初审与复核后才能回传藏宝阁。"
                if manual_editable
                else (
                    "初审上新价必须填写大于 0 的整数；品类和渠道请使用下拉选项。"
                    if editable_rows
                    else "请回到工作台按流程修改资料。"
                )
            ),
        ]
    )
    for cell in option_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    option_sheet.column_dimensions["A"].width = 22
    option_sheet.column_dimensions["B"].width = 22
    option_sheet.column_dimensions["C"].width = 56
    for row in option_sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        row[0].alignment = Alignment(vertical="top", wrap_text=True)
    option_sheet.freeze_panes = "A2"

    category_end = max(2, len(category_options) + 1)
    channel_end = max(2, len(channel_options) + 1)
    _add_defined_name(workbook, "PlanningCategoryOptions", f"'填写说明'!$A$2:$A${category_end}")
    _add_defined_name(workbook, "PlanningChannelOptions", f"'填写说明'!$B$2:$B${channel_end}")
    if rows and editable_rows:
        category_validation = DataValidation(type="list", formula1="=PlanningCategoryOptions", allow_blank=False)
        category_validation.error = "请从规则中已启用的品类选项选择。"
        category_validation.errorTitle = "品类不正确"
        category_validation.showErrorMessage = True
        channel_validation = DataValidation(type="list", formula1="=PlanningChannelOptions", allow_blank=False)
        channel_validation.error = "请从规则中已启用的渠道选项选择。"
        channel_validation.errorTitle = "渠道不正确"
        channel_validation.showErrorMessage = True
        worksheet.add_data_validation(category_validation)
        worksheet.add_data_validation(channel_validation)
        for row_index in editable_rows:
            category_validation.add(f"O{row_index}")
            channel_validation.add(f"Q{row_index}")

        invalid_price_fill = PatternFill("solid", fgColor="F7C8C3")
        worksheet.conditional_formatting.add(
            f"P2:P{len(rows) + 1}",
            FormulaRule(formula=["OR(P2<1,P2<>INT(P2))"], fill=invalid_price_fill),
        )

    widths = (13, 30, 13, 11, 14, 16, 18, 28, 24, 13, 12, 14, 14, 14, 18, 16, 18, 15)
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    worksheet.freeze_panes = "E2"
    worksheet.auto_filter.ref = f"A1:R{max(1, len(rows) + 1)}"
    worksheet.row_dimensions[1].height = 30
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins = PageMargins(left=0.2, right=0.2, top=0.4, bottom=0.4, header=0.15, footer=0.15)
    worksheet.print_title_rows = "1:1"
    # WPS and LibreOffice do not consistently honor openpyxl's combination of
    # sheet protection and unlocked cells. Initial-review workbooks and the
    # confirmed inspection copy remain unprotected; unrelated read-only
    # exports keep their original sheet protection.
    workbook_is_editable = manual_editable or bool(editable_rows)
    worksheet.protection.sheet = not workbook_is_editable
    worksheet.protection.autoFilter = True
    worksheet.protection.sort = True
    worksheet.protection.selectLockedCells = not workbook_is_editable
    worksheet.protection.selectUnlockedCells = workbook_is_editable
    selection = worksheet.sheet_view.selection[0]
    first_editable_cell = f"O{editable_rows[0]}" if editable_rows else "A1"
    selection.activeCell = first_editable_cell
    selection.sqref = first_editable_cell

    workbook.properties.title = f"商品企划中心{worksheet.title}"
    workbook.properties.subject = "上新审核工作台定价资料"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _required_integer(value, label: str, row_number: int) -> int:
    if isinstance(value, bool):
        raise ValueError(f"第 {row_number} 行“{label}”必须是整数。")
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"第 {row_number} 行“{label}”必须是整数。")
    if not number.is_integer() or number <= 0:
        raise ValueError(f"第 {row_number} 行“{label}”必须是大于 0 的整数。")
    return int(number)


def parse_initial_review_workbook(file_obj) -> list[dict]:
    try:
        workbook = load_workbook(file_obj, data_only=True, read_only=False)
    except Exception as error:
        raise ValueError(f"无法读取 Excel，请确认文件为有效的 .xlsx 工作簿：{error}") from error
    worksheet_name = next(
        (name for name in ("待初审资料", "上新审核资料") if name in workbook.sheetnames),
        None,
    )
    if worksheet_name is None:
        raise ValueError("Excel 中缺少“待初审资料”工作表，请使用系统导出的文件。")
    worksheet = workbook[worksheet_name]
    headers = tuple(str(worksheet.cell(1, index).value or "").strip() for index in range(1, len(INITIAL_REVIEW_HEADERS) + 1))
    if headers != INITIAL_REVIEW_HEADERS:
        raise ValueError("Excel 表头已被修改，请使用系统导出的原始表头。")
    if worksheet.max_row - 1 > MAX_IMPORT_ROWS:
        raise ValueError(f"单次最多导入 {MAX_IMPORT_ROWS} 条初审资料。")

    parsed: list[dict] = []
    seen_record_ids: set[int] = set()
    for row_number in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row_number, index).value for index in range(1, len(INITIAL_REVIEW_HEADERS) + 1)]
        if all(value in (None, "") for value in values):
            continue
        record_id = _required_integer(values[0], "定价记录ID", row_number)
        if record_id in seen_record_ids:
            raise ValueError(f"第 {row_number} 行的定价记录ID重复。")
        seen_record_ids.add(record_id)
        parsed.append(
            {
                "row_number": row_number,
                "record_id": record_id,
                "publication_id": str(values[1] or "").strip(),
                "source_product_id": _required_integer(values[2], "来源商品ID", row_number),
                "source_version_no": _required_integer(values[3], "来源版本", row_number),
                "category": str(values[14] or "").strip(),
                "launch_price": values[15],
                "channel": str(values[16] or "").strip(),
            }
        )
    if not parsed:
        raise ValueError("Excel 中没有可导入的初审资料。")
    return parsed
