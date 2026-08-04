from __future__ import annotations

import io
import json
import tempfile
import unittest
from base64 import b64decode
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import patch
from urllib.parse import unquote_plus
from urllib.parse import urlencode
from wsgiref.util import setup_testing_defaults

from openpyxl import Workbook, load_workbook
from PIL import Image as PillowImage

from catalog_backend import CatalogApplication, init_db
from catalog_backend import db


class CatalogAppTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "catalog.db"
        init_db(self.db_path)
        self.upload_dir = Path(self.temp_dir.name) / "uploads"
        self.app = CatalogApplication(self.db_path, self.upload_dir)

    def tearDown(self):
        self.temp_dir.cleanup()

    def request(
        self,
        path="/",
        method="GET",
        body=b"",
        content_type="application/x-www-form-urlencoded",
        cookie="",
        authorization="",
    ):
        environ = {}
        setup_testing_defaults(environ)
        environ["PATH_INFO"] = path
        environ["REQUEST_METHOD"] = method
        environ["CONTENT_LENGTH"] = str(len(body))
        environ["CONTENT_TYPE"] = content_type
        environ["wsgi.input"] = io.BytesIO(body)
        if "?" in path:
            clean_path, query_string = path.split("?", 1)
            environ["PATH_INFO"] = clean_path
            environ["QUERY_STRING"] = query_string
        if cookie:
            environ["HTTP_COOKIE"] = cookie
        if authorization:
            environ["HTTP_AUTHORIZATION"] = authorization
        captured = {}

        def start_response(status, headers):
            captured["status"] = status
            captured["headers"] = headers

        response = b"".join(self.app(environ, start_response))
        captured["body"] = response
        return captured

    def login(self, username: str, password: str):
        payload = urlencode({"username": username, "password": password}).encode("utf-8")
        response = self.request("/login", method="POST", body=payload)
        self.assertTrue(response["status"].startswith("302"))
        headers = dict(response["headers"])
        return headers["Set-Cookie"].split(";", 1)[0]

    def create_c_operator(
        self,
        username: str,
        display_name: str,
        operating_channel: str,
        billing_platform_codes=None,
    ) -> str:
        user_id = db.create_user(
            self.db_path,
            username,
            display_name,
            "C",
            "demo123",
            must_change_password=False,
            operating_channel=operating_channel,
            billing_platform_codes=billing_platform_codes,
        )
        self.assertGreater(user_id, 0)
        return self.login(username, "demo123")

    def configure_platform_bill_platforms(self, rows: list[dict]):
        return db.save_platform_bill_platform_configs(self.db_path, rows)

    def a_complete_fields_payload(self, **overrides):
        payload = {
            "shooting_date": "2026-06-18",
            "inspection_date": "2026-06-20",
            "detection_report": "已归档",
            "shipping_warehouse": "杭州一仓",
            "brand_name": "思安娜",
            "season_year": "2026夏",
            "image_url": "https://example.com/images/test-style.jpg",
            "style_color": "测试款色",
            "style_code": "TEST-001",
            "color_name": "测试颜色",
            "product_name": "测试商品",
            "category": "上衣",
            "has_accessories": "无",
            "supplier": "测试供应商",
            "cooperation_mode": "经销",
            "supply_chain_manager": "王敏",
            "tax_included_price": "129",
            "tag_price": "199",
            "size_range": "F",
            "size_f": "48",
            "material": "梭织",
            "composition_en": "SHELL: 100% POLYESTER",
            "washing_method": "建议冷水轻柔机洗",
            "washing_method_en": "Machine wash cold, gentle cycle",
            "safety_category": "B类",
            "standard_code": "GB/T 2660",
        }
        payload.update(overrides)
        return payload

    def make_product_two_a_complete(self):
        with db.get_connection(self.db_path) as connection:
            db.update_product(
                connection,
                2,
                self.a_complete_fields_payload(
                    brand_name="Studio Pine",
                    season_year="2026秋",
                    image_url="https://example.com/images/sp-8420.jpg",
                    style_color="针织开衫-米白",
                    style_code="SP-8420",
                    color_name="燕麦白",
                    product_name="毛感针织开衫",
                    category="针织衫",
                    has_accessories="有",
                    supplier="嘉兴尚品针织",
                    cooperation_mode="联营",
                    supply_chain_manager="周岚",
                    tax_included_price="188",
                    tag_price="399",
                    size_range="F",
                    size_f="48",
                    material="针织",
                    composition_en="SHELL: 46% ACRYLIC 30% POLYESTER 24% NYLON",
                    washing_method="建议平铺晾干",
                    washing_method_en="Dry flat",
                    safety_category="B类",
                    standard_code="FZ/T 73018",
                    detection_report="已归档",
                    shipping_warehouse="嘉兴二仓",
                    launch_price="269",
                    launch_channel="门店首发",
                ),
            )

    def make_brand_bill_template_bytes(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet["D1"] = "合计"
        worksheet["F1"] = "广州仓"
        worksheet["H1"] = "武汉仓"
        headers = ["年月", "平台", "店铺", "销售数量", "销售金额", "销售数量", "销售金额", "销售数量", "销售金额"]
        for index, header in enumerate(headers, start=1):
            worksheet.cell(2, index).value = header
        rows = [
            ["2026年8月", "天猫", "思安娜天猫官旗", 120, 36000, 70, 21000, 50, 15000],
            ["", "天猫", "思安娜天猫奥莱", 60, 12600, 35, 7350, 25, 5250],
            ["", "唯品会", "思安娜唯品", 80, 18800, 50, 11200, 30, 7600],
            ["", "抖音", "思安娜抖音", 45, 9900, 20, 4400, 25, 5500],
            ["", "小程序", "思安娜小程序", 30, 7200, 18, 4200, 12, 3000],
            ["", "合计", "", 335, 84500, 193, 48150, 142, 36350],
        ]
        for row_index, row_values in enumerate(rows, start=3):
            for col_index, value in enumerate(row_values, start=1):
                worksheet.cell(row_index, col_index).value = value
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def make_supplier_bill_workbook_bytes(self, rows: list[list[object]]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "供应商账单"
        worksheet.append(
            [
                "供应商编号",
                "供应商名称",
                "模式",
                "供应链经理",
                "供应商款号",
                "品牌名称",
                "款色",
                "数量",
                "含税价",
                "结算金额",
            ]
        )
        for row in rows:
            worksheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def make_supplier_master_workbook_bytes(self, rows: list[list[str]]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "供应商主档"
        worksheet.append(["供应商编号", "供应商名称", "供应链经理"])
        for row in rows:
            worksheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def make_image_mapping_workbook_bytes(self, rows: list[tuple[str, str]]):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "图片映射"
        worksheet.append(["款色", "图片文件名"])
        for style_color, image_filename in rows:
            worksheet.append([style_color, image_filename])
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def make_png_bytes(self) -> bytes:
        image = PillowImage.new("RGB", (16, 12), color=(176, 103, 45))
        output = io.BytesIO()
        image.save(output, format="PNG")
        return output.getvalue()

    def test_a_and_b_editors_show_workflow_specific_catalog_summaries(self):
        a_cookie = self.login("a_editor", "demo123")
        a_response = self.request("/products", cookie=a_cookie)
        a_body = a_response["body"].decode("utf-8")
        self.assertIn('href="/products/1/edit"', a_body)
        self.assertIn('href="/products/2/edit"', a_body)

        b_cookie = self.login("b_editor", "demo123")
        b_response = self.request("/products", cookie=b_cookie)
        b_body = b_response["body"].decode("utf-8")
        self.assertIn('class="tools bulk-tools-vertical"', b_body)
        self.assertIn('class="products-filter-submit"', b_body)
        self.assertIn(">筛选</button>", b_body)
        self.assertNotIn(">筛选资料</button>", b_body)
        self.assertIn("padding: 10px 18px 10px 10px;\n      text-align: center;", b_body)
        self.assertIn("grid-template-rows: minmax(0, 6fr) minmax(0, 14fr);", b_body)
        self.assertIn("grid-template-rows: minmax(250px, 17fr) minmax(150px, 13fr);", b_body)
        self.assertNotIn("当前资料协作节奏", b_body)

        for body in (a_body, b_body):
            self.assertIn('<div class="products-editor-dashboard">', body)
            self.assertLess(
                body.index('<section class="products-top-grid">'),
                body.index('<div class="products-main-stack">'),
            )
            self.assertLess(
                body.index('<section class="products-insights-grid products-insights-single">'),
                body.index('<div class="products-main-stack">'),
            )
            summary_start = body.index("<h2>资料概览</h2>")
            summary_block = body[summary_start:body.index("</section>", summary_start)]
            self.assertNotIn("跟单部发起资料", summary_block)
            self.assertNotIn("近 7 天运营概览", summary_block)

        a_summary_start = a_body.index("<h2>资料概览</h2>")
        a_summary_block = a_body[a_summary_start:a_body.index("</section>", a_summary_start)]
        self.assertLess(a_summary_block.index("总资料数"), a_summary_block.index("已完成"))
        self.assertLess(a_summary_block.index("已完成"), a_summary_block.index("近 7 天新增"))
        self.assertLess(a_summary_block.index("近 7 天新增"), a_summary_block.index("待商品部填写"))
        self.assertLess(a_summary_block.index("待商品部填写"), a_summary_block.index("已删除"))

        b_summary_start = b_body.index("<h2>资料概览</h2>")
        b_summary_block = b_body[b_summary_start:b_body.index("</section>", b_summary_start)]
        self.assertNotIn("总资料数", b_summary_block)
        self.assertNotIn("待商品部填写", b_summary_block)
        self.assertLess(b_summary_block.index("已完成"), b_summary_block.index("近7天新增"))
        self.assertLess(b_summary_block.index("近7天新增"), b_summary_block.index("待完成"))
        self.assertLess(b_summary_block.index("待完成"), b_summary_block.index("待接收"))
        self.assertLess(b_summary_block.index("待接收"), b_summary_block.index("近7天退回"))

    def test_b_workflow_stats_tracks_handoffs_and_returns(self):
        initial = db.b_workflow_stats(self.db_path)
        self.assertEqual(initial["completed"], 1)
        self.assertEqual(initial["recent_submitted_to_b"], 2)
        self.assertEqual(initial["pending_completion"], 1)
        self.assertEqual(initial["awaiting_receipt"], 1)
        self.assertEqual(initial["recent_returned_to_a"], 0)

        b_cookie = self.login("b_editor", "demo123")
        return_response = self.request(
            "/products/2/status",
            method="POST",
            body=urlencode({"status": "draft"}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(return_response["status"].startswith("302"))
        updated = db.b_workflow_stats(self.db_path)
        self.assertEqual(updated["pending_completion"], 0)
        self.assertEqual(updated["recent_returned_to_a"], 1)

    def test_department_filter_controls_follow_each_account_workflow(self):
        def filter_form(body: str) -> str:
            start = body.index('<form class="products-filter-form"')
            return body[start:body.index("</form>", start)]

        a_form = filter_form(self.request("/products", cookie=self.login("a_editor", "demo123"))["body"].decode("utf-8"))
        self.assertIn('class="filter-department-context"', a_form)
        self.assertIn(">跟单部</div>", a_form)
        self.assertNotIn('<select name="department">', a_form)
        self.assertIn('value="draft"', a_form)
        self.assertIn('value="pending"', a_form)
        self.assertIn('value="published"', a_form)
        self.assertNotIn('value="received"', a_form)

        b_form = filter_form(self.request("/products", cookie=self.login("b_editor", "demo123"))["body"].decode("utf-8"))
        self.assertIn(">商品部</div>", b_form)
        self.assertIn('value="pending"', b_form)
        self.assertIn('value="published"', b_form)
        self.assertNotIn('value="draft"', b_form)
        self.assertNotIn('value="received"', b_form)

        c_form = filter_form(self.request("/products", cookie=self.login("c_viewer", "demo123"))["body"].decode("utf-8"))
        self.assertIn(">运营部</div>", c_form)
        self.assertIn("全部接收状态", c_form)
        self.assertIn('value="received"', c_form)

    def test_a_summary_counts_product_after_c_receives_it(self):
        c_cookie = self.login("c_viewer", "demo123")
        receive_response = self.request(
            "/products/1/status",
            method="POST",
            body=urlencode({"status": "received"}).encode("utf-8"),
            cookie=c_cookie,
        )
        self.assertTrue(receive_response["status"].startswith("302"))
        self.assertEqual(db.get_product(self.db_path, 1)["status"], "received")

        a_cookie = self.login("a_editor", "demo123")
        a_response = self.request("/products", cookie=a_cookie)
        a_body = a_response["body"].decode("utf-8")
        summary_start = a_body.index("<h2>资料概览</h2>")
        summary_block = a_body[summary_start:a_body.index("</section>", summary_start)]
        self.assertIn("<span>已完成</span><strong>1</strong>", summary_block)

    def test_b_editor_can_upload_dashboard_excel_and_separate_flow_file(self):
        b_cookie = self.login("b_editor", "demo123")
        dashboard_response = self.request(
            "/billing/brand-bills/dashboard",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-08")],
                files=[
                    (
                        "dashboard_file",
                        "dashboard-2026-08.xlsx",
                        self.make_brand_bill_template_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.assertTrue(dashboard_response["status"].startswith("302"))

        upload_response = self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-08"),
                    ("note", "给跟单部初版"),
                ],
                files=[
                    ("brand_bill_file", "flow-2026-08-v1.xlsx", self.make_brand_bill_template_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.assertTrue(upload_response["status"].startswith("302"))

        summary = db.brand_bill_month_summary(self.db_path, "2026-08")
        self.assertTrue(summary["has_dashboard_rows"])
        self.assertEqual(len(summary["versions"]), 1)
        self.assertEqual(summary["latest_version"]["original_filename"], "flow-2026-08-v1.xlsx")

        page_response = self.request("/billing/brand-bills?month=2026-08", cookie=b_cookie)
        body = page_response["body"].decode("utf-8")
        self.assertIn('class="page-back-row brand-bills-page-back"', body)
        self.assertLess(body.index('class="page-back-row brand-bills-page-back"'), body.index("<h1>品牌月账单</h1>"))
        self.assertIn("马天奴月销看板", body)
        self.assertIn("导入excel", body)
        self.assertIn("导出excel", body)
        self.assertIn("查询", body)
        self.assertIn("销量占比", body)
        self.assertIn("销额占比", body)
        self.assertIn("brand-group-divider", body)
        self.assertIn("brand-dashboard-resize-handle", body)
        self.assertIn("brand_dashboard_column_widths_v1", body)
        self.assertIn('class="brand-dashboard-toolbar"', body)
        self.assertIn('class="brand-dashboard-import-form"', body)
        self.assertIn('class="brand-dashboard-query"', body)
        self.assertNotIn("这里用于维护马天奴月销看板数据", body)
        self.assertNotIn("brand-dashboard-note", body)
        self.assertNotIn("手工填写看板", body)
        self.assertNotIn("保存月销看板", body)
        self.assertIn("2026-08当月账单", body)
        self.assertIn('class="brand-bill-status-heading"', body)
        self.assertIn("历史账单查询", body)
        self.assertIn(">上传</button>", body)
        self.assertIn(">删除</button>", body)
        self.assertIn("下载账单", body)
        self.assertIn("flow-2026-08-v1.xlsx", body)
        self.assertIn("天猫", body)

        export_response = self.request("/billing/brand-bills/dashboard.xlsx?month=2026-08", cookie=b_cookie)
        self.assertTrue(export_response["status"].startswith("200"))
        self.assertTrue(export_response["body"].startswith(b"PK"))
        self.assertIn("brand-dashboard-2026-08.xlsx", dict(export_response["headers"]).get("Content-Disposition", ""))
        exported_workbook = load_workbook(io.BytesIO(export_response["body"]), data_only=True)
        exported_sheet = exported_workbook.active
        self.assertEqual(exported_sheet["A3"].value, "2026年8月")
        self.assertEqual(exported_sheet["B3"].value, "天猫")
        self.assertEqual(exported_sheet["D3"].number_format, "#,##0")
        self.assertEqual(exported_sheet["E3"].number_format, "#,##0.00")
        self.assertEqual(exported_sheet["H2"].value, "销量占比")
        self.assertEqual(exported_sheet["I2"].value, "销额占比")
        self.assertEqual(exported_sheet["L2"].value, "销量占比")
        self.assertEqual(exported_sheet["M2"].value, "销额占比")
        self.assertAlmostEqual(exported_sheet["H3"].value, 70 / 120)
        self.assertAlmostEqual(exported_sheet["I3"].value, 21000 / 36000)
        self.assertAlmostEqual(exported_sheet["L3"].value, 50 / 120)
        self.assertAlmostEqual(exported_sheet["M3"].value, 15000 / 36000)
        self.assertEqual(exported_sheet["H3"].number_format, "0.0%")

    def test_brand_dashboard_rejects_manual_submission_without_excel(self):
        response = self.request(
            "/billing/brand-bills/dashboard",
            method="POST",
            body=urlencode(
                {
                    "month_key": "2026-08",
                    "dashboard_mode": "manual",
                    "dashboard_month_label": "2026年8月",
                    "dashboard_platform_name": "天猫",
                }
            ).encode("utf-8"),
            cookie=self.login("b_editor", "demo123"),
        )
        self.assertTrue(response["status"].startswith("302"))
        self.assertIn("请上传看板 Excel", unquote_plus(dict(response["headers"])["Location"]))
        self.assertIsNone(db.get_brand_bill_by_key(self.db_path, "2026-08"))

    def test_successful_login_redirects_to_modules_home(self):
        response = self.request(
            "/login",
            method="POST",
            body=urlencode({"username": "a_editor", "password": "demo123"}).encode("utf-8"),
        )
        self.assertTrue(response["status"].startswith("302"))
        headers = dict(response["headers"])
        self.assertEqual(headers["Location"], "/modules")

    def test_login_page_uses_simple_mytteno_brand_layout(self):
        response = self.request("/login")
        body = response["body"].decode("utf-8")
        self.assertIn('class="login-page"', body)
        self.assertIn('class="login-mytteno"', body)
        self.assertIn(">MYTENO</div>", body)
        self.assertIn('class="login-brand-kicker"', body)
        self.assertIn(">Sianna</div>", body)
        self.assertIn('class="login-brand-main"', body)
        self.assertIn("思安娜的藏宝阁", body)
        self.assertNotIn("请输入账号与密码", body)
        self.assertIn("供应链 · 商品 · 运营", body)
        self.assertIn("多部门协同管理资料，流程加速，效率加倍", body)
        self.assertIn("开启寻宝之旅", body)
        self.assertNotIn("协同管理资料，一起寻宝", body)
        self.assertIn("grid-row: 1 / span 2;", body)
        self.assertNotIn("默认演示账号", body)
        self.assertIn('name="username"', body)
        self.assertIn('name="password"', body)

    def test_modules_home_shows_first_and_second_modules(self):
        cookie = self.login("a_editor", "demo123")
        response = self.request("/modules", cookie=cookie)
        body = response["body"].decode("utf-8")
        self.assertIn('class="modules-home"', body)
        self.assertIn('<div class="detail-grid">', body)
        self.assertIn('class="modules-home-watermark"', body)
        self.assertIn("玲珑晓楼阁，远山含黛色", body)
        self.assertIn('font-family: "STXingkaiSC-Light"', body)
        self.assertIn('V1 行书版：行楷细体', body)
        self.assertIn("板块一", body)
        self.assertIn("板块二", body)
        self.assertIn("商品资料后台", body)
        self.assertIn("账单与结算", body)
        self.assertIn('href="/billing"', body)

    def test_billing_home_limits_c_to_platform_bill_entry_only(self):
        cookie = self.login("c_viewer", "demo123")
        response = self.request("/billing", cookie=cookie)
        body = response["body"].decode("utf-8")
        self.assertIn("账单与结算", body)
        self.assertIn('href="/billing/platform-bills"', body)
        self.assertIn('class="detail-grid billing-c-home-grid"', body)
        self.assertNotIn("品牌月账单", body)
        self.assertNotIn("供应商结算", body)
        self.assertNotIn("月度状态看板", body)
        self.assertNotIn('href="/billing/brand-bills"', body)
        self.assertNotIn('href="/billing/supplier-settlements"', body)

    def test_billing_home_uses_b_specific_platform_to_brand_workboard(self):
        month_key = "2026-06"
        self.configure_platform_bill_platforms(
            [
                {"code": "tmall", "label": "天猫"},
                {"code": "jd", "label": "京东"},
                {"code": "douyin", "label": "抖音"},
                {"code": "vip", "label": "唯品"},
                {"code": "miniprogram", "label": "小程序"},
            ]
        )
        operator_cookies = {
            "tmall": self.login("c_viewer", "demo123"),
            "jd": self.create_c_operator("jd_submit", "京东运营", "tmall", ["jd"]),
            "douyin": self.create_c_operator("douyin_submit", "抖音运营", "tmall", ["douyin"]),
            "vip": self.create_c_operator("vip_submit", "唯品运营", "vip", ["vip"]),
            "miniprogram": self.create_c_operator("mini_submit", "小程序运营", "tmall", ["miniprogram"]),
        }

        def upload_and_submit(platform_code: str):
            upload_response = self.request(
                "/billing/platform-bills/upload",
                method="POST",
                body=self.build_multi_multipart(
                    fields=[("month_key", month_key), ("platform_code", platform_code)],
                    files=[
                        (
                            "upload_files",
                            f"{platform_code}.xlsx",
                            platform_code.encode("utf-8"),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    ],
                ),
                content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
                cookie=operator_cookies[platform_code],
            )
            self.assertTrue(upload_response["status"].startswith("302"))
            submit_response = self.request(
                "/billing/platform-bills/submit",
                method="POST",
                body=urlencode({"month_key": month_key, "platform_code": platform_code}).encode("utf-8"),
                cookie=operator_cookies[platform_code],
            )
            self.assertTrue(submit_response["status"].startswith("302"))

        for platform_code in ("tmall", "jd", "douyin", "vip"):
            upload_and_submit(platform_code)

        b_cookie = self.login("b_editor", "demo123")
        b_body = self.request("/billing", cookie=b_cookie)["body"].decode("utf-8")
        self.assertIn("账单与结算-月度看板", b_body)
        self.assertIn('class="page-back-row billing-b-home-back"', b_body)
        self.assertLess(b_body.index("billing-b-home-back"), b_body.index("账单与结算-月度看板"))
        self.assertIn('class="panel billing-b-workboard"', b_body)
        self.assertIn('class="detail-grid billing-b-entry-grid"', b_body)
        self.assertIn(".billing-b-workboard .catalog-table th,", b_body)
        self.assertIn(".billing-b-workboard .board-risk-pill", b_body)
        for column_name in ("月份", "平台账单", "品牌月账单", "当前节点"):
            self.assertIn(f"<th>{column_name}</th>", b_body)
        self.assertNotIn("<th>操作</th>", b_body)
        self.assertIn('href="/billing/platform-bills"', b_body)
        self.assertIn('href="/billing/brand-bills"', b_body)
        self.assertIn("进入平台账单", b_body)
        self.assertIn("进入品牌月账单", b_body)
        self.assertIn("账单文件 4 / 5 · 已提交 4 / 5", b_body)
        self.assertIn("平台账单不完整", b_body)
        self.assertNotIn("供应商结算", b_body)
        self.assertNotIn("应付合计", b_body)
        self.assertNotIn("待付合计", b_body)

        upload_and_submit("miniprogram")
        b_body = self.request("/billing", cookie=b_cookie)["body"].decode("utf-8")
        self.assertIn("待整理品牌月账单", b_body)

        return_response = self.request(
            "/billing/platform-bills/return-request",
            method="POST",
            body=urlencode({"month_key": month_key, "platform_code": "tmall", "reason": "账单金额有误"}).encode("utf-8"),
            cookie=operator_cookies["tmall"],
        )
        self.assertTrue(return_response["status"].startswith("302"))
        b_body = self.request("/billing", cookie=b_cookie)["body"].decode("utf-8")
        self.assertIn("天猫申请退回", b_body)

        pending_request = db.list_platform_bill_return_requests(self.db_path, month_key)[0]
        reject_response = self.request(
            "/billing/platform-bills/return-decision",
            method="POST",
            body=urlencode(
                {
                    "request_id": str(pending_request["id"]),
                    "month_key": month_key,
                    "platform_code": "tmall",
                    "decision": "reject",
                }
            ).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(reject_response["status"].startswith("302"))
        b_body = self.request("/billing", cookie=b_cookie)["body"].decode("utf-8")
        self.assertIn("待整理品牌月账单", b_body)

        brand_upload_response = self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", month_key), ("note", "品牌月账单初版")],
                files=[
                    (
                        "brand_bill_file",
                        "brand-2026-06.xlsx",
                        self.make_brand_bill_template_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.assertTrue(brand_upload_response["status"].startswith("302"))
        brand_submit_response = self.request(
            "/billing/brand-bills/submit",
            method="POST",
            body=urlencode({"month_key": month_key}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(brand_submit_response["status"].startswith("302"))
        b_body = self.request("/billing", cookie=b_cookie)["body"].decode("utf-8")
        self.assertIn("品牌月账单已提交", b_body)

    def test_modules_home_hides_billing_summary_cards_for_all_departments(self):
        for username in ("a_editor", "b_editor", "c_viewer", "admin_reviewer"):
            body = self.request("/modules", cookie=self.login(username, "demo123"))["body"].decode("utf-8")
            self.assertIn("账单与结算", body)
            self.assertIn('href="/billing"', body)
            self.assertNotIn("主流程部门", body)
            self.assertNotIn("C 权限", body)
            self.assertNotIn("5 平台", body)

    def test_navigation_watermark_is_available_for_all_account_pages(self):
        for username, path in (
            ("a_editor", "/products"),
            ("b_editor", "/billing"),
            ("c_viewer", "/billing"),
            ("admin_reviewer", "/modules"),
        ):
            body = self.request(path, cookie=self.login(username, "demo123"))["body"].decode("utf-8")
            self.assertIn('<nav class="nav-shell">', body)
            self.assertIn('.nav-shell::before', body)
            self.assertIn('content: "远山含黛色，玲珑晓楼阁"', body)
            self.assertIn('.nav-shell > *', body)
            self.assertIn('document.querySelectorAll("details")', body)
            self.assertIn('menu.addEventListener("pointerleave"', body)
            self.assertIn('menu.open = false;', body)

    def test_billing_home_shows_monthly_workboard(self):
        self.configure_platform_bill_platforms(
            [
                {"code": "tmall", "label": "天猫"},
                {"code": "vip", "label": "唯品"},
            ]
        )
        c_cookie = self.login("c_viewer", "demo123")
        vip_cookie = self.create_c_operator("vip_board", "唯品运营", "vip")
        for platform_code, cookie in (("tmall", c_cookie), ("vip", vip_cookie)):
            response = self.request(
                "/billing/platform-bills/upload",
                method="POST",
                body=self.build_multi_multipart(
                    fields=[
                        ("month_key", "2026-05"),
                        ("platform_code", platform_code),
                    ],
                    files=[
                        ("upload_files", f"{platform_code}.xlsx", b"board-main", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ],
                ),
                content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
                cookie=cookie,
            )
            self.assertTrue(response["status"].startswith("302"))
            submit_response = self.request(
                "/billing/platform-bills/submit",
                method="POST",
                body=urlencode({"month_key": "2026-05", "platform_code": platform_code}).encode("utf-8"),
                cookie=cookie,
            )
            self.assertTrue(submit_response["status"].startswith("302"))

        b_cookie = self.login("b_editor", "demo123")
        self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-05"),
                    ("note", "看板测试品牌账单"),
                ],
                files=[
                    ("brand_bill_file", "brand-board.xlsx", b"brand-board", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.request(
            "/billing/brand-bills/submit",
            method="POST",
            body=urlencode({"month_key": "2026-05"}).encode("utf-8"),
            cookie=b_cookie,
        )

        a_cookie = self.login("a_editor", "demo123")
        self.request(
            "/billing/supplier-settlements/suppliers",
            method="POST",
            body=urlencode(
                {
                    "supplier_code": "050",
                    "supplier_name": "看板供应商050",
                    "invoice_names": "山河",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        supplier = db.get_supplier_by_code(self.db_path, "050")
        self.request(
            "/billing/supplier-settlements/records",
            method="POST",
            body=urlencode(
                {
                    "month_key": "2026-05",
                    "supplier_id": supplier["id"],
                    "invoice_name": "山河",
                    "amount_due": "3000",
                    "payment_status": "paid",
                    "payment_date": "2026-05-30",
                    "note": "看板测试",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )

        home_response = self.request("/billing", cookie=a_cookie)
        body = home_response["body"].decode("utf-8")
        self.assertIn("账单与结算-月度看板", body)
        self.assertNotIn("月度状态看板", body)
        self.assertIn('class="detail-grid billing-a-entry-grid"', body)
        self.assertEqual(body.count('<article class="panel">'), 2)
        self.assertIn('class="a-monthly-board-form" method="post" action="/billing/monthly-board"', body)
        self.assertIn("应付供应商", body)
        self.assertIn("应付总金额", body)
        self.assertIn("12月", body)
        self.assertIn("总计", body)

    def test_a_editor_can_fill_and_save_supplier_monthly_board(self):
        a_cookie = self.login("a_editor", "demo123")
        payload = {
            "year": "2026",
            "supplier_count_12": "3",
            "payable_amount_12": "4500.50",
            "supplier_count_11": "2",
            "payable_amount_11": "1200",
        }
        save_response = self.request(
            "/billing/monthly-board",
            method="POST",
            body=urlencode(payload).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(save_response["status"].startswith("302"))
        board = db.supplier_monthly_board_for_year(self.db_path, 2026)
        self.assertEqual(board["months"][12]["payable_supplier_count"], 3)
        self.assertEqual(board["months"][11]["payable_supplier_count"], 2)
        self.assertEqual(board["payable_amount_total"], 5700.5)
        self.assertEqual(board["supplier_count_total"], 5)

        body = self.request("/billing?year=2026", cookie=a_cookie)["body"].decode("utf-8")
        self.assertIn('name="supplier_count_12"', body)
        self.assertIn('value="4500.50"', body)
        self.assertIn(">5</td>", body)
        self.assertIn(">5700.50</td>", body)

        b_cookie = self.login("b_editor", "demo123")
        forbidden = self.request(
            "/billing/monthly-board",
            method="POST",
            body=urlencode({"year": "2026"}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(forbidden["status"].startswith("403"))

    def test_a_monthly_board_is_available_when_platform_month_is_incomplete(self):
        c_cookie = self.login("c_viewer", "demo123")
        self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-06"),
                    ("platform_code", "tmall"),
                ],
                files=[
                    ("upload_files", "tmall-only.xlsx", b"tmall-only", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        a_cookie = self.login("a_editor", "demo123")
        response = self.request("/billing", cookie=a_cookie)
        body = response["body"].decode("utf-8")
        self.assertIn("账单与结算-月度看板", body)
        self.assertIn('name="supplier_count_6"', body)
        self.assertIn('name="payable_amount_6"', body)
        self.assertNotIn("待运营部补齐", body)

    def test_a_monthly_board_is_independent_from_unpaid_supplier_settlements(self):
        self.configure_platform_bill_platforms(
            [
                {"code": "tmall", "label": "天猫"},
                {"code": "vip", "label": "唯品"},
            ]
        )
        c_cookie = self.login("c_viewer", "demo123")
        vip_cookie = self.create_c_operator("vip_unpaid", "唯品运营", "vip")
        for platform_code, cookie in (("tmall", c_cookie), ("vip", vip_cookie)):
            self.request(
                "/billing/platform-bills/upload",
                method="POST",
                body=self.build_multi_multipart(
                    fields=[
                        ("month_key", "2026-09"),
                        ("platform_code", platform_code),
                    ],
                    files=[
                        ("upload_files", f"{platform_code}.xlsx", b"month-09", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ],
                ),
                content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
                cookie=cookie,
            )
            submit_response = self.request(
                "/billing/platform-bills/submit",
                method="POST",
                body=urlencode({"month_key": "2026-09", "platform_code": platform_code}).encode("utf-8"),
                cookie=cookie,
            )
            self.assertTrue(submit_response["status"].startswith("302"))

        b_cookie = self.login("b_editor", "demo123")
        self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-09"),
                    ("note", "未支付测试账单"),
                ],
                files=[
                    ("brand_bill_file", "brand-2026-09.xlsx", b"brand-09", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.request(
            "/billing/brand-bills/submit",
            method="POST",
            body=urlencode({"month_key": "2026-09"}).encode("utf-8"),
            cookie=b_cookie,
        )

        a_cookie = self.login("a_editor", "demo123")
        self.request(
            "/billing/supplier-settlements/suppliers",
            method="POST",
            body=urlencode(
                {
                    "supplier_code": "099",
                    "supplier_name": "未支付供应商099",
                    "invoice_names": "未支付抬头",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        supplier = db.get_supplier_by_code(self.db_path, "099")
        self.request(
            "/billing/supplier-settlements/records",
            method="POST",
            body=urlencode(
                {
                    "month_key": "2026-09",
                    "supplier_id": supplier["id"],
                    "invoice_name": "未支付抬头",
                    "amount_due": "6800",
                    "payment_status": "unpaid",
                    "payment_date": "",
                    "note": "待支付测试",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )

        response = self.request("/billing", cookie=a_cookie)
        body = response["body"].decode("utf-8")
        self.assertIn("账单与结算-月度看板", body)
        self.assertIn('name="supplier_count_9"', body)
        self.assertIn('name="payable_amount_9"', body)
        self.assertNotIn("本月仍有待支付金额", body)

    def test_brand_and_supplier_pages_forbid_c_viewer(self):
        cookie = self.login("c_viewer", "demo123")
        brand_response = self.request("/billing/brand-bills", cookie=cookie)
        self.assertTrue(brand_response["status"].startswith("403"))
        supplier_response = self.request("/billing/supplier-settlements", cookie=cookie)
        self.assertTrue(supplier_response["status"].startswith("403"))

    def test_c_viewer_can_upload_platform_bill_files(self):
        cookie = self.login("c_viewer", "demo123")
        multipart_body = self.build_multi_multipart(
            fields=[
                ("month_key", "2026-06"),
                ("platform_code", "tmall"),
            ],
            files=[
                ("upload_files", "tmall-main.xlsx", b"main-file-content", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("upload_files", "tmall-note.pdf", b"attachment-content", "application/pdf"),
            ],
        )
        response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=multipart_body,
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=cookie,
        )
        self.assertTrue(response["status"].startswith("302"))
        summary = db.platform_bill_month_summary(self.db_path, "2026-06")
        tmall = next(item for item in summary["platforms"] if item["platform_code"] == "tmall")
        self.assertIsNotNone(tmall["main_file"])
        self.assertEqual(len(tmall["attachments"]), 1)

    def test_c_viewer_submit_platform_locks_only_that_platform_and_cannot_download(self):
        cookie = self.login("c_viewer", "demo123")
        multipart_body = self.build_multi_multipart(
            fields=[("month_key", "2026-06"), ("platform_code", "tmall")],
            files=[
                (
                    "upload_files",
                    "tmall.xlsx",
                    b"tmall-content",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ],
        )
        response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=multipart_body,
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=cookie,
        )
        self.assertTrue(response["status"].startswith("302"))

        submit_response = self.request(
            "/billing/platform-bills/submit",
            method="POST",
            body=urlencode({"month_key": "2026-06", "platform_code": "tmall"}).encode("utf-8"),
            cookie=cookie,
        )
        self.assertTrue(submit_response["status"].startswith("302"))
        month = db.get_billing_month_by_key(self.db_path, "2026-06")
        self.assertEqual(month["status"], "partial_to_b")

        locked_body = self.build_multi_multipart(
            fields=[
                ("month_key", "2026-06"),
                ("platform_code", "tmall"),
            ],
            files=[
                ("upload_files", "tmall-more.pdf", b"new-attachment", "application/pdf"),
            ],
        )
        locked_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=locked_body,
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=cookie,
        )
        self.assertTrue(locked_response["status"].startswith("400"))

        vip_body = self.build_multi_multipart(
            fields=[
                ("month_key", "2026-06"),
                ("platform_code", "vip"),
            ],
            files=[
                ("upload_files", "vip-more.pdf", b"vip-attachment", "application/pdf"),
            ],
        )
        vip_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=vip_body,
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=cookie,
        )
        self.assertTrue(vip_response["status"].startswith("403"))

        summary = db.platform_bill_month_summary(self.db_path, "2026-06")
        tmall = next(item for item in summary["platforms"] if item["platform_code"] == "tmall")
        vip = next(item for item in summary["platforms"] if item["platform_code"] == "vip")
        self.assertTrue(tmall["submitted"])
        self.assertIsNone(vip["main_file"])
        main_file = tmall["main_file"]
        download_response = self.request(f"/billing/platform-bills/files/{main_file['id']}", cookie=cookie)
        self.assertTrue(download_response["status"].startswith("403"))

    def test_vip_operator_platform_bill_scope_filters_workspace_and_enforces_server_side(self):
        tmall_cookie = self.login("c_viewer", "demo123")
        initial_upload = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-06"), ("platform_code", "tmall")],
                files=[("upload_files", "tmall.xlsx", b"tmall", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=tmall_cookie,
        )
        self.assertTrue(initial_upload["status"].startswith("302"))
        tmall = next(
            item
            for item in db.platform_bill_month_summary(self.db_path, "2026-06")["platforms"]
            if item["platform_code"] == "tmall"
        )
        vip_cookie = self.create_c_operator("vip_bill_operator", "唯品运营李二", "vip")

        direct_page = self.request("/billing/platform-bills?month=2026-06&platform=tmall", cookie=vip_cookie)
        direct_body = direct_page["body"].decode("utf-8")
        self.assertIn('<option value="vip"', direct_body)
        self.assertNotIn('<option value="tmall"', direct_body)
        self.assertNotIn('name="platform_code" value="tmall"', direct_body)

        denied_upload = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-06"), ("platform_code", "tmall")],
                files=[("upload_files", "forbidden.xlsx", b"forbidden", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=vip_cookie,
        )
        self.assertTrue(denied_upload["status"].startswith("403"))
        denied_delete = self.request(
            "/billing/platform-bills/delete",
            method="POST",
            body=urlencode({"month_key": "2026-06", "file_id": str(tmall["main_file"]["id"])}).encode("utf-8"),
            cookie=vip_cookie,
        )
        self.assertTrue(denied_delete["status"].startswith("403"))
        denied_submit = self.request(
            "/billing/platform-bills/submit",
            method="POST",
            body=urlencode({"month_key": "2026-06", "platform_code": "tmall"}).encode("utf-8"),
            cookie=vip_cookie,
        )
        self.assertTrue(denied_submit["status"].startswith("403"))
        denied_return = self.request(
            "/billing/platform-bills/return-request",
            method="POST",
            body=urlencode({"month_key": "2026-06", "platform_code": "tmall", "reason": "测试"}).encode("utf-8"),
            cookie=vip_cookie,
        )
        self.assertTrue(denied_return["status"].startswith("403"))

        allowed_upload = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-06"), ("platform_code", "vip")],
                files=[("upload_files", "vip.xlsx", b"vip", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=vip_cookie,
        )
        self.assertTrue(allowed_upload["status"].startswith("302"))

    def test_b_and_admin_can_select_all_configured_platform_bills(self):
        for username in ("b_editor", "admin_reviewer"):
            body = self.request("/billing/platform-bills?month=2026-06", cookie=self.login(username, "demo123"))["body"].decode("utf-8")
            for platform_code in db.platform_bill_platform_codes():
                self.assertIn(f'<option value="{platform_code}"', body)

    def test_platform_bill_return_request_preserves_history_and_allows_resubmission(self):
        c_cookie = self.login("c_viewer", "demo123")
        upload_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-06"), ("platform_code", "tmall")],
                files=[("upload_files", "tmall-v1.xlsx", b"version-one", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        self.assertTrue(upload_response["status"].startswith("302"))
        submit_response = self.request(
            "/billing/platform-bills/submit",
            method="POST",
            body=urlencode({"month_key": "2026-06", "platform_code": "tmall"}).encode("utf-8"),
            cookie=c_cookie,
        )
        self.assertTrue(submit_response["status"].startswith("302"))

        c_page = self.request("/billing/platform-bills?month=2026-06&platform=tmall", cookie=c_cookie)
        self.assertIn("申请退回", c_page["body"].decode("utf-8"))
        request_response = self.request(
            "/billing/platform-bills/return-request",
            method="POST",
            body=urlencode({"month_key": "2026-06", "platform_code": "tmall", "reason": "发现文件金额有误"}).encode("utf-8"),
            cookie=c_cookie,
        )
        self.assertTrue(request_response["status"].startswith("302"))
        pending_request = db.list_platform_bill_return_requests(self.db_path, "2026-06")[0]
        self.assertEqual(pending_request["status"], "pending")

        b_cookie = self.login("b_editor", "demo123")
        b_page = self.request("/billing/platform-bills?month=2026-06&platform=tmall", cookie=b_cookie)
        b_body = b_page["body"].decode("utf-8")
        self.assertIn("退回运营部", b_body)
        self.assertIn("发现文件金额有误", b_body)
        decision_response = self.request(
            "/billing/platform-bills/return-decision",
            method="POST",
            body=urlencode(
                {
                    "request_id": str(pending_request["id"]),
                    "month_key": "2026-06",
                    "platform_code": "tmall",
                    "decision": "approve",
                }
            ).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(decision_response["status"].startswith("302"))

        returned_summary = db.platform_bill_month_summary(self.db_path, "2026-06")
        returned_tmall = next(item for item in returned_summary["platforms"] if item["platform_code"] == "tmall")
        self.assertIsNone(returned_tmall["main_file"])
        self.assertEqual(returned_tmall["history_files"][0]["original_filename"], "tmall-v1.xlsx")
        self.assertEqual(db.get_billing_month_by_key(self.db_path, "2026-06")["status"], "draft")

        returned_page = self.request("/billing/platform-bills?month=2026-06&platform=tmall", cookie=c_cookie)
        returned_body = returned_page["body"].decode("utf-8")
        self.assertIn("待重新提交", returned_body)
        self.assertIn("历史版本（1）", returned_body)
        self.assertIn("tmall-v1.xlsx", returned_body)
        self.assertNotIn("/billing/platform-bills/files/", returned_body)

        reupload_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-06"), ("platform_code", "tmall")],
                files=[("upload_files", "tmall-v2.xlsx", b"version-two", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        self.assertTrue(reupload_response["status"].startswith("302"))
        resubmitted_summary = db.platform_bill_month_summary(self.db_path, "2026-06")
        resubmitted_tmall = next(item for item in resubmitted_summary["platforms"] if item["platform_code"] == "tmall")
        self.assertEqual(resubmitted_tmall["main_file"]["version_no"], 2)
        self.assertEqual(resubmitted_tmall["main_file"]["original_filename"], "tmall-v2.xlsx")
        self.assertEqual(resubmitted_tmall["history_files"][0]["version_no"], 1)

    def test_c_viewer_submit_all_platforms_marks_month_submitted(self):
        self.configure_platform_bill_platforms(
            [
                {"code": "tmall", "label": "天猫"},
                {"code": "vip", "label": "唯品"},
            ]
        )
        tmall_cookie = self.login("c_viewer", "demo123")
        vip_cookie = self.create_c_operator("vip_submit", "唯品运营", "vip")
        for platform_code, cookie in (("tmall", tmall_cookie), ("vip", vip_cookie)):
            upload_response = self.request(
                "/billing/platform-bills/upload",
                method="POST",
                body=self.build_multi_multipart(
                    fields=[
                        ("month_key", "2026-01"),
                        ("platform_code", platform_code),
                    ],
                    files=[
                        (
                            "upload_files",
                            f"{platform_code}.xlsx",
                            f"{platform_code}-content".encode("utf-8"),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                    ],
                ),
                content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
                cookie=cookie,
            )
            self.assertTrue(upload_response["status"].startswith("302"))
            submit_response = self.request(
                "/billing/platform-bills/submit",
                method="POST",
                body=urlencode({"month_key": "2026-01", "platform_code": platform_code}).encode("utf-8"),
                cookie=cookie,
            )
            self.assertTrue(submit_response["status"].startswith("302"))

        month = db.get_billing_month_by_key(self.db_path, "2026-01")
        self.assertEqual(month["status"], "submitted_to_b")
        summary = db.platform_bill_month_summary(self.db_path, "2026-01")
        self.assertTrue(summary["all_submitted"])

    def test_c_viewer_can_delete_own_platform_bill_file_and_reupload(self):
        c_cookie = self.login("c_viewer", "demo123")
        upload_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-10"),
                    ("platform_code", "tmall"),
                ],
                files=[
                    ("upload_files", "tmall-first.xlsx", b"tmall-first", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("upload_files", "tmall-note.pdf", b"tmall-note", "application/pdf"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        self.assertTrue(upload_response["status"].startswith("302"))
        summary = db.platform_bill_month_summary(self.db_path, "2026-10")
        tmall = next(item for item in summary["platforms"] if item["platform_code"] == "tmall")
        main_file = tmall["main_file"]
        self.assertEqual(len(tmall["attachments"]), 1)
        delete_response = self.request(
            "/billing/platform-bills/delete",
            method="POST",
            body=urlencode({"month_key": "2026-10", "file_id": str(main_file["id"]), "delete_scope": "platform"}).encode("utf-8"),
            cookie=c_cookie,
        )
        self.assertTrue(delete_response["status"].startswith("302"))
        refreshed_summary = db.platform_bill_month_summary(self.db_path, "2026-10")
        refreshed_tmall = next(item for item in refreshed_summary["platforms"] if item["platform_code"] == "tmall")
        refreshed_main = refreshed_tmall["main_file"]
        self.assertIsNone(refreshed_main)
        self.assertEqual(len(refreshed_tmall["attachments"]), 0)

    def test_c_viewer_cannot_see_other_c_file_name_or_override_it(self):
        c_cookie = self.login("c_viewer", "demo123")
        second_c_cookie = self.create_c_operator("c_uploader_2", "运营部李四", "tmall")

        first_upload = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-11"),
                    ("platform_code", "tmall"),
                ],
                files=[
                    ("upload_files", "tmall-li-si.xlsx", b"tmall-lisi", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=second_c_cookie,
        )
        self.assertTrue(first_upload["status"].startswith("302"))

        page_response = self.request("/billing/platform-bills?month=2026-11&platform=tmall", cookie=c_cookie)
        body = page_response["body"].decode("utf-8")
        self.assertIn("已由其他同事上传", body)
        self.assertNotIn("tmall-li-si.xlsx", body)

        override_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-11"),
                    ("platform_code", "tmall"),
                ],
                files=[
                    ("upload_files", "tmall-zhang-san.xlsx", b"tmall-zhangsan", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        self.assertTrue(override_response["status"].startswith("403"))

    def test_platform_bill_page_shows_month_overview_cards(self):
        c_cookie = self.login("c_viewer", "demo123")
        for month_key in ("2026-03", "2026-04"):
            self.request(
                "/billing/platform-bills/upload",
                method="POST",
                body=self.build_multi_multipart(
                    fields=[
                        ("month_key", month_key),
                        ("platform_code", "tmall"),
                    ],
                    files=[
                        ("upload_files", f"{month_key}-tmall.xlsx", b"tmall-month", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ],
                ),
                content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
                cookie=c_cookie,
            )
        page_response = self.request("/billing/platform-bills?month=2026-04&platform=tmall", cookie=c_cookie)
        body = page_response["body"].decode("utf-8")
        self.assertIn("平台账单", body)
        self.assertIn("2026-03", body)
        self.assertIn("2026-04", body)
        self.assertIn("当月工作台", body)
        self.assertIn("当前状态", body)
        self.assertIn("账单文件", body)
        self.assertIn("已上传", body)
        self.assertIn("确认提交", body)
        self.assertNotIn("已上传平台", body)
        self.assertNotIn("已提交平台", body)
        self.assertIn("billing-status-chip", body)
        self.assertIn("当前月份已确认提交平台：0 / 1", body)
        self.assertLess(
            body.index('<div class="detail-grid billing-month-grid"'),
            body.index('<form method="get" action="/billing/platform-bills" class="billing-month-picker">'),
        )
        self.assertNotIn("待启动", body)
        self.assertNotIn("待处理平台", body)
        self.assertNotIn("主文件状态", body)
        self.assertNotIn("当前处理", body)
        self.assertNotIn("附件", body)

        b_cookie = self.login("b_editor", "demo123")
        b_body = self.request("/billing/platform-bills?month=2026-04", cookie=b_cookie)["body"].decode("utf-8")
        self.assertIn("billing-month-card", b_body)
        self.assertNotIn("已上传平台", b_body)
        self.assertNotIn("已提交平台", b_body)
        self.assertLess(
            b_body.index('<div class="detail-grid billing-month-grid"'),
            b_body.index('<form method="get" action="/billing/platform-bills" class="billing-month-picker">'),
        )

    def test_platform_bill_workspace_requires_selection_and_keeps_selected_platform_after_upload(self):
        c_cookie = self.login("c_viewer", "demo123")
        empty_response = self.request("/billing/platform-bills?month=2026-04", cookie=c_cookie)
        empty_body = empty_response["body"].decode("utf-8")
        self.assertIn('aria-label="选择平台"', empty_body)
        self.assertIn("请选择平台", empty_body)
        self.assertIn("请选择平台后查看状态、账单文件和操作。", empty_body)
        self.assertNotIn("打开工作台", empty_body)

        selected_response = self.request("/billing/platform-bills?month=2026-04&platform=tmall", cookie=c_cookie)
        selected_body = selected_response["body"].decode("utf-8")
        self.assertIn('value="tmall" selected', selected_body)
        self.assertIn('name="platform_code" value="tmall"', selected_body)
        self.assertNotIn('name="platform_code" value="jd"', selected_body)

        upload_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-04"), ("platform_code", "tmall")],
                files=[("upload_files", "tmall.xlsx", b"tmall", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        self.assertTrue(upload_response["status"].startswith("302"))
        self.assertIn("month=2026-04", dict(upload_response["headers"]).get("Location", ""))
        self.assertIn("platform=tmall", dict(upload_response["headers"]).get("Location", ""))

    def test_admin_can_remove_unused_platform_from_platform_settings(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        month_key = "2026-04"
        configs = db.platform_bill_platform_configs_for_month(self.db_path, month_key)
        removed_code = "mini_program"
        payload = {"month_key": month_key, "selected_platform": "tmall", "remove_platform_code": removed_code}
        for index, item in enumerate(configs):
            payload[f"platform_code__{index}"] = item["code"]
            payload[f"platform_label__{index}"] = item["label"]
        response = self.request(
            "/billing/platform-bills/platforms",
            method="POST",
            body=urlencode(payload).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(response["status"].startswith("302"))
        self.assertIn("platform=tmall", dict(response["headers"]).get("Location", ""))
        updated_codes = {item["code"] for item in db.platform_bill_platform_configs_for_month(self.db_path, month_key)}
        self.assertNotIn(removed_code, updated_codes)

    def test_c_viewer_platform_bill_upload_limits_to_three_files(self):
        c_cookie = self.login("c_viewer", "demo123")
        response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-12"),
                    ("platform_code", "tmall"),
                ],
                files=[
                    ("upload_files", "f1.xlsx", b"1", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("upload_files", "f2.xlsx", b"2", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("upload_files", "f3.xlsx", b"3", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("upload_files", "f4.xlsx", b"4", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        self.assertTrue(response["status"].startswith("302"))
        redirected = dict(response["headers"]).get("Location", "")
        self.assertIn("month=2026-12", redirected)
        self.assertIn("%E5%8D%95%E6%AC%A1%E6%9C%80%E5%A4%9A%E4%B8%8A%E4%BC%A0+3+%E4%B8%AA%E6%96%87%E4%BB%B6", redirected)

    def test_c_viewer_platform_bill_total_file_count_cannot_exceed_three(self):
        c_cookie = self.login("c_viewer", "demo123")
        first_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-12"),
                    ("platform_code", "tmall"),
                ],
                files=[
                    ("upload_files", "tmall-1.xlsx", b"1", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("upload_files", "tmall-2.xlsx", b"2", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        self.assertTrue(first_response["status"].startswith("302"))
        second_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-12"),
                    ("platform_code", "tmall"),
                ],
                files=[
                    ("upload_files", "tmall-3.xlsx", b"3", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("upload_files", "tmall-4.xlsx", b"4", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        self.assertTrue(second_response["status"].startswith("302"))
        redirected = dict(second_response["headers"]).get("Location", "")
        self.assertIn(
            "%E8%AF%A5%E5%B9%B3%E5%8F%B0%E6%9C%80%E5%A4%9A%E4%BF%9D%E7%95%99+3+%E4%B8%AA%E8%B4%A6%E5%8D%95%E6%96%87%E4%BB%B6",
            redirected,
        )
        summary = db.platform_bill_month_summary(self.db_path, "2026-12")
        tmall = next(item for item in summary["platforms"] if item["platform_code"] == "tmall")
        self.assertIsNotNone(tmall["main_file"])
        self.assertEqual(len(tmall["attachments"]), 1)

    def test_c_viewer_platform_bill_submit_redirects_back_with_notice_when_incomplete(self):
        c_cookie = self.login("c_viewer", "demo123")
        c_user = next(item for item in db.list_users(self.db_path) if item["username"] == "c_viewer")
        with db.get_connection(self.db_path) as connection:
            db.get_or_create_billing_month(connection, "2026-02", c_user["id"])
        submit_response = self.request(
            "/billing/platform-bills/submit",
            method="POST",
            body=urlencode({"month_key": "2026-02", "platform_code": "tmall"}).encode("utf-8"),
            cookie=c_cookie,
        )
        self.assertTrue(submit_response["status"].startswith("302"))
        redirected = dict(submit_response["headers"]).get("Location", "")
        self.assertIn("/billing/platform-bills?", redirected)
        self.assertIn("month=2026-02", redirected)
        self.assertIn("%E5%A4%A9%E7%8C%AB%E8%BF%98%E6%B2%A1%E6%9C%89%E4%B8%8A%E4%BC%A0%E8%B4%A6%E5%8D%95%E6%96%87%E4%BB%B6", redirected)

    def test_b_editor_can_view_platform_bill_status_and_download(self):
        c_cookie = self.login("c_viewer", "demo123")
        upload_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-07"),
                    ("platform_code", "tmall"),
                ],
                files=[
                    ("upload_files", "tmall-main.xlsx", b"tmall-main", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=c_cookie,
        )
        self.assertTrue(upload_response["status"].startswith("302"))

        b_cookie = self.login("b_editor", "demo123")
        page_response = self.request("/billing/platform-bills?month=2026-07&platform=tmall", cookie=b_cookie)
        body = page_response["body"].decode("utf-8")
        self.assertIn("2026-07", body)
        self.assertIn("tmall-main.xlsx", body)
        self.assertIn("/billing/platform-bills/files/", body)

        summary = db.platform_bill_month_summary(self.db_path, "2026-07")
        tmall_main = next(item for item in summary["platforms"] if item["platform_code"] == "tmall")["main_file"]
        download_response = self.request(f"/billing/platform-bills/files/{tmall_main['id']}", cookie=b_cookie)
        self.assertTrue(download_response["status"].startswith("200"))
        self.assertEqual(download_response["body"], b"tmall-main")

    def test_b_editor_can_upload_brand_bill_versions_and_submit_to_a(self):
        b_cookie = self.login("b_editor", "demo123")
        dashboard_response = self.request(
            "/billing/brand-bills/dashboard",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-08")],
                files=[
                    (
                        "dashboard_file",
                        "dashboard-2026-08.xlsx",
                        self.make_brand_bill_template_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.assertTrue(dashboard_response["status"].startswith("302"))
        first_upload = self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-08"),
                    ("note", "初版汇总"),
                ],
                files=[
                    ("brand_bill_file", "brand-2026-08-v1.xlsx", self.make_brand_bill_template_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.assertTrue(first_upload["status"].startswith("302"))
        second_upload = self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-08"),
                    ("note", "修正版"),
                ],
                files=[
                    ("brand_bill_file", "brand-2026-08-v2.xlsx", self.make_brand_bill_template_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.assertTrue(second_upload["status"].startswith("302"))
        summary = db.brand_bill_month_summary(self.db_path, "2026-08")
        self.assertEqual(len(summary["versions"]), 2)
        self.assertEqual(summary["latest_version"]["original_filename"], "brand-2026-08-v2.xlsx")
        page_response = self.request("/billing/brand-bills?month=2026-08", cookie=b_cookie)
        body = page_response["body"].decode("utf-8")
        self.assertIn("马天奴月销看板", body)
        self.assertIn("2026-08当月账单", body)
        self.assertIn('class="brand-bill-status-heading"', body)
        self.assertIn("历史账单查询", body)
        self.assertNotIn("账单明细与流转", body)
        self.assertIn("年月", body)
        self.assertIn("平台", body)
        self.assertIn("店铺", body)
        self.assertIn("销售金额", body)
        self.assertIn("广州仓", body)
        self.assertIn("武汉仓", body)
        self.assertIn("天猫", body)
        self.assertIn("唯品会", body)
        self.assertIn("导入excel", body)
        self.assertIn("导出excel", body)
        self.assertIn("查询", body)
        self.assertNotIn("手工填写看板", body)
        self.assertNotIn("保存月销看板", body)
        self.assertIn(">上传</button>", body)
        self.assertIn(">删除</button>", body)
        self.assertIn("提交给跟单部", body)
        self.assertNotIn("<th>版本</th>", body)
        self.assertNotIn("<th>文件</th>", body)

        submit_response = self.request(
            "/billing/brand-bills/submit",
            method="POST",
            body=urlencode({"month_key": "2026-08"}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(submit_response["status"].startswith("302"))
        brand_bill = db.get_brand_bill_by_key(self.db_path, "2026-08")
        self.assertEqual(brand_bill["status"], "submitted_to_a")

    def test_b_editor_can_delete_only_current_unsubmitted_brand_bill_version(self):
        month_key = "2026-12"
        b_cookie = self.login("b_editor", "demo123")
        empty_page = self.request(f"/billing/brand-bills?month={month_key}", cookie=b_cookie)["body"].decode("utf-8")
        self.assertIn('class="brand-bill-delete-form"><button type="button" class="ghost-button" disabled>删除</button>', empty_page)
        self.assertIn('<button type="submit" disabled>提交给跟单部</button>', empty_page)
        for filename, content in (("brand-2026-12-v1.xlsx", b"brand-v1"), ("brand-2026-12-v2.xlsx", b"brand-v2")):
            upload_response = self.request(
                "/billing/brand-bills/upload",
                method="POST",
                body=self.build_multi_multipart(
                    fields=[("month_key", month_key)],
                    files=[
                        (
                            "brand_bill_file",
                            filename,
                            content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    ],
                ),
                content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
                cookie=b_cookie,
            )
            self.assertTrue(upload_response["status"].startswith("302"))

        page_response = self.request(f"/billing/brand-bills?month={month_key}", cookie=b_cookie)
        body = page_response["body"].decode("utf-8")
        self.assertIn('action="/billing/brand-bills/delete"', body)
        self.assertIn(">上传</button>", body)
        self.assertNotIn("上传新版本", body)
        history_start = body.index('<aside class="brand-bill-side">')
        history_end = body.index("</aside>", history_start)
        self.assertNotIn(">月份<", body[history_start:history_end])

        before_delete = db.brand_bill_month_summary(self.db_path, month_key)
        v2 = before_delete["latest_version"]
        self.assertEqual(v2["original_filename"], "brand-2026-12-v2.xlsx")
        a_cookie = self.login("a_editor", "demo123")
        forbidden_delete = self.request(
            "/billing/brand-bills/delete",
            method="POST",
            body=urlencode({"month_key": month_key}).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(forbidden_delete["status"].startswith("403"))

        delete_response = self.request(
            "/billing/brand-bills/delete",
            method="POST",
            body=urlencode({"month_key": month_key}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(delete_response["status"].startswith("302"))
        after_delete = db.brand_bill_month_summary(self.db_path, month_key)
        self.assertEqual(len(after_delete["versions"]), 1)
        self.assertEqual(after_delete["latest_version"]["original_filename"], "brand-2026-12-v1.xlsx")
        self.assertFalse((self.upload_dir / v2["stored_path"]).exists())

        submit_response = self.request(
            "/billing/brand-bills/submit",
            method="POST",
            body=urlencode({"month_key": month_key}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(submit_response["status"].startswith("302"))
        locked_delete = self.request(
            "/billing/brand-bills/delete",
            method="POST",
            body=urlencode({"month_key": month_key}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(locked_delete["status"].startswith("302"))
        self.assertEqual(len(db.brand_bill_month_summary(self.db_path, month_key)["versions"]), 1)

    def test_brand_bill_return_keeps_v1_and_allows_b_to_resubmit_v2(self):
        month_key = "2026-11"
        b_cookie = self.login("b_editor", "demo123")
        first_upload = self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", month_key)],
                files=[
                    (
                        "brand_bill_file",
                        "brand-2026-11-v1.xlsx",
                        b"brand-v1",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.assertTrue(first_upload["status"].startswith("302"))
        submit_response = self.request(
            "/billing/brand-bills/submit",
            method="POST",
            body=urlencode({"month_key": month_key}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(submit_response["status"].startswith("302"))

        b_page = self.request(f"/billing/brand-bills?month={month_key}", cookie=b_cookie)
        self.assertIn("申请退回", b_page["body"].decode("utf-8"))
        request_response = self.request(
            "/billing/brand-bills/return-request",
            method="POST",
            body=urlencode({"month_key": month_key, "reason": "发现汇总金额有误"}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(request_response["status"].startswith("302"))
        return_request = db.list_brand_bill_return_requests(self.db_path, month_key)[0]
        self.assertEqual(return_request["status"], "pending")

        forbidden_decision = self.request(
            "/billing/brand-bills/return-decision",
            method="POST",
            body=urlencode(
                {
                    "request_id": str(return_request["id"]),
                    "month_key": month_key,
                    "decision": "approve",
                }
            ).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(forbidden_decision["status"].startswith("403"))

        a_cookie = self.login("a_editor", "demo123")
        a_page = self.request(f"/billing/brand-bills?month={month_key}", cookie=a_cookie)
        a_body = a_page["body"].decode("utf-8")
        self.assertIn("退回商品部", a_body)
        self.assertIn("发现汇总金额有误", a_body)
        approve_response = self.request(
            "/billing/brand-bills/return-decision",
            method="POST",
            body=urlencode(
                {
                    "request_id": str(return_request["id"]),
                    "month_key": month_key,
                    "decision": "approve",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(approve_response["status"].startswith("302"))

        returned_summary = db.brand_bill_month_summary(self.db_path, month_key)
        self.assertEqual(returned_summary["brand_bill"]["status"], "draft")
        self.assertEqual(len(returned_summary["versions"]), 1)
        self.assertEqual(db.list_brand_bill_return_requests(self.db_path, month_key)[0]["status"], "approved")
        v1_download = self.request(f"/billing/brand-bills/files/{returned_summary['latest_version']['id']}", cookie=b_cookie)
        self.assertTrue(v1_download["status"].startswith("200"))
        self.assertEqual(v1_download["body"], b"brand-v1")

        second_upload = self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", month_key)],
                files=[
                    (
                        "brand_bill_file",
                        "brand-2026-11-v2.xlsx",
                        b"brand-v2",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.assertTrue(second_upload["status"].startswith("302"))
        resubmit_response = self.request(
            "/billing/brand-bills/submit",
            method="POST",
            body=urlencode({"month_key": month_key}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(resubmit_response["status"].startswith("302"))
        final_summary = db.brand_bill_month_summary(self.db_path, month_key)
        self.assertEqual(len(final_summary["versions"]), 2)
        self.assertEqual(final_summary["latest_version"]["original_filename"], "brand-2026-11-v2.xlsx")
        self.assertEqual(final_summary["brand_bill"]["status"], "submitted_to_a")

    def test_a_editor_can_view_and_download_brand_bill_but_cannot_upload(self):
        b_cookie = self.login("b_editor", "demo123")
        upload_response = self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[
                    ("month_key", "2026-09"),
                    ("note", "给跟单部的月账单"),
                ],
                files=[
                    ("brand_bill_file", "brand-2026-09-v1.xlsx", self.make_brand_bill_template_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=b_cookie,
        )
        self.assertTrue(upload_response["status"].startswith("302"))

        a_cookie = self.login("a_editor", "demo123")
        page_response = self.request("/billing/brand-bills?month=2026-09", cookie=a_cookie)
        body = page_response["body"].decode("utf-8")
        self.assertIn("品牌月账单", body)
        self.assertIn("马天奴月销看板", body)
        self.assertIn("brand-2026-09-v1.xlsx", body)
        self.assertIn("/billing/brand-bills/files/", body)
        self.assertIn("下载账单", body)
        self.assertIn("历史账单查询", body)

        version = db.brand_bill_month_summary(self.db_path, "2026-09")["latest_version"]
        download_response = self.request(f"/billing/brand-bills/files/{version['id']}", cookie=a_cookie)
        self.assertTrue(download_response["status"].startswith("200"))
        self.assertTrue(download_response["body"].startswith(b"PK"))
        self.assertIn("brand-2026-09-v1.xlsx", dict(download_response["headers"]).get("Content-Disposition", ""))

        template_response = self.request("/billing/brand-bills/template.xlsx", cookie=a_cookie)
        self.assertTrue(template_response["status"].startswith("200"))
        self.assertTrue(template_response["body"].startswith(b"PK"))
        self.assertIn("brand-bill-template.xlsx", dict(template_response["headers"]).get("Content-Disposition", ""))
        template_sheet = load_workbook(io.BytesIO(template_response["body"]), data_only=True).active
        self.assertEqual(template_sheet["F1"].value, "广州仓")
        self.assertEqual(template_sheet["J1"].value, "武汉仓")
        self.assertEqual(template_sheet["H2"].value, "销量占比")
        self.assertEqual(template_sheet["M2"].value, "销额占比")

        forbidden_upload = self.request(
            "/billing/brand-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-09")],
                files=[
                    ("brand_bill_file", "forbidden.xlsx", b"forbidden", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=a_cookie,
        )
        self.assertTrue(forbidden_upload["status"].startswith("403"))

    def test_a_editor_can_create_supplier_master_and_monthly_settlement(self):
        a_cookie = self.login("a_editor", "demo123")
        create_supplier = self.request(
            "/billing/supplier-settlements/suppliers",
            method="POST",
            body=urlencode(
                {
                    "supplier_code": "001",
                    "supplier_name": "原始供应商001",
                    "invoice_names": "青山\n绿水",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(create_supplier["status"].startswith("302"))
        supplier = db.get_supplier_by_code(self.db_path, "001")
        self.assertIsNotNone(supplier)
        self.assertEqual(supplier["supplier_name"], "原始供应商001")
        invoice_names = [item["invoice_name"] for item in supplier["invoice_names"]]
        self.assertEqual(invoice_names, ["青山", "绿水"])

        save_settlement = self.request(
            "/billing/supplier-settlements/records",
            method="POST",
            body=urlencode(
                {
                    "month_key": "2026-10",
                    "supplier_id": supplier["id"],
                    "invoice_name": "青山",
                    "amount_due": "12800",
                    "payment_status": "paid",
                    "payment_date": "2026-10-25",
                    "note": "10月已结清",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(save_settlement["status"].startswith("302"))
        month_summary = db.supplier_settlement_month_summary(self.db_path, "2026-10")
        self.assertEqual(len(month_summary["items"]), 1)
        item = month_summary["items"][0]
        self.assertEqual(item["supplier_code"], "001")
        self.assertEqual(item["invoice_name"], "青山")
        self.assertEqual(float(item["amount_due"]), 12800.0)
        self.assertEqual(float(item["amount_paid"]), 12800.0)
        self.assertEqual(item["payment_status"], "paid")

    def test_supplier_year_summary_accumulates_paid_and_unpaid(self):
        a_cookie = self.login("a_editor", "demo123")
        self.request(
            "/billing/supplier-settlements/suppliers",
            method="POST",
            body=urlencode(
                {
                    "supplier_code": "002",
                    "supplier_name": "原始供应商002",
                    "invoice_names": "海川",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        supplier = db.get_supplier_by_code(self.db_path, "002")
        self.assertIsNotNone(supplier)

        for payload in (
            {
                "month_key": "2026-01",
                "supplier_id": supplier["id"],
                "invoice_name": "海川",
                "amount_due": "1000",
                "payment_status": "paid",
                "payment_date": "2026-01-18",
                "note": "",
            },
            {
                "month_key": "2026-02",
                "supplier_id": supplier["id"],
                "invoice_name": "海川",
                "amount_due": "2500",
                "payment_status": "unpaid",
                "payment_date": "",
                "note": "",
            },
        ):
            response = self.request(
                "/billing/supplier-settlements/records",
                method="POST",
                body=urlencode(payload).encode("utf-8"),
                cookie=a_cookie,
            )
            self.assertTrue(response["status"].startswith("302"))

        year_summary = db.supplier_year_summary(self.db_path, supplier["id"], 2026)
        self.assertEqual(len(year_summary["items"]), 2)
        self.assertEqual(year_summary["total_due"], 3500.0)
        self.assertEqual(year_summary["total_paid"], 1000.0)
        self.assertEqual(year_summary["total_unpaid"], 2500.0)

    def test_b_and_c_users_cannot_access_supplier_settlement(self):
        for username in ("b_editor", "c_viewer"):
            cookie = self.login(username, "demo123")
            page_response = self.request("/billing/supplier-settlements", cookie=cookie)
            self.assertTrue(page_response["status"].startswith("403"))
            master_response = self.request(
                "/billing/supplier-settlements/master",
                method="POST",
                body=urlencode(
                    {
                        "supplier_code": "S001",
                        "supplier_name": "无权限供应商",
                        "supply_chain_manager": "测试",
                    }
                ).encode("utf-8"),
                cookie=cookie,
            )
            self.assertTrue(master_response["status"].startswith("403"))
            export_response = self.request(
                "/billing/supplier-settlements/bills/export.xlsx?start_month=2026-05&end_month=2026-05",
                cookie=cookie,
            )
            self.assertTrue(export_response["status"].startswith("403"))

    def test_a_editor_can_download_supplier_settlement_template_and_export(self):
        a_cookie = self.login("a_editor", "demo123")
        template_response = self.request("/billing/supplier-settlements/template.xlsx", cookie=a_cookie)
        self.assertTrue(template_response["status"].startswith("200"))
        template_workbook = load_workbook(io.BytesIO(template_response["body"]))
        template_headers = [cell.value for cell in template_workbook.active[1]]
        self.assertEqual(
            template_headers,
            ["供应商编码", "供应商名称", "开票抬头", "应付金额", "支付状态", "支付日期", "备注"],
        )

        self.request(
            "/billing/supplier-settlements/suppliers",
            method="POST",
            body=urlencode(
                {
                    "supplier_code": "010",
                    "supplier_name": "模板导出供应商",
                    "invoice_names": "春山",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        supplier = db.get_supplier_by_code(self.db_path, "010")
        self.request(
            "/billing/supplier-settlements/records",
            method="POST",
            body=urlencode(
                {
                    "month_key": "2026-12",
                    "supplier_id": supplier["id"],
                    "invoice_name": "春山",
                    "amount_due": "4321",
                    "payment_status": "unpaid",
                    "payment_date": "",
                    "note": "导出测试",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        export_response = self.request("/billing/supplier-settlements/export.xlsx?month=2026-12", cookie=a_cookie)
        self.assertTrue(export_response["status"].startswith("200"))
        export_workbook = load_workbook(io.BytesIO(export_response["body"]))
        self.assertEqual(export_workbook.active.max_row, 2)
        self.assertEqual(export_workbook.active["A2"].value, "010")
        self.assertEqual(export_workbook.active["B2"].value, "模板导出供应商")
        self.assertEqual(export_workbook.active["D2"].value, 4321)

    def test_a_editor_can_import_supplier_settlement_workbook(self):
        a_cookie = self.login("a_editor", "demo123")
        self.request(
            "/billing/supplier-settlements/suppliers",
            method="POST",
            body=urlencode(
                {
                    "supplier_code": "020",
                    "supplier_name": "导入供应商020",
                    "invoice_names": "青石",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["供应商编码", "供应商名称", "开票抬头", "应付金额", "支付状态", "支付日期", "备注"])
        sheet.append(["020", "导入供应商020", "青石", 5600, "已支付", "2026-12-20", "导入成功"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        import_response = self.request(
            "/billing/supplier-settlements/import",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-12")],
                files=[
                    ("settlement_workbook", "supplier-import.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=a_cookie,
        )
        self.assertTrue(import_response["status"].startswith("302"))
        month_summary = db.supplier_settlement_month_summary(self.db_path, "2026-12")
        self.assertEqual(len(month_summary["items"]), 1)
        item = month_summary["items"][0]
        self.assertEqual(item["supplier_code"], "020")
        self.assertEqual(item["invoice_name"], "青石")
        self.assertEqual(item["payment_status"], "paid")
        self.assertEqual(float(item["amount_paid"]), 5600.0)

    def test_a_supplier_bill_query_supports_code_names_and_latest_month_version(self):
        a_cookie = self.login("a_editor", "demo123")
        for supplier_name in ("广美舟", "广美舟贸易"):
            response = self.request(
                "/billing/supplier-settlements/master",
                method="POST",
                body=urlencode(
                    {
                        "supplier_code": "S001",
                        "supplier_name": supplier_name,
                        "supply_chain_manager": "王敏",
                    }
                ).encode("utf-8"),
                cookie=a_cookie,
            )
            self.assertTrue(response["status"].startswith("302"))

        master_names = db.list_supplier_master_names(self.db_path, "S001")
        self.assertEqual([item["supplier_name"] for item in master_names], ["广美舟", "广美舟贸易"])
        initial_query_response = self.request("/billing/supplier-settlements", cookie=a_cookie)
        initial_query_body = initial_query_response["body"].decode("utf-8")
        self.assertIn('.supplier-name-option[hidden]', initial_query_body)
        self.assertIn('data-supplier-code="S001" hidden', initial_query_body)
        first_version = self.make_supplier_bill_workbook_bytes(
            [
                ["S001", "广美舟", "经销", "王敏", "GM-01", "马天奴", "连衣裙-红", 3, 120, 360],
                ["S001", "广美舟贸易", "经销", "王敏", "GM-02", "马天奴", "衬衫-白", 2, 90, 180],
            ]
        )
        import_response = self.request(
            "/billing/supplier-settlements/bills/import",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("period_month", "2026-05")],
                files=[
                    (
                        "bill_workbook",
                        "supplier-bill-v1.xlsx",
                        first_version,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type="multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest",
            cookie=a_cookie,
        )
        self.assertTrue(import_response["status"].startswith("302"))

        selected_ids = ",".join(str(item["id"]) for item in master_names)
        query_response = self.request(
            "/billing/supplier-settlements?"
            + urlencode(
                {
                    "view": "bills",
                    "start_month": "2026-05",
                    "end_month": "2026-05",
                    "supplier_code": "S001",
                    "supplier_name_ids": selected_ids,
                }
            ),
            cookie=a_cookie,
        )
        query_body = query_response["body"].decode("utf-8")
        self.assertTrue(query_response["status"].startswith("200"))
        self.assertLess(query_body.index('class="page-back-row"'), query_body.index('class="supplier-settlement-layout"'))
        self.assertIn('class="supplier-settlement-columns"', query_body)
        self.assertIn('class="panel supplier-settlement-main-panel"', query_body)
        self.assertIn('class="panel supplier-bill-query-panel"', query_body)
        self.assertIn('class="panel supplier-master-panel"', query_body)
        self.assertIn("<h2>账单导入</h2>", query_body)
        self.assertIn("<h2>账单查询</h2>", query_body)
        self.assertIn("<h2>供应商管理</h2>", query_body)
        self.assertIn("<h3>查询汇总</h3>", query_body)
        self.assertIn('input id="supplier-code-filter"', query_body)
        self.assertIn('datalist id="supplier-code-suggestions"', query_body)
        self.assertNotIn('select id="supplier-code-filter"', query_body)
        self.assertNotIn('data-supplier-code="S001" hidden', query_body)
        self.assertIn("540.00", query_body)
        self.assertIn(">5<", query_body)
        self.assertNotIn("GM-01</td>", query_body)

        export_response = self.request(
            "/billing/supplier-settlements/bills/export.xlsx?"
            + urlencode(
                {
                    "start_month": "2026-05",
                    "end_month": "2026-05",
                    "supplier_code": "S001",
                    "supplier_name_ids": str(master_names[0]["id"]),
                }
            ),
            cookie=a_cookie,
        )
        self.assertTrue(export_response["status"].startswith("200"))
        export_sheet = load_workbook(io.BytesIO(export_response["body"]), data_only=True).active
        self.assertEqual(export_sheet["A1"].value, "所属月份")
        self.assertEqual(export_sheet["F2"].value, "GM-01")
        self.assertEqual(export_sheet["H2"].value, "连衣裙-红")

        second_version = self.make_supplier_bill_workbook_bytes(
            [["S001", "广美舟", "经销", "王敏", "GM-03", "马天奴", "半裙-黑", 4, 150, 600]]
        )
        replacement_response = self.request(
            "/billing/supplier-settlements/bills/import",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("period_month", "2026-05")],
                files=[
                    (
                        "bill_workbook",
                        "supplier-bill-v2.xlsx",
                        second_version,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type="multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest",
            cookie=a_cookie,
        )
        self.assertTrue(replacement_response["status"].startswith("302"))
        result = db.query_supplier_bill_lines(self.db_path, "2026-05", "2026-05", "S001")
        self.assertEqual(result["quantity_total"], 4)
        self.assertEqual(result["settlement_amount_total"], 600.0)
        self.assertEqual(result["items"][0]["supplier_style_code"], "GM-03")

    def test_a_supplier_bill_delete_and_reupload_only_within_initial_30_days(self):
        a_cookie = self.login("a_editor", "demo123")
        self.request(
            "/billing/supplier-settlements/master",
            method="POST",
            body=urlencode(
                {
                    "supplier_code": "S930",
                    "supplier_name": "删除测试供应商",
                    "supply_chain_manager": "王敏",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        initial_page = self.request("/billing/supplier-settlements", cookie=a_cookie)
        initial_body = initial_page["body"].decode("utf-8")
        self.assertIn('formaction="/billing/supplier-settlements/bills/delete"', initial_body)
        self.assertIn("首次导入后 30 天内可删除并重新上传", initial_body)

        first_workbook = self.make_supplier_bill_workbook_bytes(
            [["S930", "删除测试供应商", "经销", "王敏", "S930-01", "马天奴", "测试款色", 2, 100, 200]]
        )
        first_import = self.request(
            "/billing/supplier-settlements/bills/import",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("period_month", "2026-07")],
                files=[
                    (
                        "bill_workbook",
                        "supplier-bill-v1.xlsx",
                        first_workbook,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type="multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest",
            cookie=a_cookie,
        )
        self.assertTrue(first_import["status"].startswith("302"))

        delete_response = self.request(
            "/billing/supplier-settlements/bills/delete",
            method="POST",
            body=urlencode({"period_month": "2026-07"}).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(delete_response["status"].startswith("302"))
        self.assertIn("当前账单已删除", unquote_plus(dict(delete_response["headers"])["Location"]))
        self.assertEqual(db.query_supplier_bill_lines(self.db_path, "2026-07", "2026-07", "S930")["items"], [])

        replacement_workbook = self.make_supplier_bill_workbook_bytes(
            [["S930", "删除测试供应商", "经销", "王敏", "S930-02", "马天奴", "更新款色", 3, 120, 360]]
        )
        replacement_import = self.request(
            "/billing/supplier-settlements/bills/import",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("period_month", "2026-07")],
                files=[
                    (
                        "bill_workbook",
                        "supplier-bill-v2.xlsx",
                        replacement_workbook,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type="multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest",
            cookie=a_cookie,
        )
        self.assertTrue(replacement_import["status"].startswith("302"))
        self.assertIn("V2", unquote_plus(dict(replacement_import["headers"])["Location"]))
        current_bills = db.query_supplier_bill_lines(self.db_path, "2026-07", "2026-07", "S930")["items"]
        self.assertEqual(current_bills[0]["supplier_style_code"], "S930-02")

        expired_timestamp = (datetime.now(timezone.utc) - timedelta(days=31)).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        with db.get_connection(self.db_path) as connection:
            connection.execute(
                "UPDATE supplier_bill_batches SET imported_at = ? WHERE period_month = ? AND version_no = 1",
                (expired_timestamp, "2026-07"),
            )

        expired_delete = self.request(
            "/billing/supplier-settlements/bills/delete",
            method="POST",
            body=urlencode({"period_month": "2026-07"}).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(expired_delete["status"].startswith("302"))
        self.assertIn("超过首次导入后 30 天", unquote_plus(dict(expired_delete["headers"])["Location"]))

        expired_reupload = self.request(
            "/billing/supplier-settlements/bills/import",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("period_month", "2026-07")],
                files=[
                    (
                        "bill_workbook",
                        "supplier-bill-v3.xlsx",
                        first_workbook,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type="multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest",
            cookie=a_cookie,
        )
        self.assertTrue(expired_reupload["status"].startswith("302"))
        self.assertIn("超过首次导入后 30 天", unquote_plus(dict(expired_reupload["headers"])["Location"]))
        self.assertEqual(
            db.query_supplier_bill_lines(self.db_path, "2026-07", "2026-07", "S930")["items"][0]["supplier_style_code"],
            "S930-02",
        )

        b_cookie = self.login("b_editor", "demo123")
        forbidden_delete = self.request(
            "/billing/supplier-settlements/bills/delete",
            method="POST",
            body=urlencode({"period_month": "2026-07"}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(forbidden_delete["status"].startswith("403"))

    def test_a_supplier_bill_import_rejects_fractional_quantity(self):
        a_cookie = self.login("a_editor", "demo123")
        self.request(
            "/billing/supplier-settlements/master",
            method="POST",
            body=urlencode(
                {
                    "supplier_code": "S002",
                    "supplier_name": "数量校验供应商",
                    "supply_chain_manager": "李四",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        workbook = self.make_supplier_bill_workbook_bytes(
            [["S002", "数量校验供应商", "经销", "李四", "S-01", "马天奴", "测试款", 1.5, 100, 150]]
        )
        response = self.request(
            "/billing/supplier-settlements/bills/import",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("period_month", "2026-06")],
                files=[
                    (
                        "bill_workbook",
                        "fractional-quantity.xlsx",
                        workbook,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ],
            ),
            content_type="multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest",
            cookie=a_cookie,
        )
        self.assertTrue(response["status"].startswith("302"))
        self.assertIn("数量必须是非负整数", unquote_plus(dict(response["headers"])["Location"]))
        result = db.query_supplier_bill_lines(self.db_path, "2026-06", "2026-06", "S002")
        self.assertEqual(result["items"], [])

    def test_a_supplier_master_excel_import_and_bidirectional_lookup(self):
        a_cookie = self.login("a_editor", "demo123")
        template_response = self.request(
            "/billing/supplier-settlements/master/template.xlsx",
            cookie=a_cookie,
        )
        self.assertTrue(template_response["status"].startswith("200"))
        template_sheet = load_workbook(io.BytesIO(template_response["body"]), data_only=True).active
        self.assertEqual([cell.value for cell in template_sheet[1]], ["供应商编号", "供应商名称", "供应链经理"])

        workbook = self.make_supplier_master_workbook_bytes(
            [
                ["S900", "青山制衣", "王敏"],
                ["S900", "绿水制衣", "王敏"],
            ]
        )
        import_response = self.request(
            "/billing/supplier-settlements/master/import",
            method="POST",
            body=self.build_multi_multipart(
                files=[
                    (
                        "supplier_master_workbook",
                        "supplier-masters.xlsx",
                        workbook,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                ]
            ),
            content_type="multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest",
            cookie=a_cookie,
        )
        self.assertTrue(import_response["status"].startswith("302"))
        self.assertIn("已导入 2 条供应商信息", unquote_plus(dict(import_response["headers"])["Location"]))

        name_query_response = self.request(
            "/billing/supplier-settlements?master_name=%E9%9D%92%E5%B1%B1",
            cookie=a_cookie,
        )
        name_query_body = name_query_response["body"].decode("utf-8")
        self.assertIn("青山制衣", name_query_body)
        self.assertIn("S900", name_query_body)
        self.assertIn("王敏", name_query_body)
        self.assertIn("新建供应商", name_query_body)
        self.assertIn("按供应商名称", name_query_body)
        self.assertIn("按供应商编号", name_query_body)
        self.assertIn('/billing/supplier-settlements/master/new', name_query_body)

        code_query_response = self.request(
            "/billing/supplier-settlements?master_code=S900",
            cookie=a_cookie,
        )
        code_query_body = code_query_response["body"].decode("utf-8")
        self.assertIn("青山制衣", code_query_body)
        self.assertIn("绿水制衣", code_query_body)
        self.assertIn("王敏", code_query_body)

        form_response = self.request("/billing/supplier-settlements/master/new", cookie=a_cookie)
        self.assertTrue(form_response["status"].startswith("200"))
        self.assertIn("单个供应商录入", form_response["body"].decode("utf-8"))
        single_save_response = self.request(
            "/billing/supplier-settlements/master",
            method="POST",
            body=urlencode(
                {
                    "return_to": "master_form",
                    "supplier_code": "S901",
                    "supplier_name": "云帆制衣",
                    "supply_chain_manager": "李四",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(single_save_response["status"].startswith("302"))
        self.assertTrue(dict(single_save_response["headers"])["Location"].startswith("/billing/supplier-settlements/master/new?"))

    def test_a_editor_can_edit_supplier_master_and_sync_bill_reference(self):
        a_cookie = self.login("a_editor", "demo123")
        create_response = self.request(
            "/billing/supplier-settlements/master",
            method="POST",
            body=urlencode(
                {
                    "supplier_code": "S910",
                    "supplier_name": "原供应商名称",
                    "supply_chain_manager": "原负责人",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))
        master = db.list_supplier_master_names(self.db_path, "S910")[0]
        a_user = next(item for item in db.list_users(self.db_path) if item["username"] == "a_editor")
        with db.get_connection(self.db_path) as connection:
            db.create_supplier_bill_batch(
                connection,
                "2026-07",
                "supplier-master-edit.xlsx",
                "billing/test/supplier-master-edit.xlsx",
                [
                    {
                        "supplier_master_name_id": master["id"],
                        "supplier_code": "S910",
                        "supplier_name": "原供应商名称",
                        "supply_chain_manager": "原负责人",
                        "mode": "经销",
                        "supplier_style_code": "S910-01",
                        "brand_name": "马天奴",
                        "style_color": "测试款色",
                        "quantity": 2,
                        "tax_included_price": 100,
                        "settlement_amount": 200,
                        "source_row_no": 1,
                    }
                ],
                a_user["id"],
            )

        query_response = self.request("/billing/supplier-settlements?master_code=S910", cookie=a_cookie)
        query_body = query_response["body"].decode("utf-8")
        edit_url = f'/billing/supplier-settlements/master/edit?id={master["id"]}'
        self.assertIn(edit_url, query_body)

        edit_page = self.request(edit_url, cookie=a_cookie)
        self.assertTrue(edit_page["status"].startswith("200"))
        self.assertIn("编辑供应商", edit_page["body"].decode("utf-8"))

        save_response = self.request(
            "/billing/supplier-settlements/master/edit",
            method="POST",
            body=urlencode(
                {
                    "supplier_master_name_id": master["id"],
                    "supplier_code": "S911",
                    "supplier_name": "更正供应商名称",
                    "supply_chain_manager": "新负责人",
                }
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(save_response["status"].startswith("302"))
        self.assertEqual(db.list_supplier_master_names(self.db_path, "S910"), [])
        updated_master = db.list_supplier_master_names(self.db_path, "S911")[0]
        self.assertEqual(updated_master["supplier_name"], "更正供应商名称")
        self.assertEqual(updated_master["supply_chain_manager"], "新负责人")

        updated_bills = db.query_supplier_bill_lines(self.db_path, "2026-07", "2026-07", "S911")["items"]
        self.assertEqual(len(updated_bills), 1)
        self.assertEqual(updated_bills[0]["supplier_name"], "更正供应商名称")
        self.assertEqual(updated_bills[0]["supply_chain_manager"], "新负责人")

        b_cookie = self.login("b_editor", "demo123")
        forbidden_edit = self.request(edit_url, cookie=b_cookie)
        self.assertTrue(forbidden_edit["status"].startswith("403"))

    def test_c_viewer_cannot_see_hidden_price_field_in_list_or_api(self):
        cookie = self.login("c_viewer", "demo123")
        list_response = self.request("/products", cookie=cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertNotIn("<th>上新价格</th>", list_body)
        self.assertNotIn("新建资料", list_body)
        self.assertIn("已完成", list_body)
        self.assertNotIn(">待B填写<", list_body)
        self.assertNotIn("近 7 天运营概览", list_body)
        self.assertIn("历时天数", list_body)

        api_response = self.request("/api/products", cookie=cookie)
        payload = json.loads(api_response["body"].decode("utf-8"))
        self.assertEqual(payload["role"], "C")
        self.assertEqual(payload["count"], 1)
        first_item = payload["items"][0]
        self.assertNotIn("launch_price", first_item)
        self.assertNotIn("supplier", first_item)
        self.assertEqual(first_item["status"], "published")
        self.assertEqual(first_item["status_label"], "已完成")
        self.assertIn("elapsed_days", first_item)
        self.assertIn("elapsed_days_label", first_item)

    def test_c_viewer_can_bulk_receive_completed_products(self):
        cookie = self.login("c_viewer", "demo123")
        list_response = self.request("/products", cookie=cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn("批量接收资料", list_body)
        self.assertIn(">接收</button>", list_body)
        self.assertLess(
            list_body.index('<section class="products-insights-grid products-insights-single">'),
            list_body.index('<div class="products-main-stack">'),
        )
        for label in ("总接收", "近7天新增", "待接收"):
            self.assertIn(label, list_body)
        summary_start = list_body.index("<h2>资料概览</h2>")
        summary_block = list_body[summary_start:list_body.index("</section>", summary_start)]
        self.assertNotIn("总资料数", summary_block)
        self.assertLess(summary_block.index("<span>总接收</span>"), summary_block.index("<span>近7天新增</span>"))
        self.assertLess(summary_block.index("<span>近7天新增</span>"), summary_block.index("<span>待接收</span>"))
        self.assertIn("grid-template-rows: minmax(288px, 7fr) minmax(150px, 3fr);", list_body)

        receive_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "1"), ("bulk_action", "receive_selected")]).encode("utf-8"),
            cookie=cookie,
        )
        self.assertTrue(receive_response["status"].startswith("302"))
        product = db.get_product(self.db_path, 1)
        c_user = next(item for item in db.list_users(self.db_path) if item["username"] == "c_viewer")
        self.assertEqual(product["status"], "received")
        self.assertTrue(
            db.c_product_received_by_user(
                self.db_path,
                product["id"],
                c_user["id"],
                product["c_release_no"],
            )
        )
        self.assertEqual(
            db.c_user_receipt_stats(self.db_path, c_user),
            {"total": 1, "received": 1, "pending": 0, "recent_created": 1},
        )

        received_list = self.request("/products?status=received", cookie=cookie)
        received_body = received_list["body"].decode("utf-8")
        self.assertIn("已接收", received_body)
        self.assertNotIn(">接收</button>", received_body)
        self.assertNotIn('href="/products/1/edit"', received_body)

    def test_c_viewer_can_receive_completed_product_from_detail(self):
        cookie = self.login("c_viewer", "demo123")
        detail_response = self.request("/products/1", cookie=cookie)
        detail_body = detail_response["body"].decode("utf-8")
        self.assertIn("接收资料", detail_body)

        receive_response = self.request(
            "/products/1/status",
            method="POST",
            body=urlencode({"status": "received"}).encode("utf-8"),
            cookie=cookie,
        )
        self.assertTrue(receive_response["status"].startswith("302"))
        product = db.get_product(self.db_path, 1)
        c_user = next(item for item in db.list_users(self.db_path) if item["username"] == "c_viewer")
        self.assertEqual(product["status"], "received")
        self.assertTrue(
            db.c_product_received_by_user(
                self.db_path,
                product["id"],
                c_user["id"],
                product["c_release_no"],
            )
        )

    def test_c_channel_scope_isolates_products_and_tracks_same_style_receipts_per_account(self):
        vip_user_id = db.create_user(
            self.db_path,
            "vip_operator",
            "唯品运营李二",
            "C",
            "demo123",
            must_change_password=False,
            operating_channel="vip",
        )
        users = {item["username"]: item for item in db.list_users(self.db_path)}
        tmall_user = users["c_viewer"]
        a_user = users["a_editor"]
        b_user = users["b_editor"]
        with db.get_connection(self.db_path) as connection:
            created_ids = []
            for index, (launch_channel, product_name) in enumerate(
                (("唯品", "唯品渠道资料"), ("同款", "同款渠道资料")),
                start=1,
            ):
                product_id = db.create_product(
                    connection,
                    self.a_complete_fields_payload(
                        style_color=f"渠道款色-{index}",
                        style_code=f"CHANNEL-{index}",
                        product_name=product_name,
                        launch_price="299",
                        launch_channel=launch_channel,
                        completion_flag="Y",
                    ),
                    a_user["id"],
                    "A",
                )
                db.change_product_status(connection, product_id, "pending", a_user["id"], "提交给商品部填写", "测试渠道流转。")
                db.change_product_status(connection, product_id, "published", b_user["id"], "填写完成，开放给运营部", "测试渠道流转。")
                created_ids.append(product_id)
        vip_product_id, same_product_id = created_ids
        tmall_cookie = self.login("c_viewer", "demo123")
        vip_cookie = self.login("vip_operator", "demo123")

        tmall_list = self.request("/products", cookie=tmall_cookie)["body"].decode("utf-8")
        self.assertNotIn("唯品渠道资料", tmall_list)
        self.assertIn("同款渠道资料", tmall_list)
        self.assertIn("上新渠道", tmall_list)
        vip_list = self.request("/products", cookie=vip_cookie)["body"].decode("utf-8")
        self.assertIn("唯品渠道资料", vip_list)
        self.assertIn("同款渠道资料", vip_list)
        self.assertTrue(self.request(f"/products/{vip_product_id}", cookie=tmall_cookie)["status"].startswith("403"))

        receive_response = self.request(
            f"/products/{same_product_id}/status",
            method="POST",
            body=urlencode({"status": "received"}).encode("utf-8"),
            cookie=tmall_cookie,
        )
        self.assertTrue(receive_response["status"].startswith("302"))
        same_product = db.get_product(self.db_path, same_product_id)
        self.assertEqual(same_product["status"], "received")
        self.assertTrue(
            db.c_product_received_by_user(
                self.db_path,
                same_product_id,
                tmall_user["id"],
                same_product["c_release_no"],
            )
        )
        self.assertFalse(
            db.c_product_received_by_user(
                self.db_path,
                same_product_id,
                vip_user_id,
                same_product["c_release_no"],
            )
        )
        vip_same_detail = self.request(f"/products/{same_product_id}", cookie=vip_cookie)["body"].decode("utf-8")
        self.assertIn("接收资料", vip_same_detail)
        vip_receive_response = self.request(
            f"/products/{same_product_id}/status",
            method="POST",
            body=urlencode({"status": "received"}).encode("utf-8"),
            cookie=vip_cookie,
        )
        self.assertTrue(vip_receive_response["status"].startswith("302"))
        self.assertTrue(
            db.c_product_received_by_user(
                self.db_path,
                same_product_id,
                vip_user_id,
                same_product["c_release_no"],
            )
        )
        self.assertEqual(db.get_product(self.db_path, same_product_id)["status"], "received")

    def test_catalog_export_omits_hidden_size_and_total_fields(self):
        cookie = self.login("a_editor", "demo123")
        response = self.request("/export.xlsx", cookie=cookie)
        workbook = load_workbook(io.BytesIO(response["body"]))
        headers = [cell.value for cell in workbook.active[1]]
        for hidden_header in ("F", "S", "M", "L", "XL", "2XL", "3XL", "合计"):
            self.assertNotIn(hidden_header, headers)

    def test_admin_must_assign_operating_channel_for_c_account(self):
        cookie = self.login("admin_reviewer", "demo123")
        incomplete_response = self.request(
            "/users",
            method="POST",
            body=urlencode(
                {
                    "username": "missing_scope",
                    "display_name": "未归属运营",
                    "department": "C",
                    "password": "demo123",
                }
            ).encode("utf-8"),
            cookie=cookie,
        )
        self.assertTrue(incomplete_response["status"].startswith("400"))
        self.assertIn("必须选择天猫类或唯品类", incomplete_response["body"].decode("utf-8"))

        create_response = self.request(
            "/users",
            method="POST",
            body=urlencode(
                {
                    "username": "tmall_operator",
                    "display_name": "天猫运营王大",
                    "department": "C",
                    "operating_channel": "tmall",
                    "billing_platform_codes__tmall": "tmall",
                    "password": "demo123",
                }
            ).encode("utf-8"),
            cookie=cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))
        created_user = next(item for item in db.list_users(self.db_path) if item["username"] == "tmall_operator")
        self.assertEqual(created_user["operating_channel"], "tmall")
        self.assertEqual(json.loads(created_user["billing_platforms_json"]), ["tmall"])

    def test_c_account_separates_product_channel_from_multi_platform_bill_permissions(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        create_response = self.request(
            "/users",
            method="POST",
            body=urlencode(
                {
                    "username": "d_platform_operator",
                    "display_name": "天猫运营王大",
                    "department": "C",
                    "operating_channel": "tmall",
                    "billing_platform_codes__tmall": "tmall",
                    "billing_platform_codes__jd": "jd",
                    "password": "demo123",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))
        managed_user = next(item for item in db.list_users(self.db_path) if item["username"] == "d_platform_operator")
        self.assertEqual(managed_user["operating_channel"], "tmall")
        self.assertEqual(json.loads(managed_user["billing_platforms_json"]), ["tmall", "jd"])

        d_cookie = self.login("d_platform_operator", "demo123")
        page_body = self.request("/billing/platform-bills?month=2026-06&platform=jd", cookie=d_cookie)["body"].decode("utf-8")
        self.assertIn('<option value="tmall"', page_body)
        self.assertIn('<option value="jd" selected', page_body)
        self.assertNotIn('<option value="vip"', page_body)
        self.assertIn('name="platform_code" value="jd"', page_body)

        upload_response = self.request(
            "/billing/platform-bills/upload",
            method="POST",
            body=self.build_multi_multipart(
                fields=[("month_key", "2026-06"), ("platform_code", "jd")],
                files=[("upload_files", "jd.xlsx", b"jd", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=d_cookie,
        )
        self.assertTrue(upload_response["status"].startswith("302"))

    def test_b_channel_alias_is_normalized_to_tmall(self):
        b_user = next(item for item in db.list_users(self.db_path) if item["username"] == "b_editor")
        product = db.get_product(self.db_path, 1)
        form = self.app.normalized_form_for_stage(
            b_user,
            product,
            {"launch_channel": "天猫/京东/抖音"},
        )
        self.assertEqual(form["launch_channel"], "天猫")

    def test_products_list_merges_owner_and_creator_into_initiator_column(self):
        with db.get_connection(self.db_path) as connection:
            connection.execute(
                "UPDATE users SET display_name = ? WHERE username = ?",
                ("张三", "a_editor"),
            )
        cookie = self.login("a_editor", "demo123")
        response = self.request("/products", cookie=cookie)
        body = response["body"].decode("utf-8")
        self.assertIn("<th>发起人</th>", body)
        self.assertNotIn("<th>发起部门</th>", body)
        self.assertNotIn("<th>录入人</th>", body)
        self.assertIn('class="table-initiator-cell"', body)
        self.assertIn("张三", body)

    def test_c_viewer_products_page_does_not_error_when_no_published_products(self):
        with db.get_connection(self.db_path) as connection:
            connection.execute("UPDATE products SET status = 'draft'")
        cookie = self.login("c_viewer", "demo123")
        response = self.request("/products", cookie=cookie)
        body = response["body"].decode("utf-8")
        self.assertTrue(response["status"].startswith("200"))
        self.assertIn("当前账号为运营部", body)
        self.assertIn("可查看资料", body)

    def test_api_requires_login_or_valid_c_token(self):
        anonymous_response = self.request("/api/products")
        self.assertTrue(anonymous_response["status"].startswith("401"))
        anonymous_payload = json.loads(anonymous_response["body"].decode("utf-8"))
        self.assertEqual(anonymous_payload["error"], "unauthorized")

        admin_cookie = self.login("admin_reviewer", "demo123")
        token_update = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("field_keys__style_code", "style_code"),
                    ("rotate_token", "1"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(token_update["status"].startswith("302"))
        token = db.get_setting(self.db_path, "c_api_token")
        self.assertTrue(token)

        invalid_response = self.request("/api/products", authorization="Bearer invalid-token")
        self.assertTrue(invalid_response["status"].startswith("401"))

        valid_response = self.request("/api/products", authorization=f"Bearer {token}")
        self.assertTrue(valid_response["status"].startswith("200"))
        valid_payload = json.loads(valid_response["body"].decode("utf-8"))
        self.assertEqual(valid_payload["role"], "C")
        self.assertEqual(valid_payload["count"], 1)
        self.assertIn("style_code", valid_payload["items"][0])
        self.assertNotIn("supplier", valid_payload["items"][0])

        query_token_response = self.request(f"/api/products?access_token={token}")
        self.assertTrue(query_token_response["status"].startswith("200"))
        query_payload = json.loads(query_token_response["body"].decode("utf-8"))
        self.assertEqual(query_payload["role"], "C")

    def test_healthz_returns_basic_runtime_status(self):
        response = self.request("/healthz")
        self.assertTrue(response["status"].startswith("200"))
        payload = json.loads(response["body"].decode("utf-8"))
        self.assertEqual(payload["status"], "ok")
        self.assertTrue(payload["db_exists"])
        self.assertEqual(payload["user_count"], 4)

    def test_login_is_temporarily_locked_after_repeated_failures(self):
        for _ in range(db.LOGIN_FAILURE_LIMIT - 1):
            response = self.request(
                "/login",
                method="POST",
                body=urlencode({"username": "a_editor", "password": "wrong-pass"}).encode("utf-8"),
            )
            self.assertTrue(response["status"].startswith("401"))
        lock_response = self.request(
            "/login",
            method="POST",
            body=urlencode({"username": "a_editor", "password": "wrong-pass"}).encode("utf-8"),
        )
        self.assertTrue(lock_response["status"].startswith("423"))
        body = lock_response["body"].decode("utf-8")
        self.assertIn("临时锁定", body)

        blocked_correct_response = self.request(
            "/login",
            method="POST",
            body=urlencode({"username": "a_editor", "password": "demo123"}).encode("utf-8"),
        )
        self.assertTrue(blocked_correct_response["status"].startswith("423"))

    def test_successful_login_clears_failure_counter(self):
        wrong_response = self.request(
            "/login",
            method="POST",
            body=urlencode({"username": "a_editor", "password": "wrong-pass"}).encode("utf-8"),
        )
        self.assertTrue(wrong_response["status"].startswith("401"))
        status_before = db.get_login_attempt_status(self.db_path, "a_editor")
        self.assertEqual(status_before["failure_count"], 1)

        login_response = self.request(
            "/login",
            method="POST",
            body=urlencode({"username": "a_editor", "password": "demo123"}).encode("utf-8"),
        )
        self.assertTrue(login_response["status"].startswith("302"))
        status_after = db.get_login_attempt_status(self.db_path, "a_editor")
        self.assertEqual(status_after["failure_count"], 0)

    def test_init_db_can_skip_demo_seed_and_create_bootstrap_admin(self):
        clean_db_path = Path(self.temp_dir.name) / "clean-catalog.db"
        init_db(
            clean_db_path,
            seed_demo=False,
            seed_samples=False,
            bootstrap_admin={
                "username": "owner_admin",
                "display_name": "正式管理员",
                "password": "OwnerPass123",
                "must_change_password": True,
            },
        )
        users = db.list_users(clean_db_path)
        self.assertEqual(len(users), 1)
        self.assertEqual(users[0]["username"], "owner_admin")
        self.assertEqual(users[0]["department"], "ADMIN")
        self.assertIsNotNone(db.authenticate_user(clean_db_path, "owner_admin", "OwnerPass123"))

    def test_init_db_without_demo_seed_does_not_create_sample_products(self):
        clean_db_path = Path(self.temp_dir.name) / "no-demo-catalog.db"
        init_db(clean_db_path, seed_demo=False, seed_samples=False)
        self.assertEqual(db.list_users(clean_db_path), [])
        self.assertEqual(db.list_products(clean_db_path), [])

    def test_admin_can_disable_c_api_token(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        generate_response = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("rotate_token", "1"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(generate_response["status"].startswith("302"))
        token = db.get_setting(self.db_path, "c_api_token")
        self.assertTrue(token)

        missing_confirmation_response = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("disable_token", "1"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(missing_confirmation_response["status"].startswith("400"))

        disable_response = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("confirm_text", "DISABLE"),
                    ("disable_token", "1"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(disable_response["status"].startswith("302"))
        self.assertEqual(db.get_setting(self.db_path, "c_api_token"), "")

        revoked_response = self.request("/api/products", authorization=f"Bearer {token}")
        self.assertTrue(revoked_response["status"].startswith("401"))

    def test_c_viewer_export_contains_only_open_fields(self):
        cookie = self.login("c_viewer", "demo123")
        response = self.request("/export.xlsx", cookie=cookie)
        workbook = load_workbook(io.BytesIO(response["body"]))
        headers = [cell.value for cell in workbook.active[1]]
        self.assertIn("历时天数", headers)
        self.assertIn("资料完成", headers)
        self.assertIn("商品名称", headers)
        self.assertNotIn("上新价格", headers)
        self.assertNotIn("供应商", headers)
        self.assertEqual(workbook.active.max_row, 2)

    def test_c_viewer_can_export_only_selected_products(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        self.make_product_two_a_complete()
        publish_response = self.request(
            "/products/2/status",
            method="POST",
            body=urlencode({"status": "published", "review_note": "用于运营部勾选导出测试"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(publish_response["status"].startswith("302"))

        c_cookie = self.login("c_viewer", "demo123")
        list_response = self.request("/products", cookie=c_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn('name="product_ids"', list_body)
        self.assertIn("导出 Excel", list_body)
        self.assertIn("导出全部资料", list_body)
        self.assertIn("导出勾选资料", list_body)
        self.assertIn('id="toggle-all-products"', list_body)

        selected_response = self.request("/export.xlsx?selected=2", cookie=c_cookie)
        self.assertTrue(selected_response["status"].startswith("200"))
        workbook = load_workbook(io.BytesIO(selected_response["body"]))
        headers = [cell.value for cell in workbook.active[1]]
        self.assertIn("历时天数", headers)
        self.assertIn("资料完成", headers)
        self.assertNotIn("上新价格", headers)
        self.assertEqual(workbook.active.max_row, 2)
        product_name_col = headers.index("商品名称") + 1
        self.assertEqual(workbook.active.cell(2, product_name_col).value, "毛感针织开衫")

    def test_c_viewer_export_selected_requires_checked_items(self):
        c_cookie = self.login("c_viewer", "demo123")
        response = self.request("/export.xlsx?mode=selected", cookie=c_cookie)
        self.assertTrue(response["status"].startswith("302"))
        headers = dict(response["headers"])
        notice = unquote_plus(headers["Location"])
        self.assertIn("请先勾选至少一条资料", notice)

    def test_export_with_images_embeds_primary_product_images(self):
        image_bytes = self.make_png_bytes()
        cookie = self.login("a_editor", "demo123")
        with patch.object(self.app, "fetch_export_image", return_value=image_bytes) as fetch_image:
            response = self.request("/export.xlsx?mode=selected&include_images=1&selected=1", cookie=cookie)
        self.assertTrue(response["status"].startswith("200"))
        self.assertIn("catalog-export-with-images.xlsx", dict(response["headers"])["Content-Disposition"])
        worksheet = load_workbook(io.BytesIO(response["body"])).active
        self.assertEqual(fetch_image.call_count, 1)
        self.assertEqual(len(worksheet._images), 1)
        image_column = [cell.value for cell in worksheet[1]].index("图片") + 1
        self.assertIsNone(worksheet.cell(2, image_column).value)
        self.assertGreaterEqual(worksheet.row_dimensions[2].height, 96)

    def test_export_menu_is_available_for_all_logged_in_roles(self):
        a_cookie = self.login("a_editor", "demo123")
        a_response = self.request("/products", cookie=a_cookie)
        a_body = a_response["body"].decode("utf-8")
        self.assertIn("导出 Excel", a_body)
        self.assertIn("导出全部资料", a_body)
        self.assertIn("导出勾选资料", a_body)
        self.assertIn("导出勾选含图片", a_body)
        self.assertIn('data-base-href="/export.xlsx?mode=selected&amp;include_images=1"', a_body)
        self.assertIn(".products-overview-card {", a_body)
        self.assertIn("overflow: visible", a_body)
        self.assertIn(".products-overview-card {\n      overflow: visible;\n      z-index: 1;", a_body)
        self.assertIn(".nav-shell {", a_body)
        self.assertIn("top: 12px;\n      z-index: 20;", a_body)
        self.assertIn('data-export-selected="1"', a_body)
        self.assertIn('menu.addEventListener("pointerleave"', a_body)
        self.assertIn('menu.addEventListener("pointerenter"', a_body)
        self.assertIn(".export-menu-panel::before", a_body)
        self.assertIn("menu.open = false", a_body)

        admin_cookie = self.login("admin_reviewer", "demo123")
        admin_response = self.request("/products", cookie=admin_cookie)
        admin_body = admin_response["body"].decode("utf-8")
        self.assertIn("导出 Excel", admin_body)
        self.assertIn("导出全部资料", admin_body)
        self.assertIn("导出勾选资料", admin_body)
        self.assertIn("列表字段设置", admin_body)

    def test_nav_uses_export_menu_for_logged_in_user(self):
        a_cookie = self.login("a_editor", "demo123")
        response = self.request("/products", cookie=a_cookie)
        body = response["body"].decode("utf-8")
        self.assertNotIn("export-menu-summary export-menu-summary-compact", body)
        self.assertNotIn('<li class="nav-chip"><a href="/export.xlsx">导出 Excel</a></li>', body)
        self.assertNotIn('<li class="nav-chip"><a href="/products/new">新建资料</a></li>', body)
        self.assertNotIn('<li class="nav-chip"><a href="/import">导入 Excel</a></li>', body)
        self.assertNotIn('<li class="nav-chip"><a href="/api/products">JSON 调用</a></li>', body)

        admin_cookie = self.login("admin_reviewer", "demo123")
        admin_response = self.request("/products", cookie=admin_cookie)
        admin_body = admin_response["body"].decode("utf-8")
        self.assertNotIn('<li class="nav-chip"><a href="/products/review">流转看板</a></li>', admin_body)

    def test_department_can_open_and_save_list_layout_settings(self):
        a_cookie = self.login("a_editor", "demo123")
        settings_response = self.request("/settings/list-layout", cookie=a_cookie)
        settings_body = settings_response["body"].decode("utf-8")
        self.assertIn("列表字段设置", settings_body)
        self.assertIn("资料完成", settings_body)
        self.assertIn("completion_flag", settings_body)
        self.assertIn("品牌名称", settings_body)
        self.assertIn("款号", settings_body)
        self.assertIn("送拍时间", settings_body)
        self.assertIn("合作模式", settings_body)
        self.assertNotIn(">F<", settings_body)
        self.assertNotIn(">S<", settings_body)
        self.assertNotIn(">M<", settings_body)
        self.assertNotIn(">L<", settings_body)
        self.assertNotIn(">XL<", settings_body)
        self.assertNotIn(">2XL<", settings_body)
        self.assertNotIn(">3XL<", settings_body)
        self.assertNotIn(">合计<", settings_body)

        save_response = self.request(
            "/settings/list-layout",
            method="POST",
            body=urlencode(
                [
                    ("field_order__1", "style_code"),
                    ("field_order__2", "brand_name"),
                    ("field_order__3", "category"),
                    ("field_order__4", "product_name"),
                    ("field_rank__style_code", "1"),
                    ("field_rank__brand_name", "2"),
                    ("field_rank__category", "4"),
                    ("field_rank__product_name", "3"),
                    ("field_keys__style_code", "style_code"),
                    ("field_keys__brand_name", "brand_name"),
                    ("field_keys__product_name", "product_name"),
                ]
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(save_response["status"].startswith("302"))

        list_response = self.request("/products", cookie=a_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn("<th>款号</th><th>品牌名称</th><th>商品名称</th>", list_body)
        self.assertNotIn("<th>品类</th><th>状态</th>", list_body)
        self.assertEqual(db.get_setting(self.db_path, "list_layout_customized_A"), "1")

    def test_executive_role_matches_a_visibility_but_is_read_only(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        users_page = self.request("/users", cookie=admin_cookie)
        users_body = users_page["body"].decode("utf-8")
        self.assertIn('value="EXECUTIVE"', users_body)
        self.assertIn(">总经办</option>", users_body)
        create_form = users_body.split('<form method="post" action="/users">', 1)[1].split("</form>", 1)[0]
        self.assertIn("account-create-grid", create_form)
        self.assertNotIn("<span>显示名称</span>", create_form)
        self.assertLess(create_form.index("用户名"), create_form.index("初始密码"))
        self.assertLess(create_form.index("初始密码"), create_form.index("首次登录"))
        self.assertLess(create_form.index("首次登录"), create_form.index("角色/部门"))
        self.assertLess(create_form.index("角色/部门"), create_form.index("渠道属性"))
        self.assertLess(create_form.index("渠道属性"), create_form.index("账单属性（可多选）"))

        create_response = self.request(
            "/users",
            method="POST",
            body=urlencode(
                {
                    "username": "executive_viewer",
                    "department": "EXECUTIVE",
                    "password": "demo123",
                    "must_change_password": "",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))
        executive_user = next(item for item in db.list_users(self.db_path) if item["username"] == "executive_viewer")
        self.assertEqual(executive_user["display_name"], "executive_viewer")
        executive_cookie = self.login("executive_viewer", "demo123")

        products_response = self.request("/products", cookie=executive_cookie)
        products_body = products_response["body"].decode("utf-8")
        self.assertTrue(products_response["status"].startswith("200"))
        self.assertIn("总经办只读模式", products_body)
        self.assertIn("<strong>executive_viewer</strong>", products_body)
        self.assertIn('<div class="nav-role-note">总经办</div>', products_body)
        self.assertNotIn('href="/products/new"', products_body)
        self.assertNotIn('href="/import"', products_body)
        self.assertNotIn('href="/settings/list-layout"', products_body)
        self.assertNotIn("<th>图片</th>", products_body)
        self.assertNotIn("<th>上新价格</th>", products_body)

        self.assertTrue(self.request("/export.xlsx", cookie=executive_cookie)["status"].startswith("200"))
        self.assertTrue(self.request("/billing", cookie=executive_cookie)["status"].startswith("200"))
        self.assertTrue(self.request("/billing/brand-bills", cookie=executive_cookie)["status"].startswith("200"))
        self.assertTrue(
            self.request("/billing/supplier-settlements/master/template.xlsx", cookie=executive_cookie)["status"].startswith("200")
        )

        supplier_response = self.request("/billing/supplier-settlements", cookie=executive_cookie)
        supplier_body = supplier_response["body"].decode("utf-8")
        self.assertIn("总经办账号仅可查询与导出账单明细", supplier_body)
        self.assertNotIn('action="/billing/supplier-settlements/bills/import"', supplier_body)
        self.assertNotIn('action="/billing/supplier-settlements/master/import"', supplier_body)

        self.assertTrue(
            self.request("/import", cookie=executive_cookie)["status"].startswith("403")
        )
        self.assertTrue(
            self.request("/products/bulk", method="POST", body=b"", cookie=executive_cookie)["status"].startswith("403")
        )
        self.assertTrue(
            self.request("/billing/monthly-board", method="POST", body=b"", cookie=executive_cookie)["status"].startswith("403")
        )
        self.assertTrue(
            self.request("/settings/list-layout", method="POST", body=b"", cookie=executive_cookie)["status"].startswith("403")
        )

    def test_default_list_layout_follows_catalog_export_header_order(self):
        a_cookie = self.login("a_editor", "demo123")
        response = self.request("/products", cookie=a_cookie)
        body = response["body"].decode("utf-8")
        self.assertIn("<th>资料完成</th>", body)
        self.assertIn("<th>资料完成</th><th>送拍时间</th><th>送检时间</th>", body)
        self.assertIn("<th>送拍时间</th>", body)
        self.assertIn("<th>送检时间</th>", body)
        self.assertIn("<th>检测报告</th><th>尺寸表</th><th>发货仓库</th>", body)
        self.assertIn("<th>品牌名称</th><th>年份季节</th><th>款色</th>", body)
        self.assertIn("<th>款色</th><th>款号</th><th>颜色名称</th>", body)
        self.assertIn("<th>合作模式</th>", body)
        self.assertIn("<th>供应链经理</th>", body)
        self.assertIn("<th>含税价</th>", body)
        self.assertIn("<th>成分(英文)</th>", body)
        self.assertIn("<th>69码</th>", body)
        self.assertIn("<th>尺寸表</th><th>发货仓库</th><th>品牌名称</th>", body)
        self.assertNotIn("<th>图片</th>", body)
        self.assertNotIn("<th>上新价格</th>", body)
        self.assertNotIn("<th>F</th>", body)
        self.assertNotIn("<th>S</th>", body)
        self.assertNotIn("<th>M</th>", body)
        self.assertNotIn("<th>L</th>", body)
        self.assertNotIn("<th>XL</th>", body)
        self.assertNotIn("<th>2XL</th>", body)
        self.assertNotIn("<th>3XL</th>", body)
        self.assertNotIn("<th>合计</th>", body)

    def test_list_layout_settings_can_hide_completion_flag_when_department_does_not_select_it(self):
        a_cookie = self.login("a_editor", "demo123")
        save_response = self.request(
            "/settings/list-layout",
            method="POST",
            body=urlencode(
                [
                    ("field_order__1", "style_code"),
                    ("field_order__2", "brand_name"),
                    ("field_order__3", "product_name"),
                    ("field_rank__style_code", "1"),
                    ("field_rank__brand_name", "2"),
                    ("field_rank__product_name", "3"),
                    ("field_keys__style_code", "style_code"),
                    ("field_keys__brand_name", "brand_name"),
                    ("field_keys__product_name", "product_name"),
                ]
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(save_response["status"].startswith("302"))

        list_response = self.request("/products", cookie=a_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertNotIn("<th>资料完成</th>", list_body)

    def test_c_list_layout_defaults_to_current_visible_fields_until_customized(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("field_keys__style_code", "style_code"),
                    ("field_keys__launch_price", "launch_price"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )

        c_cookie = self.login("c_viewer", "demo123")
        list_response = self.request("/products", cookie=c_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn("<th>款号</th><th>商品名称</th><th>上新价格</th>", list_body)
        self.assertNotIn("<th>资料完成</th>", list_body)
        self.assertNotIn("<th>品牌名称</th>", list_body)

        save_layout = self.request(
            "/settings/list-layout",
            method="POST",
            body=urlencode(
                [
                    ("field_rank__launch_price", "1"),
                    ("field_rank__style_code", "2"),
                    ("field_rank__product_name", "3"),
                    ("field_keys__launch_price", "launch_price"),
                    ("field_keys__style_code", "style_code"),
                ]
            ).encode("utf-8"),
            cookie=c_cookie,
        )
        self.assertTrue(save_layout["status"].startswith("302"))
        self.assertEqual(db.get_setting(self.db_path, "list_layout_customized_C"), "1")

        customized_response = self.request("/products", cookie=c_cookie)
        customized_body = customized_response["body"].decode("utf-8")
        self.assertIn("<th>上新价格</th><th>款号</th>", customized_body)
        self.assertNotIn("<th>商品名称</th><th>款号</th><th>上新价格</th>", customized_body)

    def test_product_list_keeps_empty_cells_blank_instead_of_unfilled_label(self):
        with db.get_connection(self.db_path) as connection:
            connection.execute("UPDATE products SET cooperation_mode = '' WHERE id = 2")

        a_cookie = self.login("a_editor", "demo123")
        list_response = self.request("/products", cookie=a_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn('class="table-cell-empty"', list_body)
        self.assertNotIn('<span class="meta">未填写</span>', list_body)

    def test_placeholder_excel_date_does_not_break_products_page(self):
        with db.get_connection(self.db_path) as connection:
            connection.execute("UPDATE products SET inspection_date = '1900' WHERE id = 2")

        product = db.get_product(self.db_path, 2)
        self.assertIsNone(db.parse_utc("1900"))
        self.assertIn("inspection_date", db.completion_missing_field_keys(product))

        a_cookie = self.login("a_editor", "demo123")
        list_response = self.request("/products", cookie=a_cookie)
        self.assertTrue(list_response["status"].startswith("200"))
        list_body = list_response["body"].decode("utf-8")
        self.assertNotIn(">1900<", list_body)

    def test_catalog_export_headers_can_be_mapped_by_import_parser(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "历时天数",
                "资料完成",
                "送拍时间",
                "送检时间",
                "检测报告",
                "尺寸表",
                "发货仓库",
                "品牌\n名称",
                "年份季节",
                "图片",
                "款色",
                "款号",
                "颜色\n名称",
                "商品名称",
                "品类",
                "是否有配饰",
                "供应商",
                "合作模式",
                "供应链经理",
                "含税价",
                "吊牌价",
                "上新价格",
                "上新渠道",
                "尺码段",
                "材质",
                "成分(英文)",
                "洗涤方式",
                "洗涤方式(英文)",
                "安全技术类别",
                "执行标准",
                "69码",
            ]
        )
        sheet.append(
            [
                "0 天",
                "",
                "2026-06-18",
                "2026-06-20",
                "有",
                "",
                "武汉仓",
                "马天奴",
                "2026春夏",
                "",
                "VESQ21SWH0220",
                "VESQ21SWH0",
                "灰色",
                "两件套（罩衫）",
                "套装",
                "",
                "",
                "联营",
                "王主管",
                "699",
                "1580",
                "329",
                "唯品",
                "F",
                "面料：莱赛尔83.7%锦纶16.3%",
                "SHELL: LYOCELL 83.7%NYLON 16.3%",
                "手洗",
                "HAND WASH",
                "GB 18401-2010 B类",
                "Q/MTN 008-2023",
                "12",
            ]
        )
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        from catalog_backend.excel import parse_workbook

        products = parse_workbook(buffer)
        self.assertEqual(len(products), 1)
        first = products[0]
        self.assertEqual(first["shooting_date"], "2026-06-18")
        self.assertEqual(first["inspection_date"], "2026-06-20")
        self.assertEqual(first["cooperation_mode"], "联营")
        self.assertEqual(first["supply_chain_manager"], "王主管")
        self.assertEqual(first["tax_included_price"], 699.0)
        self.assertEqual(first["composition_en"], "SHELL: LYOCELL 83.7%NYLON 16.3%")
        self.assertEqual(first["size_69"], 12)

    def test_composition_field_is_removed_and_legacy_header_maps_to_material(self):
        from catalog_backend.excel import parse_workbook, workbook_bytes
        from catalog_backend.fields import PRODUCT_FIELD_MAP, PRODUCT_FIELDS

        self.assertNotIn("composition", PRODUCT_FIELD_MAP)
        template = load_workbook(io.BytesIO(workbook_bytes([], PRODUCT_FIELDS)), data_only=True)
        headers = [cell.value for cell in template.active[1]]
        self.assertNotIn("成分", headers)
        self.assertIn("材质", headers)
        self.assertIn("成分(英文)", headers)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["商品名称", "款号", "成分"])
        worksheet.append(["旧模板商品", "LEGACY-001", "100%棉"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        products = parse_workbook(buffer)
        self.assertEqual(products[0]["material"], "100%棉")
        self.assertNotIn("composition", products[0])

    def test_import_updates_owned_record(self):
        cookie = self.login("a_editor", "demo123")
        workbook = Workbook()
        sheet = workbook.active
        headers = [
            "检测报告", "发货仓库", "品牌\n名称", "年份季节", "图片", "款色", "款号", "颜色\n名称",
            "商品名称", "品类", "是否有配饰", "供应商", "吊牌价", "上新价格", "上新渠道", "尺码段", "F", "S", "M",
            "L", "XL", "2XL", "3XL", "合计", "材质", "成分", "洗涤方式", "洗涤方式 （英文）",
            "安全技术类别", "执行标准", "尺寸表",
        ]
        sheet.append(headers)
        sheet.append([
            "已归档", "杭州一仓", "North Harbor", "2026夏", "https://example.com/images/nh-2601.jpg",
            "短袖连衣裙-蓝", "NH-2601", "海盐蓝", "褶皱短袖连衣裙", "连衣裙", "无", "杭州云锦供应链",
            499, 359, "直播首发", "S-XL", "", 20, 28, 18, 10, "", "", 76, "梭织",
            "面料 85%棉 15%锦纶", "建议冷水轻柔机洗", "Machine wash cold, gentle cycle", "B类",
            "GB/T 2660", "S: 肩宽37 / 胸围92 / 衣长112",
        ])
        buffer = io.BytesIO()
        workbook.save(buffer)
        file_body = self.build_multipart("workbook", "import.xlsx", buffer.getvalue())
        response = self.request(
            "/import",
            method="POST",
            body=file_body,
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=cookie,
        )
        body = response["body"].decode("utf-8")
        self.assertIn("新增 0 条，更新 1 条", body)

    def test_product_form_hides_size_chart_and_shows_launch_channel(self):
        cookie = self.login("a_editor", "demo123")
        response = self.request("/products/1/edit", cookie=cookie)
        body = response["body"].decode("utf-8")
        self.assertNotIn('name="launch_channel"', body)
        self.assertNotIn('name="completion_flag"', body)
        self.assertNotIn('name="size_chart"', body)

    def test_b_editor_only_sees_launch_fields_in_edit_form(self):
        cookie = self.login("b_editor", "demo123")
        response = self.request("/products/2/edit", cookie=cookie)
        body = response["body"].decode("utf-8")
        self.assertIn("上新价格", body)
        self.assertIn("上新渠道", body)
        self.assertIn("资料完成", body)
        self.assertIn('name="launch_price"', body)
        self.assertIn('name="launch_channel"', body)
        self.assertIn('name="completion_flag"', body)
        self.assertNotIn('name="brand_name"', body)
        self.assertNotIn('name="size_chart"', body)

    def test_admin_can_review_and_view_logs(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        self.make_product_two_a_complete()
        queue_response = self.request("/products/review", cookie=admin_cookie)
        queue_body = queue_response["body"].decode("utf-8")
        self.assertIn("资料流转看板", queue_body)
        self.assertIn("#2", queue_body)
        self.assertIn("待商品部填写", queue_body)
        self.assertIn("批量完成", queue_body)
        self.assertIn("最近提交", queue_body)

        publish_response = self.request(
            "/products/2/status",
            method="POST",
            body=urlencode({"status": "published", "review_note": "管理员确认可直接开放给 C"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(publish_response["status"].startswith("302"))

        logs_response = self.request("/products/2/logs", cookie=admin_cookie)
        logs_body = logs_response["body"].decode("utf-8")
        self.assertIn("管理员代为完成", logs_body)
        self.assertIn("提交给商品部填写", logs_body)
        self.assertIn("管理员确认可直接开放给 C", logs_body)

    def test_review_queue_supports_filter_and_bulk_publish(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        self.make_product_two_a_complete()
        filtered_response = self.request("/products/review?department=A&q=毛感", cookie=admin_cookie)
        filtered_body = filtered_response["body"].decode("utf-8")
        self.assertIn("毛感针织开衫", filtered_body)
        self.assertNotIn("褶皱短袖连衣裙", filtered_body)

        bulk_response = self.request(
            "/products/review/bulk",
            method="POST",
            body=urlencode([("product_ids", "2")]).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(bulk_response["status"].startswith("302"))
        updated_product = db.get_product(self.db_path, 2)
        self.assertEqual(updated_product["status"], "published")

        logs_response = self.request("/products/2/logs", cookie=admin_cookie)
        logs_body = logs_response["body"].decode("utf-8")
        self.assertIn("管理员代为完成", logs_body)

    def test_status_forms_show_review_note_input_and_log_editor_note(self):
        editor_cookie = self.login("a_editor", "demo123")
        create_response = self.request(
            "/products/new",
            method="POST",
            body=urlencode(
                self.a_complete_fields_payload(
                    product_name="备注测试新款",
                    style_code="NOTE-001",
                    brand_name="Note Brand",
                    category="上衣",
                    style_color="备注测试款色",
                    color_name="云朵白",
                    supplier="备注供应商",
                    image_url="https://example.com/images/note-001.jpg",
                )
            ).encode("utf-8"),
            cookie=editor_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))

        created_product = db.list_products(self.db_path, query="NOTE-001")[0]
        detail_response = self.request(f"/products/{created_product['id']}", cookie=editor_cookie)
        detail_body = detail_response["body"].decode("utf-8")
        self.assertIn("阶段流转", detail_body)
        self.assertIn('name="review_note"', detail_body)
        self.assertIn("交接说明（选填）", detail_body)

        pending_response = self.request(
            f"/products/{created_product['id']}/status",
            method="POST",
            body=urlencode({"status": "pending", "review_note": "主体资料已完成，请 B 部门补价格和渠道"}).encode("utf-8"),
            cookie=editor_cookie,
        )
        self.assertTrue(pending_response["status"].startswith("302"))

        logs = db.get_product_logs(self.db_path, created_product["id"])
        submit_log = next(item for item in logs if item["action"] == "status:pending")
        self.assertIn("交接说明", submit_log["details"])
        self.assertIn("主体资料已完成，请 B 部门补价格和渠道", submit_log["details"])

    def test_admin_reject_note_is_visible_in_logs(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        reject_response = self.request(
            "/products/2/status",
            method="POST",
            body=urlencode({"status": "draft", "review_note": "请补充检测报告后重新提交"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(reject_response["status"].startswith("302"))

        logs_response = self.request("/products/2/logs", cookie=admin_cookie)
        logs_body = logs_response["body"].decode("utf-8")
        self.assertIn("管理员退回跟单部", logs_body)
        self.assertIn("处理说明", logs_body)
        self.assertIn("请补充检测报告后重新提交", logs_body)

    def test_admin_can_manage_users(self):
        admin_cookie = self.login("admin_reviewer", "demo123")

        create_response = self.request(
            "/users",
            method="POST",
            body=urlencode(
                {
                    "username": "new_editor",
                    "display_name": "新建资料员",
                    "department": "A",
                    "password": "start123",
                    "must_change_password": "on",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))

        users_page = self.request("/users", cookie=admin_cookie)
        users_body = users_page["body"].decode("utf-8")
        self.assertIn("new_editor", users_body)
        self.assertNotIn("<th>显示名称</th>", users_body)
        self.assertIn('class="user-account-primary-action"', users_body)
        self.assertIn('class="user-account-reset-action"', users_body)

        new_user = db.list_users(self.db_path)[-1]
        self.assertEqual(new_user["username"], "new_editor")
        self.assertEqual(new_user["department"], "A")

        update_response = self.request(
            f"/users/{new_user['id']}/edit",
            method="POST",
            body=urlencode({"display_name": "改名后的资料员", "department": "B"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(update_response["status"].startswith("302"))
        updated_user = db.get_user_record(self.db_path, new_user["id"])
        self.assertEqual(updated_user["display_name"], "改名后的资料员")
        self.assertEqual(updated_user["department"], "B")

        toggle_response = self.request(
            f"/users/{new_user['id']}/toggle",
            method="POST",
            body=urlencode({"confirm_text": "DISABLE"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(toggle_response["status"].startswith("302"))
        disabled_user = db.get_user_record(self.db_path, new_user["id"])
        self.assertEqual(disabled_user["is_active"], 0)
        self.assertIsNone(db.authenticate_user(self.db_path, "new_editor", "start123"))

        reenable_response = self.request(
            f"/users/{new_user['id']}/toggle",
            method="POST",
            body=b"",
            cookie=admin_cookie,
        )
        self.assertTrue(reenable_response["status"].startswith("302"))

        reset_response = self.request(
            f"/users/{new_user['id']}/reset-password",
            method="POST",
            body=urlencode({"new_password": "fresh456", "confirm_text": "RESET"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(reset_response["status"].startswith("302"))
        self.assertIsNotNone(db.authenticate_user(self.db_path, "new_editor", "fresh456"))

    def test_high_risk_user_actions_require_confirmation(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        create_response = self.request(
            "/users",
            method="POST",
            body=urlencode(
                {
                    "username": "confirm_user",
                    "display_name": "确认测试账号",
                    "department": "A",
                    "password": "start123",
                    "must_change_password": "on",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))
        new_user = db.list_users(self.db_path)[-1]

        missing_disable_confirmation = self.request(
            f"/users/{new_user['id']}/toggle",
            method="POST",
            body=b"",
            cookie=admin_cookie,
        )
        self.assertTrue(missing_disable_confirmation["status"].startswith("400"))

        missing_reset_confirmation = self.request(
            f"/users/{new_user['id']}/reset-password",
            method="POST",
            body=urlencode({"new_password": "fresh456"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(missing_reset_confirmation["status"].startswith("400"))

    def test_user_must_change_password_before_accessing_products(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        create_response = self.request(
            "/users",
            method="POST",
            body=urlencode(
                {
                    "username": "force_change_user",
                    "display_name": "强制改密测试",
                    "department": "A",
                    "password": "start123",
                    "must_change_password": "on",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))

        login_response = self.request(
            "/login",
            method="POST",
            body=urlencode({"username": "force_change_user", "password": "start123"}).encode("utf-8"),
        )
        headers = dict(login_response["headers"])
        self.assertEqual(headers["Location"], "/profile/password")
        cookie = headers["Set-Cookie"].split(";", 1)[0]

        redirect_response = self.request("/products", cookie=cookie)
        redirect_headers = dict(redirect_response["headers"])
        self.assertEqual(redirect_headers["Location"], "/profile/password")

        password_page = self.request("/profile/password", cookie=cookie)
        password_body = password_page["body"].decode("utf-8")
        self.assertIn("必须修改密码", password_body)

        change_response = self.request(
            "/profile/password",
            method="POST",
            body=urlencode(
                {
                    "current_password": "start123",
                    "new_password": "fresh456",
                    "confirm_password": "fresh456",
                }
            ).encode("utf-8"),
            cookie=cookie,
        )
        self.assertTrue(change_response["status"].startswith("302"))
        change_headers = dict(change_response["headers"])
        self.assertIn("/products?notice=", change_headers["Location"])

        updated_user = db.authenticate_user(self.db_path, "force_change_user", "fresh456")
        self.assertIsNotNone(updated_user)
        self.assertEqual(updated_user["must_change_password"], 0)

    def test_user_can_change_password_from_profile(self):
        cookie = self.login("a_editor", "demo123")
        page_response = self.request("/profile/password", cookie=cookie)
        page_body = page_response["body"].decode("utf-8")
        self.assertIn("修改登录密码", page_body)

        bad_response = self.request(
            "/profile/password",
            method="POST",
            body=urlencode(
                {
                    "current_password": "wrongpass",
                    "new_password": "newpass123",
                    "confirm_password": "newpass123",
                }
            ).encode("utf-8"),
            cookie=cookie,
        )
        self.assertTrue(bad_response["status"].startswith("400"))
        self.assertIn("当前密码不正确", bad_response["body"].decode("utf-8"))

        change_response = self.request(
            "/profile/password",
            method="POST",
            body=urlencode(
                {
                    "current_password": "demo123",
                    "new_password": "newpass123",
                    "confirm_password": "newpass123",
                }
            ).encode("utf-8"),
            cookie=cookie,
        )
        self.assertTrue(change_response["status"].startswith("302"))
        self.assertIsNotNone(db.authenticate_user(self.db_path, "a_editor", "newpass123"))

    def test_editor_can_upload_product_image_and_preview_media(self):
        cookie = self.login("a_editor", "demo123")
        png_bytes = b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Wn3FoAAAAAASUVORK5CYII="
        )
        multipart_body = self.build_multipart(
            "image_upload",
            "sample.png",
            png_bytes,
            extra_fields={
                "product_name": "上传图片测试款",
                "style_code": "PIC-001",
                "brand_name": "Image Brand",
                "category": "上衣",
            },
            file_content_type="image/png",
        )
        response = self.request(
            "/products/new",
            method="POST",
            body=multipart_body,
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=cookie,
        )
        self.assertTrue(response["status"].startswith("302"))

        created_product = db.list_products(self.db_path, query="PIC-001")[0]
        self.assertTrue(str(created_product["image_url"]).startswith("/media/"))

        detail_response = self.request(f"/products/{created_product['id']}", cookie=cookie)
        detail_body = detail_response["body"].decode("utf-8")
        self.assertIn(created_product["image_url"], detail_body)

        media_response = self.request(created_product["image_url"], cookie=cookie)
        self.assertTrue(media_response["status"].startswith("200"))
        headers = dict(media_response["headers"])
        self.assertEqual(headers["Content-Type"], "image/png")
        self.assertGreater(len(media_response["body"]), 10)

    def test_editor_can_upload_multiple_images_and_reorder_gallery(self):
        cookie = self.login("a_editor", "demo123")
        png_bytes = b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Wn3FoAAAAAASUVORK5CYII="
        )
        multipart_body = self.build_multi_multipart(
            fields=[
                ("product_name", "多图测试款"),
                ("style_code", "PIC-MULTI-001"),
                ("brand_name", "Gallery Brand"),
                ("category", "上衣"),
            ],
            files=[
                ("image_uploads", "gallery-1.png", png_bytes, "image/png"),
                ("image_uploads", "gallery-2.png", png_bytes, "image/png"),
            ],
        )
        response = self.request(
            "/products/new",
            method="POST",
            body=multipart_body,
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=cookie,
        )
        self.assertTrue(response["status"].startswith("302"))

        created_product = db.list_products(self.db_path, query="PIC-MULTI-001")[0]
        gallery = json.loads(created_product["image_gallery_json"])
        self.assertEqual(len(gallery), 2)
        self.assertEqual(created_product["image_url"], gallery[0])

        detail_response = self.request(f"/products/{created_product['id']}", cookie=cookie)
        detail_body = detail_response["body"].decode("utf-8")
        self.assertIn("第 1 张", detail_body)
        self.assertIn("第 2 张", detail_body)

        reorder_body = self.build_multi_multipart(
            fields=[
                ("product_name", "多图测试款"),
                ("style_code", "PIC-MULTI-001"),
                ("brand_name", "Gallery Brand"),
                ("category", "上衣"),
                ("style_color", ""),
                ("color_name", ""),
                ("season_year", ""),
                ("has_accessories", ""),
                ("supplier", ""),
                ("tag_price", ""),
                ("size_range", ""),
                ("size_f", ""),
                ("size_s", ""),
                ("size_m", ""),
                ("size_l", ""),
                ("size_xl", ""),
                ("size_2xl", ""),
                ("size_3xl", ""),
                ("total_quantity", ""),
                ("material", ""),
                ("composition", ""),
                ("washing_method", ""),
                ("washing_method_en", ""),
                ("safety_category", ""),
                ("standard_code", ""),
                ("detection_report", ""),
                ("shipping_warehouse", ""),
                ("image_gallery_existing__0", gallery[1]),
                ("image_gallery_manual__0", gallery[1]),
                ("image_gallery_existing__1", gallery[0]),
                ("image_gallery_manual__1", gallery[0]),
            ],
            files=[],
        )
        reorder_response = self.request(
            f"/products/{created_product['id']}/edit",
            method="POST",
            body=reorder_body,
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=cookie,
        )
        self.assertTrue(reorder_response["status"].startswith("302"))

        updated_product = db.get_product(self.db_path, created_product["id"])
        updated_gallery = json.loads(updated_product["image_gallery_json"])
        self.assertEqual(updated_gallery[0], gallery[1])
        self.assertEqual(updated_product["image_url"], gallery[1])

        api_response = self.request("/api/products", cookie=cookie)
        payload = json.loads(api_response["body"].decode("utf-8"))
        target_item = next(item for item in payload["items"] if item["style_code"] == "PIC-MULTI-001")
        self.assertEqual(target_item["image_gallery"][0], gallery[1])
        self.assertEqual(len(target_item["image_gallery"]), 2)

    def test_b_editor_can_import_images_with_mapping_workbook(self):
        cookie = self.login("b_editor", "demo123")
        png_bytes = b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Wn3FoAAAAAASUVORK5CYII="
        )
        workbook_bytes = self.make_image_mapping_workbook_bytes(
            [("针织开衫-米白", "mapped-image.png")]
        )
        response = self.request(
            "/import-images",
            method="POST",
            body=self.build_multi_multipart(
                files=[
                    ("mapping_workbook", "image-map.xlsx", workbook_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("image_files", "mapped-image.png", png_bytes, "image/png"),
                ],
            ),
            content_type='multipart/form-data; boundary=----WebKitFormBoundaryCatalogTest',
            cookie=cookie,
        )
        self.assertTrue(response["status"].startswith("200"))
        body = response["body"].decode("utf-8")
        self.assertIn("Excel 映射", body)
        self.assertIn("成功更新", body)

        updated_product = db.get_product(self.db_path, 2)
        self.assertTrue(str(updated_product["image_url"]).startswith("/media/"))
        gallery = json.loads(updated_product["image_gallery_json"])
        self.assertEqual(updated_product["image_url"], gallery[0])

        media_response = self.request(updated_product["image_url"], cookie=cookie)
        self.assertTrue(media_response["status"].startswith("200"))
        headers = dict(media_response["headers"])
        self.assertEqual(headers["Content-Type"], "image/png")

    def test_product_detail_includes_quick_copy_blocks(self):
        cookie = self.login("a_editor", "demo123")
        detail_response = self.request("/products/1", cookie=cookie)
        detail_body = detail_response["body"].decode("utf-8")
        self.assertIn("快捷复制与调用片段", detail_body)
        self.assertIn("复制核心信息", detail_body)
        self.assertIn("复制 JSON", detail_body)
        self.assertIn("复制图片地址", detail_body)
        self.assertIn("商品名称: 褶皱短袖连衣裙", detail_body)
        self.assertIn('"style_code": "NH-2601"', detail_body)
        self.assertIn("历时天数", detail_body)
        self.assertIn("资料完成", detail_body)
        self.assertIn("修改版本", detail_body)
        self.assertIn("V1", detail_body)

    def test_initial_product_has_version_one_snapshot(self):
        product = db.get_product(self.db_path, 1)
        self.assertEqual(product["current_version_no"], 1)
        versions = db.list_product_versions(self.db_path, 1)
        self.assertEqual(len(versions), 1)
        self.assertEqual(versions[0]["version_no"], 1)
        self.assertEqual(versions[0]["change_count"], 0)

    def test_elapsed_days_is_calculated_from_first_create_time(self):
        with db.get_connection(self.db_path) as connection:
            connection.execute(
                "UPDATE products SET created_at = ?, updated_at = ?, completed_to_c_at = ? WHERE id = 1",
                ("2026-06-10T00:00:00Z", "2026-06-10T00:00:00Z", "2026-06-12T00:00:00Z"),
            )
        expected_days = 3
        cookie = self.login("a_editor", "demo123")
        list_response = self.request("/products", cookie=cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn(f'title="{expected_days} 天">{expected_days} 天</span>', list_body)

        detail_response = self.request("/products/1", cookie=cookie)
        detail_body = detail_response["body"].decode("utf-8")
        self.assertIn(f"历时天数</span><strong>{expected_days} 天</strong>", detail_body)

        api_response = self.request("/api/products", cookie=cookie)
        payload = json.loads(api_response["body"].decode("utf-8"))
        target_item = next(item for item in payload["items"] if item["id"] == 1)
        self.assertEqual(target_item["elapsed_days"], expected_days)
        self.assertEqual(target_item["elapsed_days_label"], f"{expected_days} 天")

    def test_completion_flag_is_auto_calculated_and_exported(self):
        cookie = self.login("a_editor", "demo123")
        api_response = self.request("/api/products", cookie=cookie)
        payload = json.loads(api_response["body"].decode("utf-8"))
        first_item = next(item for item in payload["items"] if item["id"] == 1)
        self.assertEqual(first_item["completion_flag"], "Y")

        second_item = next(item for item in payload["items"] if item["id"] == 2)
        self.assertEqual(second_item["completion_flag"], "")

        export_response = self.request("/export.xlsx", cookie=cookie)
        workbook = load_workbook(io.BytesIO(export_response["body"]))
        headers = [cell.value for cell in workbook.active[1]]
        product_name_col = headers.index("商品名称") + 1
        completion_col = headers.index("资料完成") + 1
        target_row = None
        for row_index in range(2, workbook.active.max_row + 1):
            if workbook.active.cell(row_index, product_name_col).value == "褶皱短袖连衣裙":
                target_row = row_index
                break
        self.assertIsNotNone(target_row)
        self.assertEqual(workbook.active.cell(target_row, completion_col).value, "Y")

    def test_c_product_detail_quick_json_respects_visible_fields(self):
        cookie = self.login("c_viewer", "demo123")
        detail_response = self.request("/products/1", cookie=cookie)
        detail_body = detail_response["body"].decode("utf-8")
        self.assertIn("复制 JSON", detail_body)
        self.assertIn('"product_name": "褶皱短袖连衣裙"', detail_body)
        self.assertNotIn('"launch_price":', detail_body)
        self.assertNotIn('"supplier":', detail_body)

    def test_admin_can_configure_c_visible_fields(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        settings_page = self.request("/settings/c-fields", cookie=admin_cookie)
        settings_body = settings_page["body"].decode("utf-8")
        self.assertIn("C 部门 API 令牌", settings_body)

        update_response = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("field_keys__style_code", "style_code"),
                    ("field_keys__launch_price", "launch_price"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(update_response["status"].startswith("302"))
        self.assertEqual(
            db.get_setting(self.db_path, "c_visible_field_keys"),
            "product_name,style_code,launch_price",
        )

        c_cookie = self.login("c_viewer", "demo123")
        list_response = self.request("/products", cookie=c_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn("<th>上新价格</th>", list_body)
        self.assertNotIn("<th>品牌</th>", list_body)

        api_response = self.request("/api/products", cookie=c_cookie)
        payload = json.loads(api_response["body"].decode("utf-8"))
        first_item = payload["items"][0]
        self.assertIn("launch_price", first_item)
        self.assertNotIn("brand_name", first_item)

    def test_c_field_settings_sync_with_current_a_list_fields(self):
        db.set_setting(
            self.db_path,
            "list_layout_fields_A",
            "shooting_date,style_code,material",
        )
        db.set_setting(self.db_path, "list_layout_customized_A", "1")
        db.set_setting(
            self.db_path,
            "c_visible_field_keys",
            "shooting_date,style_code,size_chart,size_f,total_quantity",
        )

        self.assertEqual(
            self.app.configured_c_field_keys(),
            ["shooting_date", "style_code"],
        )

        admin_cookie = self.login("admin_reviewer", "demo123")
        settings_page = self.request("/settings/c-fields", cookie=admin_cookie)
        settings_body = settings_page["body"].decode("utf-8")

        self.assertIn('name="field_keys__shooting_date"', settings_body)
        self.assertIn('name="field_keys__style_code"', settings_body)
        self.assertIn('name="field_keys__material"', settings_body)
        self.assertIn('name="field_keys__launch_price"', settings_body)
        self.assertIn('name="field_keys__launch_channel"', settings_body)
        self.assertIn('name="field_keys__completion_flag"', settings_body)
        self.assertNotIn('name="field_keys__supplier"', settings_body)
        self.assertNotIn('name="field_keys__size_chart"', settings_body)
        self.assertNotIn('name="field_keys__size_f"', settings_body)
        self.assertNotIn('name="field_keys__total_quantity"', settings_body)
        self.assertEqual(settings_body.count('class="panel c-field-access-card"'), 1)
        self.assertGreater(settings_body.count('class="c-field-group"'), 1)

    def test_admin_can_save_apply_and_delete_c_field_template(self):
        admin_cookie = self.login("admin_reviewer", "demo123")

        save_template_response = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("field_keys__style_code", "style_code"),
                    ("field_keys__launch_price", "launch_price"),
                    ("template_name", "价格视图"),
                    ("save_template", "1"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(save_template_response["status"].startswith("302"))
        self.assertEqual(
            db.get_setting(self.db_path, "c_visible_field_keys"),
            "product_name,style_code,launch_price",
        )
        saved_templates = json.loads(db.get_setting(self.db_path, "c_field_templates_json"))
        self.assertEqual(
            saved_templates["价格视图"]["field_keys"],
            ["product_name", "style_code", "launch_price"],
        )

        self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("field_keys__brand_name", "brand_name"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertEqual(
            db.get_setting(self.db_path, "c_visible_field_keys"),
            "product_name,brand_name",
        )

        apply_template_response = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                {
                    "template_name": "价格视图",
                    "apply_template": "1",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(apply_template_response["status"].startswith("302"))
        self.assertEqual(
            db.get_setting(self.db_path, "c_visible_field_keys"),
            "product_name,style_code,launch_price",
        )

        c_cookie = self.login("c_viewer", "demo123")
        list_response = self.request("/products", cookie=c_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn("<th>上新价格</th>", list_body)
        self.assertNotIn("<th>品牌</th>", list_body)

        api_response = self.request("/api/products", cookie=c_cookie)
        payload = json.loads(api_response["body"].decode("utf-8"))
        first_item = payload["items"][0]
        self.assertIn("launch_price", first_item)
        self.assertNotIn("brand_name", first_item)

        delete_template_response = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                {
                    "template_name": "价格视图",
                    "delete_template": "1",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(delete_template_response["status"].startswith("302"))
        deleted_templates = json.loads(db.get_setting(self.db_path, "c_field_templates_json"))
        self.assertNotIn("价格视图", deleted_templates)

    def test_c_field_templates_apply_same_column_scope_to_all_published_products(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        save_template_response = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("field_keys__style_code", "style_code"),
                    ("field_keys__brand_name", "brand_name"),
                    ("field_keys__launch_price", "launch_price"),
                    ("template_name", "统一运营视图"),
                    ("save_template", "1"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(save_template_response["status"].startswith("302"))
        saved_templates = json.loads(db.get_setting(self.db_path, "c_field_templates_json"))
        self.assertEqual(
            saved_templates["统一运营视图"]["field_keys"],
            ["product_name", "style_code", "brand_name", "launch_price"],
        )

        apply_template_response = self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                {
                    "template_name": "统一运营视图",
                    "apply_template": "1",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(apply_template_response["status"].startswith("302"))

        self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("field_keys__style_code", "style_code"),
                    ("field_keys__brand_name", "brand_name"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )

        self.make_product_two_a_complete()
        c_cookie = self.login("c_viewer", "demo123")
        self.request(
            "/products/2/status",
            method="POST",
            body=urlencode({"status": "published", "review_note": "用于自动模板命中测试"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        api_response_b = self.request("/api/products", cookie=c_cookie)
        payload_b = json.loads(api_response_b["body"].decode("utf-8"))
        a_item = next(item for item in payload_b["items"] if item["id"] == 1)
        b_item = next(item for item in payload_b["items"] if item["id"] == 2)
        self.assertNotIn("launch_price", a_item)
        self.assertNotIn("launch_price", b_item)

    def test_b_stage_update_logs_field_level_diff(self):
        with db.get_connection(self.db_path) as connection:
            db.change_product_status(
                connection,
                1,
                "pending",
                1,
                "提交给B填写",
                "测试：A 已提交给 B。",
            )
        cookie = self.login("b_editor", "demo123")
        update_response = self.request(
            "/products/1/edit",
            method="POST",
            body=urlencode(
                {
                    "brand_name": "North Harbor",
                    "product_name": "褶皱短袖连衣裙",
                    "style_code": "NH-2601",
                    "launch_price": "359",
                    "launch_channel": "电商首发",
                }
            ).encode("utf-8"),
            cookie=cookie,
        )
        self.assertTrue(update_response["status"].startswith("302"))

        logs = db.get_product_logs(self.db_path, 1)
        update_log = next(item for item in logs if item["action"] == "update")
        diff_items = update_log["diff_items"]
        self.assertTrue(any(item["field_key"] == "launch_price" for item in diff_items))
        self.assertTrue(any(item["field_key"] == "launch_channel" for item in diff_items))
        self.assertLessEqual(len(diff_items), 3)

        logs_response = self.request("/products/1/logs", cookie=cookie)
        logs_body = logs_response["body"].decode("utf-8")
        self.assertIn("本次共修改 2 项，仅摘要展示 2 项", logs_body)
        self.assertIn("上新价格", logs_body)

        versions = db.list_product_versions(self.db_path, 1)
        self.assertEqual(versions[0]["version_no"], 2)
        self.assertEqual(versions[0]["change_count"], 2)

    def test_admin_can_view_versions_and_restore_old_version(self):
        with db.get_connection(self.db_path) as connection:
            db.change_product_status(
                connection,
                1,
                "pending",
                1,
                "提交给B填写",
                "测试版本恢复：先流转给商品部。",
            )
            db.update_product(
                connection,
                1,
                {
                    "brand_name": "North Harbor",
                    "product_name": "褶皱短袖连衣裙改版",
                    "style_code": "NH-2601",
                    "launch_price": "399",
                    "launch_channel": "直播首发",
                },
                2,
            )

        admin_cookie = self.login("admin_reviewer", "demo123")
        versions_response = self.request("/products/1/versions", cookie=admin_cookie)
        versions_body = versions_response["body"].decode("utf-8")
        self.assertIn("修改版本", versions_body)
        self.assertIn("V2", versions_body)
        self.assertIn("恢复为当前版本", versions_body)
        self.assertIn("当前版本", versions_body)
        self.assertIn("初始版本", versions_body)

        restore_response = self.request(
            "/products/1/versions/1/restore",
            method="POST",
            body=b"",
            cookie=admin_cookie,
        )
        self.assertTrue(restore_response["status"].startswith("302"))
        restored_product = db.get_product(self.db_path, 1)
        self.assertEqual(restored_product["current_version_no"], 3)
        self.assertEqual(restored_product["product_name"], "褶皱短袖连衣裙")

        logs = db.get_product_logs(self.db_path, 1)
        restore_log = next(item for item in logs if item["action"] == "version_restore")
        self.assertIn("恢复到 V1", restore_log["details"])

        restored_versions = db.list_product_versions(self.db_path, 1)
        self.assertEqual(restored_versions[0]["version_no"], 3)
        self.assertEqual(restored_versions[0]["source_version_no"], 1)
        self.assertGreaterEqual(restored_versions[0]["change_count"], 1)

        refreshed_versions_response = self.request("/products/1/versions", cookie=admin_cookie)
        refreshed_versions_body = refreshed_versions_response["body"].decode("utf-8")
        self.assertIn("恢复自 V1", refreshed_versions_body)
        self.assertIn("修改项数量", refreshed_versions_body)

    def test_product_logs_support_filtering_and_csv_export(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        self.make_product_two_a_complete()
        publish_response = self.request(
            "/products/2/status",
            method="POST",
            body=urlencode({"status": "published", "review_note": "审核通过用于日志导出测试"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(publish_response["status"].startswith("302"))

        logs_response = self.request("/products/2/logs?action=%E7%AE%A1%E7%90%86%E5%91%98%E4%BB%A3%E4%B8%BA%E5%AE%8C%E6%88%90", cookie=admin_cookie)
        logs_body = logs_response["body"].decode("utf-8")
        self.assertIn("导出 CSV", logs_body)
        self.assertIn("管理员代为完成", logs_body)
        self.assertNotIn("<td>提交给B填写</td>", logs_body)

        export_response = self.request(
            "/products/2/logs/export.csv?action=%E7%AE%A1%E7%90%86%E5%91%98%E4%BB%A3%E4%B8%BA%E5%AE%8C%E6%88%90",
            cookie=admin_cookie,
        )
        self.assertTrue(export_response["status"].startswith("200"))
        export_headers = dict(export_response["headers"])
        self.assertEqual(export_headers["Content-Type"], "text/csv; charset=utf-8")
        csv_body = export_response["body"].decode("utf-8-sig")
        self.assertIn("动作,说明", csv_body)
        self.assertIn("管理员代为完成", csv_body)
        self.assertNotIn("提交给B填写", csv_body)

    def test_product_logs_can_filter_by_actor_keyword(self):
        editor_cookie = self.login("a_editor", "demo123")
        detail_response = self.request("/products/1/logs?actor=%E8%B7%9F%E5%8D%95%E9%83%A8", cookie=editor_cookie)
        detail_body = detail_response["body"].decode("utf-8")
        self.assertIn("跟单部录入员", detail_body)
        self.assertIn("创建资料", detail_body)

        export_response = self.request("/products/1/logs/export.csv?actor=%E8%B7%9F%E5%8D%95%E9%83%A8", cookie=editor_cookie)
        csv_body = export_response["body"].decode("utf-8-sig")
        self.assertIn("跟单部录入员", csv_body)
        self.assertNotIn("系统管理员", csv_body)

    def test_global_logs_center_respects_role_scope(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        admin_response = self.request("/logs", cookie=admin_cookie)
        admin_body = admin_response["body"].decode("utf-8")
        self.assertIn("全局操作日志中心", admin_body)
        self.assertIn("褶皱短袖连衣裙", admin_body)
        self.assertIn("毛感针织开衫", admin_body)

        editor_cookie = self.login("a_editor", "demo123")
        editor_response = self.request("/logs", cookie=editor_cookie)
        editor_body = editor_response["body"].decode("utf-8")
        self.assertIn("褶皱短袖连衣裙", editor_body)
        self.assertIn("毛感针织开衫", editor_body)

        c_cookie = self.login("c_viewer", "demo123")
        forbidden_response = self.request("/logs", cookie=c_cookie)
        self.assertTrue(forbidden_response["status"].startswith("403"))

    def test_global_logs_center_supports_filter_and_export(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        self.make_product_two_a_complete()
        publish_response = self.request(
            "/products/2/status",
            method="POST",
            body=urlencode({"status": "published", "review_note": "全局日志中心测试"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(publish_response["status"].startswith("302"))

        filtered_response = self.request("/logs?action=%E7%AE%A1%E7%90%86%E5%91%98%E4%BB%A3%E4%B8%BA%E5%AE%8C%E6%88%90", cookie=admin_cookie)
        filtered_body = filtered_response["body"].decode("utf-8")
        self.assertIn("管理员代为完成", filtered_body)
        self.assertNotIn("<td>提交给B填写</td>", filtered_body)

        export_response = self.request("/logs/export.csv?action=%E7%AE%A1%E7%90%86%E5%91%98%E4%BB%A3%E4%B8%BA%E5%AE%8C%E6%88%90", cookie=admin_cookie)
        self.assertTrue(export_response["status"].startswith("200"))
        export_headers = dict(export_response["headers"])
        self.assertEqual(export_headers["Content-Type"], "text/csv; charset=utf-8")
        csv_body = export_response["body"].decode("utf-8-sig")
        self.assertIn("资料ID,商品名称,款号", csv_body)
        self.assertIn("管理员代为完成", csv_body)
        self.assertNotIn("提交给B填写", csv_body)

    def test_admin_actions_are_included_in_global_audit_center(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("rotate_token", "1"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.request(
            "/settings/c-fields",
            method="POST",
            body=urlencode(
                [
                    ("field_keys__product_name", "product_name"),
                    ("confirm_text", "DISABLE"),
                    ("disable_token", "1"),
                ]
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        create_response = self.request(
            "/users",
            method="POST",
            body=urlencode(
                {
                    "username": "audit_user",
                    "display_name": "审计测试账号",
                    "department": "A",
                    "password": "start123",
                    "must_change_password": "on",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))
        new_user = db.list_users(self.db_path)[-1]
        self.request(
            f"/users/{new_user['id']}/reset-password",
            method="POST",
            body=urlencode({"new_password": "fresh456", "confirm_text": "RESET"}).encode("utf-8"),
            cookie=admin_cookie,
        )

        logs_response = self.request("/logs?action=%E9%87%8D%E7%BD%AE%E5%AF%86%E7%A0%81", cookie=admin_cookie)
        logs_body = logs_response["body"].decode("utf-8")
        self.assertIn("重置密码", logs_body)
        self.assertIn("audit_user", logs_body)
        self.assertIn("管理审计", logs_body)

        export_response = self.request("/logs/export.csv?action=%E9%87%8D%E7%BD%AE%E5%AF%86%E7%A0%81", cookie=admin_cookie)
        export_body = export_response["body"].decode("utf-8-sig")
        self.assertIn("重置密码", export_body)
        self.assertIn("audit_user", export_body)

    def test_admin_can_bulk_publish_and_archive_from_products_list(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        self.make_product_two_a_complete()
        list_response = self.request("/products", cookie=admin_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn("批量操作", list_body)
        self.assertIn('name="product_ids"', list_body)

        publish_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "2"), ("bulk_action", "publish_selected")]).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(publish_response["status"].startswith("302"))
        published_product = db.get_product(self.db_path, 2)
        self.assertEqual(published_product["status"], "published")

        archive_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "1"), ("bulk_action", "archive_selected")]).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(archive_response["status"].startswith("302"))
        archived_product = db.get_product(self.db_path, 1)
        self.assertEqual(archived_product["lifecycle_status"], "archived")

    def test_a_editor_can_bulk_submit_owned_drafts_to_b(self):
        editor_cookie = self.login("a_editor", "demo123")
        create_response = self.request(
            "/products/new",
            method="POST",
            body=urlencode(
                {
                    "product_name": "跟单部批量提交流转测试",
                    "style_code": "A-BULK-001",
                    "style_color": "测试款色",
                    "color_name": "测试颜色",
                    "brand_name": "思安娜",
                    "category": "上衣",
                }
            ).encode("utf-8"),
            cookie=editor_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))
        list_response = self.request("/products", cookie=editor_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn("批量流转", list_body)
        self.assertIn("批量提交给商品部填写", list_body)
        self.assertIn('name="product_ids"', list_body)

        created_product = next(item for item in db.list_products(self.db_path) if item["style_code"] == "A-BULK-001")
        submit_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", str(created_product["id"])), ("bulk_action", "submit_to_b_selected")]).encode("utf-8"),
            cookie=editor_cookie,
        )
        self.assertTrue(submit_response["status"].startswith("302"))
        product = db.get_product(self.db_path, created_product["id"])
        self.assertEqual(product["status"], "draft")

    def test_a_editor_cannot_submit_to_b_until_a_required_fields_are_complete(self):
        editor_cookie = self.login("a_editor", "demo123")
        create_response = self.request(
            "/products/new",
            method="POST",
            body=urlencode(
                {
                    "product_name": "待完善资料",
                    "style_code": "A-CHECK-001",
                    "brand_name": "Check Brand",
                    "category": "上衣",
                }
            ).encode("utf-8"),
            cookie=editor_cookie,
        )
        self.assertTrue(create_response["status"].startswith("302"))
        created_product = next(item for item in db.list_products(self.db_path) if item["style_code"] == "A-CHECK-001")

        submit_response = self.request(
            f"/products/{created_product['id']}/status",
            method="POST",
            body=urlencode({"status": "pending"}).encode("utf-8"),
            cookie=editor_cookie,
        )
        self.assertTrue(submit_response["status"].startswith("400"))
        body = submit_response["body"].decode("utf-8")
        self.assertIn("暂不能提交给商品部", body)
        product = db.get_product(self.db_path, created_product["id"])
        self.assertEqual(product["status"], "draft")

    def test_a_editor_bulk_submit_skips_non_owned_or_ineligible_products(self):
        editor_cookie = self.login("a_editor", "demo123")
        submit_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "2"), ("bulk_action", "submit_to_b_selected")]).encode("utf-8"),
            cookie=editor_cookie,
        )
        self.assertTrue(submit_response["status"].startswith("302"))
        headers = dict(submit_response["headers"])
        notice = unquote_plus(headers["Location"])
        self.assertIn("成功 0 条", notice)
        self.assertIn("跳过 1 条", notice)
        product = db.get_product(self.db_path, 2)
        self.assertEqual(product["status"], "pending")

    def test_b_editor_can_bulk_complete_pending_products_to_c(self):
        b_cookie = self.login("b_editor", "demo123")
        list_response = self.request("/products", cookie=b_cookie)
        list_body = list_response["body"].decode("utf-8")
        self.assertIn("批量完成", list_body)
        self.assertIn("批量完成并开放给运营部", list_body)
        self.assertIn("批量退回跟单部修改", list_body)
        self.assertIn('name="product_ids"', list_body)

        complete_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "2"), ("bulk_action", "complete_to_c_selected")]).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(complete_response["status"].startswith("302"))
        product = db.get_product(self.db_path, 2)
        self.assertEqual(product["status"], "pending")

    def test_b_editor_can_bulk_return_pending_products_to_a(self):
        b_cookie = self.login("b_editor", "demo123")
        return_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "2"), ("bulk_action", "return_to_a_selected")]).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(return_response["status"].startswith("302"))
        notice = unquote_plus(dict(return_response["headers"])["Location"])
        self.assertIn("批量退回跟单部修改完成：成功 1 条", notice)
        product = db.get_product(self.db_path, 2)
        self.assertEqual(product["status"], "draft")

    def test_b_editor_bulk_return_skips_non_pending_products(self):
        b_cookie = self.login("b_editor", "demo123")
        return_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "1"), ("bulk_action", "return_to_a_selected")]).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(return_response["status"].startswith("302"))
        notice = unquote_plus(dict(return_response["headers"])["Location"])
        self.assertIn("成功 0 条", notice)
        self.assertIn("跳过 1 条", notice)
        self.assertIn("不能退回给跟单部", notice)
        product = db.get_product(self.db_path, 1)
        self.assertEqual(product["status"], "published")

    def test_b_editor_cannot_publish_without_launch_fields(self):
        with db.get_connection(self.db_path) as connection:
            connection.execute(
                "UPDATE products SET launch_price = NULL, launch_channel = NULL WHERE id = 2"
            )
        b_cookie = self.login("b_editor", "demo123")
        response = self.request(
            "/products/2/status",
            method="POST",
            body=urlencode({"status": "published"}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(response["status"].startswith("400"))
        body = response["body"].decode("utf-8")
        self.assertIn("暂不能流转给运营部", body)
        product = db.get_product(self.db_path, 2)
        self.assertEqual(product["status"], "pending")

    def test_b_editor_bulk_complete_skips_non_pending_products(self):
        b_cookie = self.login("b_editor", "demo123")
        complete_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "1"), ("bulk_action", "complete_to_c_selected")]).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(complete_response["status"].startswith("302"))
        headers = dict(complete_response["headers"])
        notice = unquote_plus(headers["Location"])
        self.assertIn("成功 0 条", notice)
        self.assertIn("跳过 1 条", notice)
        product = db.get_product(self.db_path, 1)
        self.assertEqual(product["status"], "published")

    def test_b_editor_bulk_complete_notice_includes_validation_reason(self):
        self.make_product_two_a_complete()
        with db.get_connection(self.db_path) as connection:
            connection.execute(
                "UPDATE products SET completion_flag = 'Y', inspection_date = '' WHERE id = 2"
            )
        b_cookie = self.login("b_editor", "demo123")
        complete_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "2"), ("bulk_action", "complete_to_c_selected")]).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(complete_response["status"].startswith("302"))
        notice = unquote_plus(dict(complete_response["headers"])["Location"])
        self.assertIn("跳过 1 条", notice)
        self.assertIn("送检时间", notice)
        product = db.get_product(self.db_path, 2)
        self.assertEqual(product["status"], "pending")

    def test_a_editor_can_edit_after_submit_and_marks_revision_for_b_and_c(self):
        a_cookie = self.login("a_editor", "demo123")
        submit_response = self.request(
            "/products/2/edit",
            method="POST",
            body=urlencode(
                self.a_complete_fields_payload(
                    brand_name="Studio Pine",
                    season_year="2026秋",
                    image_url="https://example.com/images/sp-8420.jpg",
                    style_color="针织开衫-米白",
                    style_code="SP-8420",
                    color_name="燕麦白",
                    product_name="毛感针织开衫（已更新）",
                    category="针织衫",
                    has_accessories="有",
                    supplier="嘉兴尚品针织",
                    cooperation_mode="联营",
                    supply_chain_manager="周岚",
                    tax_included_price="188",
                    tag_price="399",
                    size_range="F",
                    size_f="48",
                    material="针织",
                    composition_en="SHELL: 46% ACRYLIC 30% POLYESTER 24% NYLON",
                    washing_method="建议平铺晾干",
                    washing_method_en="Dry flat",
                    safety_category="B类",
                    standard_code="FZ/T 73018",
                    detection_report="已归档",
                    shipping_warehouse="嘉兴二仓",
                )
            ).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(submit_response["status"].startswith("302"))
        updated_product = db.get_product(self.db_path, 2)
        self.assertEqual(updated_product["status"], "pending")
        self.assertEqual(updated_product["revision_flag"], 1)
        self.assertEqual(updated_product["product_name"], "毛感针织开衫（已更新）")

        b_cookie = self.login("b_editor", "demo123")
        b_list_response = self.request("/products", cookie=b_cookie)
        b_list_body = b_list_response["body"].decode("utf-8")
        self.assertIn("已更新", b_list_body)
        self.assertIn("毛感针织开衫（已更新）", b_list_body)

        admin_cookie = self.login("admin_reviewer", "demo123")
        publish_response = self.request(
            "/products/2/status",
            method="POST",
            body=urlencode({"status": "published", "review_note": "用于验证已更新标记"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(publish_response["status"].startswith("302"))

        c_cookie = self.login("c_viewer", "demo123")
        c_list_response = self.request("/products", cookie=c_cookie)
        c_list_body = c_list_response["body"].decode("utf-8")
        self.assertIn("已更新", c_list_body)
        self.assertIn("毛感针织开衫（已更新）", c_list_body)

    def test_bulk_actions_skip_ineligible_products_and_forbid_non_admin(self):
        editor_cookie = self.login("a_editor", "demo123")
        forbidden_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "1"), ("bulk_action", "archive_selected")]).encode("utf-8"),
            cookie=editor_cookie,
        )
        self.assertTrue(forbidden_response["status"].startswith("302"))
        headers = dict(forbidden_response["headers"])
        notice = unquote_plus(headers["Location"])
        self.assertIn("成功 1 条", notice)
        self.assertIn("跳过 0 条", notice)

        admin_cookie = self.login("admin_reviewer", "demo123")
        skip_response = self.request(
            "/products/bulk",
            method="POST",
            body=urlencode([("product_ids", "1"), ("bulk_action", "publish_selected")]).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(skip_response["status"].startswith("302"))
        headers = dict(skip_response["headers"])
        self.assertIn("0+", headers["Location"])
        self.assertIn("1+", headers["Location"])

    def test_editor_can_archive_own_product_and_admin_can_restore_deleted_product(self):
        editor_cookie = self.login("a_editor", "demo123")
        archive_response = self.request(
            "/products/1/lifecycle",
            method="POST",
            body=urlencode({"lifecycle_status": "archived"}).encode("utf-8"),
            cookie=editor_cookie,
        )
        self.assertTrue(archive_response["status"].startswith("302"))
        product = db.get_product(self.db_path, 1)
        self.assertEqual(product["lifecycle_status"], "archived")

        logs = db.get_product_logs(self.db_path, 1)
        self.assertTrue(any(item["action"] == "lifecycle:archived" for item in logs))

        c_cookie = self.login("c_viewer", "demo123")
        forbidden_response = self.request("/products/1", cookie=c_cookie)
        self.assertTrue(forbidden_response["status"].startswith("403"))

        admin_cookie = self.login("admin_reviewer", "demo123")
        delete_response = self.request(
            "/products/1/lifecycle",
            method="POST",
            body=urlencode({"lifecycle_status": "deleted", "confirm_text": "DELETE"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(delete_response["status"].startswith("302"))
        deleted_product = db.get_product(self.db_path, 1)
        self.assertEqual(deleted_product["lifecycle_status"], "deleted")

        restore_response = self.request(
            "/products/1/lifecycle",
            method="POST",
            body=urlencode({"lifecycle_status": "active"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(restore_response["status"].startswith("302"))
        restored_product = db.get_product(self.db_path, 1)
        self.assertEqual(restored_product["lifecycle_status"], "active")

    def test_delete_lifecycle_requires_confirmation(self):
        admin_cookie = self.login("admin_reviewer", "demo123")
        delete_response = self.request(
            "/products/1/lifecycle",
            method="POST",
            body=urlencode({"lifecycle_status": "deleted"}).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(delete_response["status"].startswith("400"))
        product = db.get_product(self.db_path, 1)
        self.assertEqual(product["lifecycle_status"], "active")

    def test_only_a_creator_or_admin_can_delete_product_records(self):
        b_cookie = self.login("b_editor", "demo123")
        b_list_response = self.request("/products", cookie=b_cookie)
        self.assertNotIn(b'data-delete-button="1"', b_list_response["body"])

        forbidden_response = self.request(
            "/products/2/lifecycle",
            method="POST",
            body=urlencode({"lifecycle_status": "deleted", "confirm_text": "DELETE"}).encode("utf-8"),
            cookie=b_cookie,
        )
        self.assertTrue(forbidden_response["status"].startswith("403"))
        self.assertEqual(db.get_product(self.db_path, 2)["lifecycle_status"], "active")

        a_cookie = self.login("a_editor", "demo123")
        a_list_response = self.request("/products", cookie=a_cookie)
        a_list_body = a_list_response["body"].decode("utf-8")
        self.assertLess(a_list_body.index('data-delete-button="1"'), a_list_body.index('/products/1/logs'))
        delete_response = self.request(
            "/products/2/lifecycle",
            method="POST",
            body=urlencode({"lifecycle_status": "deleted", "confirm_text": "DELETE"}).encode("utf-8"),
            cookie=a_cookie,
        )
        self.assertTrue(delete_response["status"].startswith("302"))
        self.assertEqual(db.get_product(self.db_path, 2)["lifecycle_status"], "deleted")

    def build_multipart(
        self,
        field_name: str,
        filename: str,
        file_bytes: bytes,
        extra_fields: dict | None = None,
        file_content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ) -> bytes:
        boundary = "----WebKitFormBoundaryCatalogTest"
        parts = []
        for key, value in (extra_fields or {}).items():
            parts.extend(
                [
                    f"--{boundary}\r\n".encode("utf-8"),
                    f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode("utf-8"),
                    str(value).encode("utf-8"),
                    b"\r\n",
                ]
            )
        parts.extend(
            [
                f"--{boundary}\r\n".encode("utf-8"),
                (
                    f'Content-Disposition: form-data; name="{field_name}"; filename="{filename}"\r\n'
                    f"Content-Type: {file_content_type}\r\n\r\n"
                ).encode("utf-8"),
                file_bytes,
                b"\r\n",
                f"--{boundary}--\r\n".encode("utf-8"),
            ]
        )
        return b"".join(parts)

    def build_multi_multipart(
        self,
        fields: list[tuple[str, str]] | None = None,
        files: list[tuple[str, str, bytes, str]] | None = None,
    ) -> bytes:
        boundary = "----WebKitFormBoundaryCatalogTest"
        parts = []
        for key, value in (fields or []):
            parts.extend(
                [
                    f"--{boundary}\r\n".encode("utf-8"),
                    f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode("utf-8"),
                    str(value).encode("utf-8"),
                    b"\r\n",
                ]
            )
        for field_name, filename, file_bytes, file_content_type in (files or []):
            parts.extend(
                [
                    f"--{boundary}\r\n".encode("utf-8"),
                    (
                        f'Content-Disposition: form-data; name="{field_name}"; filename="{filename}"\r\n'
                        f"Content-Type: {file_content_type}\r\n\r\n"
                    ).encode("utf-8"),
                    file_bytes,
                    b"\r\n",
                ]
            )
        parts.append(f"--{boundary}--\r\n".encode("utf-8"))
        return b"".join(parts)


if __name__ == "__main__":
    unittest.main()
