from __future__ import annotations

import io
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch
from urllib.parse import urlencode
from wsgiref.util import setup_testing_defaults

from openpyxl import load_workbook

from catalog_backend import CatalogApplication, init_db as init_catalog_db
from catalog_backend import db as catalog_db
from catalog_backend.policies import editable_field_keys_for_user
from planning_center import PlanningApplication
from planning_center import db as planning_db


class PlanningCenterTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        root = Path(self.temp.name)
        self.catalog_db_path = root / "catalog.db"
        self.planning_db_path = root / "planning.db"
        init_catalog_db(self.catalog_db_path)
        planning_db.init_db(self.planning_db_path)

    def tearDown(self):
        self.temp.cleanup()

    def wsgi_request(self, app, path, method="GET", body=b"", content_type="application/x-www-form-urlencoded", cookie="", authorization=""):
        environ = {}
        setup_testing_defaults(environ)
        if "?" in path:
            path, environ["QUERY_STRING"] = path.split("?", 1)
        environ.update({"PATH_INFO": path, "REQUEST_METHOD": method, "CONTENT_LENGTH": str(len(body)), "CONTENT_TYPE": content_type, "wsgi.input": io.BytesIO(body)})
        if cookie:
            environ["HTTP_COOKIE"] = cookie
        if authorization:
            environ["HTTP_AUTHORIZATION"] = authorization
        captured = {}

        def start_response(status, headers):
            captured["status"] = status
            captured["headers"] = headers

        captured["body"] = b"".join(app(environ, start_response))
        return captured

    def login_cookie(self, app, username):
        response = self.wsgi_request(app, "/login", method="POST", body=urlencode({"username": username, "password": "demo123"}).encode())
        return dict(response["headers"])["Set-Cookie"].split(";", 1)[0]

    def workbook_multipart(self, workbook_bytes: bytes, filename: str = "initial-review.xlsx") -> tuple[bytes, str]:
        boundary = "----PlanningCenterExcelTest"
        body = (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="workbook"; filename="{filename}"\r\n'
            "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        ).encode("utf-8") + workbook_bytes + f"\r\n--{boundary}--\r\n".encode("utf-8")
        return body, f"multipart/form-data; boundary={boundary}"

    def test_internal_api_requires_token_and_publishes_with_version_check(self):
        app = CatalogApplication(self.catalog_db_path, Path(self.temp.name) / "uploads", planning_api_token="planning-secret")
        pending_product = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "pending"
        )
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            connection.execute(
                "UPDATE products SET tax_included_price = 150 WHERE id = ?",
                (pending_product["id"],),
            )
        response = self.wsgi_request(app, "/api/internal/planning/products")
        self.assertTrue(response["status"].startswith("401"))
        response = self.wsgi_request(app, "/api/internal/planning/products", authorization="Bearer planning-secret")
        payload = json.loads(response["body"])
        self.assertEqual(payload["count"], 1)
        self.assertTrue(payload["workflow_gate"])
        self.assertTrue(payload["image_gate"])
        self.assertTrue(payload["cost_gate"])
        self.assertEqual(payload["eligibility_gate_version"], 1)
        self.assertEqual({item["status"] for item in payload["items"]}, {"pending"})
        self.assertEqual({item["lifecycle_status"] for item in payload["items"]}, {"active"})
        self.assertEqual({item["submitted_to_merchandise"] for item in payload["items"]}, {True})
        self.assertTrue(payload["withdrawn_ids"])
        source = payload["items"][0]
        b_user = next(user for user in catalog_db.list_users(self.catalog_db_path) if user["username"] == "b_editor")
        before_image_version = int(source["image_version_no"])
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            catalog_db.update_product(
                connection,
                source["id"],
                {
                    "image_url": "https://example.com/images/pending-professional.jpg",
                    "image_gallery_json": json.dumps(["https://example.com/images/pending-professional.jpg"]),
                },
                b_user["id"],
            )
        refreshed_source = catalog_db.get_product(self.catalog_db_path, source["id"])
        self.assertEqual(refreshed_source["current_version_no"], source["source_version_no"])
        self.assertEqual(refreshed_source["image_version_no"], before_image_version + 1)
        source = catalog_db.planning_source_payloads(self.catalog_db_path, source["id"])[0]
        publication = {
            "publication_id": "PC-TEST-001",
            "source_version_no": source["source_version_no"],
            "category": "连衣裙",
            "launch_channel": "唯品",
            "launch_price": 599,
            "fixed_multiplier": 4,
            "supplier_coefficient": 1,
            "raw_price": 600,
            "operator_name": "测试企划员",
        }
        protected_before = {
            field.key: refreshed_source.get(field.key)
            for field in catalog_db.PRODUCT_FIELDS
            if field.key not in catalog_db.PLANNING_PRODUCT_MUTABLE_FIELD_KEYS
        }
        forbidden_source_changes = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{source['id']}/price-publication",
            method="POST",
            body=json.dumps(
                dict(
                    publication,
                    publication_id="PC-TEST-FORBIDDEN-SOURCE",
                    style_code="FORGED-STYLE",
                    tax_included_price=1,
                )
            ).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertTrue(forbidden_source_changes["status"].startswith("400"))
        self.assertIn("不允许回传字段", forbidden_source_changes["body"].decode("utf-8"))
        unchanged = catalog_db.get_product(self.catalog_db_path, source["id"])
        self.assertEqual(unchanged["style_code"], source["style_code"])
        self.assertEqual(unchanged["tax_included_price"], source["tax_included_price"])
        fractional = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{source['id']}/price-publication",
            method="POST",
            body=json.dumps(dict(publication, publication_id="PC-TEST-FRACTION", launch_price=599.5)).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertTrue(fractional["status"].startswith("400"))
        self.assertIn("必须是大于 0 的整数", fractional["body"].decode("utf-8"))
        response = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{source['id']}/price-publication",
            method="POST",
            body=json.dumps(publication).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertTrue(response["status"].startswith("200"))
        published_product = catalog_db.get_product(self.catalog_db_path, source["id"])
        self.assertEqual(published_product["launch_price"], 599)
        self.assertEqual(published_product["launch_channel"], "唯品")
        self.assertEqual(published_product["category"], "连衣裙")
        protected_after = {
            field.key: published_product.get(field.key)
            for field in catalog_db.PRODUCT_FIELDS
            if field.key not in catalog_db.PLANNING_PRODUCT_MUTABLE_FIELD_KEYS
        }
        self.assertEqual(protected_after, protected_before)
        retry = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{source['id']}/price-publication",
            method="POST",
            body=json.dumps(publication).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertEqual(json.loads(retry["body"])["status"], "already_published")
        stale = dict(publication, publication_id="PC-TEST-002")
        response = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{source['id']}/price-publication",
            method="POST",
            body=json.dumps(stale).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertTrue(response["status"].startswith("409"))

        completed_product = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "published"
        )
        rejected_completed = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{completed_product['id']}/price-publication",
            method="POST",
            body=json.dumps(
                dict(
                    publication,
                    publication_id="PC-COMPLETED-001",
                    source_version_no=completed_product["current_version_no"],
                )
            ).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertTrue(rejected_completed["status"].startswith("400"))
        self.assertIn("已完成资料", rejected_completed["body"].decode("utf-8"))

    def test_internal_planning_image_api_serves_pending_product_media(self):
        upload_dir = Path(self.temp.name) / "uploads"
        upload_dir.mkdir()
        image_bytes = b"\x89PNG\r\n\x1a\nplanning-image"
        (upload_dir / "planning-test.png").write_bytes(image_bytes)
        product = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "pending"
        )
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            connection.execute(
                "UPDATE products SET image_url = ? WHERE id = ?",
                ("/media/planning-test.png", product["id"]),
            )
        app = CatalogApplication(
            self.catalog_db_path,
            upload_dir,
            planning_api_token="planning-secret",
        )

        unauthorized = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{product['id']}/image",
        )
        self.assertTrue(unauthorized["status"].startswith("401"))
        response = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{product['id']}/image",
            authorization="Bearer planning-secret",
        )

        self.assertTrue(response["status"].startswith("200"))
        self.assertEqual(response["body"], image_bytes)
        self.assertEqual(dict(response["headers"])["Content-Type"], "image/png")

    def test_internal_planning_image_api_proxies_saved_external_image(self):
        product = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "pending"
        )
        image_url = "http://image-host.test/styles/new-arrival.jpg"
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            connection.execute(
                "UPDATE products SET image_url = ? WHERE id = ?",
                (image_url, product["id"]),
            )
        app = CatalogApplication(
            self.catalog_db_path,
            Path(self.temp.name) / "uploads",
            planning_api_token="planning-secret",
        )
        image_bytes = b"external-catalog-image"

        class ExternalImageResponse(io.BytesIO):
            headers = {"Content-Type": "image/jpeg"}

            def geturl(self):
                return image_url

        with patch("catalog_backend.web.urlopen", return_value=ExternalImageResponse(image_bytes)) as mocked_urlopen:
            response = self.wsgi_request(
                app,
                f"/api/internal/planning/products/{product['id']}/image",
                authorization="Bearer planning-secret",
            )

        self.assertTrue(response["status"].startswith("200"))
        self.assertEqual(response["body"], image_bytes)
        self.assertEqual(dict(response["headers"])["Content-Type"], "image/jpeg")
        request = mocked_urlopen.call_args.args[0]
        self.assertEqual(request.full_url, image_url)
        self.assertEqual(request.get_header("Accept"), "image/*")

    def test_internal_planning_image_api_sniffs_opaque_external_jpeg(self):
        product = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "pending"
        )
        image_url = "https://image-host.test/asset?id=professional-001"
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            connection.execute(
                "UPDATE products SET image_url = ? WHERE id = ?",
                (image_url, product["id"]),
            )
        app = CatalogApplication(
            self.catalog_db_path,
            Path(self.temp.name) / "uploads",
            planning_api_token="planning-secret",
        )

        class OpaqueImageResponse(io.BytesIO):
            headers = {"Content-Type": "application/octet-stream"}

            def geturl(self):
                return image_url

        with patch(
            "catalog_backend.web.urlopen",
            return_value=OpaqueImageResponse(b"\xff\xd8\xffopaque-jpeg"),
        ):
            response = self.wsgi_request(
                app,
                f"/api/internal/planning/products/{product['id']}/image",
                authorization="Bearer planning-secret",
            )

        self.assertTrue(response["status"].startswith("200"))
        self.assertEqual(response["body"], b"\xff\xd8\xffopaque-jpeg")
        self.assertEqual(dict(response["headers"])["Content-Type"], "image/jpeg")

    def test_planning_revision_publication_updates_same_catalog_product(self):
        app = CatalogApplication(
            self.catalog_db_path,
            Path(self.temp.name) / "uploads",
            planning_api_token="planning-secret",
        )
        source = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "pending"
        )
        initial = {
            "publication_id": "PC-REVISION-INITIAL",
            "source_version_no": source["current_version_no"],
            "category": "其他",
            "launch_channel": "唯品",
            "launch_price": 599,
        }
        response = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{source['id']}/price-publication",
            method="POST",
            body=json.dumps(initial).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertTrue(response["status"].startswith("200"))
        b_user = next(user for user in catalog_db.list_users(self.catalog_db_path) if user["username"] == "b_editor")
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            catalog_db.change_product_status(
                connection,
                source["id"],
                "published",
                b_user["id"],
                "填写完成，开放给运营部",
                "测试同款修订回传。",
            )
        completed = catalog_db.get_product(self.catalog_db_path, source["id"])
        revision = dict(
            initial,
            publication_id="PC-REVISION-SECOND",
            source_version_no=completed["current_version_no"],
            category="连衣裙",
            launch_channel="天猫",
            launch_price=629,
            revision=True,
        )
        response = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{source['id']}/price-publication",
            method="POST",
            body=json.dumps(revision).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertTrue(response["status"].startswith("200"))
        updated = catalog_db.get_product(self.catalog_db_path, source["id"])
        self.assertEqual(updated["status"], "published")
        self.assertEqual(updated["current_version_no"], completed["current_version_no"] + 1)
        self.assertEqual(updated["category"], "连衣裙")
        self.assertEqual(updated["launch_channel"], "天猫")
        self.assertEqual(updated["launch_price"], 629)
        self.assertEqual(updated["workflow_restart_required"], 1)
        self.assertEqual(updated["c_published_version_no"], completed["c_published_version_no"])
        c_formal = catalog_db.products_for_c_published_versions(self.catalog_db_path, [updated])[0]
        self.assertEqual(c_formal["launch_channel"], "唯品")
        self.assertEqual(c_formal["launch_price"], 599)
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            catalog_db.change_product_status(
                connection,
                source["id"],
                "published",
                b_user["id"],
                "重新提交给运营部",
                "测试企划字段触发后的新运营批次。",
            )
        republished = catalog_db.get_product(self.catalog_db_path, source["id"])
        self.assertEqual(republished["workflow_restart_required"], 0)
        self.assertEqual(republished["c_release_no"], completed["c_release_no"] + 1)
        c_republished = catalog_db.products_for_c_published_versions(self.catalog_db_path, [republished])[0]
        self.assertEqual(c_republished["launch_channel"], "天猫")
        self.assertEqual(c_republished["launch_price"], 629)

    def test_configured_planning_channel_is_accepted_by_catalog_callback(self):
        planning_db.save_channel_option(self.planning_db_path, "直播首发", 40)
        app = CatalogApplication(
            self.catalog_db_path,
            Path(self.temp.name) / "uploads",
            planning_api_token="planning-secret",
        )
        source = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "pending"
        )
        response = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{source['id']}/price-publication",
            method="POST",
            body=json.dumps(
                {
                    "publication_id": "PC-CONFIGURED-CHANNEL",
                    "source_version_no": source["current_version_no"],
                    "category": "其他",
                    "launch_channel": "直播首发",
                    "launch_price": 599,
                }
            ).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertTrue(response["status"].startswith("200"))
        self.assertEqual(catalog_db.get_product(self.catalog_db_path, source["id"])["launch_channel"], "直播首发")

    def test_catalog_image_gate_and_image_only_refresh_preserve_workflow(self):
        catalog_app = CatalogApplication(
            self.catalog_db_path,
            Path(self.temp.name) / "uploads",
            planning_api_token="planning-secret",
        )
        pending = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "pending"
        )
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            connection.execute(
                "UPDATE products SET image_url = '', image_gallery_json = '[]' WHERE id = ?",
                (pending["id"],),
            )
        response = self.wsgi_request(
            catalog_app,
            "/api/internal/planning/products",
            authorization="Bearer planning-secret",
        )
        payload = json.loads(response["body"])
        self.assertTrue(payload["image_gate"])
        self.assertNotIn(pending["id"], {item["id"] for item in payload["items"]})
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            connection.execute(
                "UPDATE products SET image_gallery_json = ?, image_url = '' WHERE id = ?",
                (json.dumps(["https://example.com/images/gallery-only.jpg"]), pending["id"]),
            )
        response = self.wsgi_request(
            catalog_app,
            "/api/internal/planning/products",
            authorization="Bearer planning-secret",
        )
        payload = json.loads(response["body"])
        gallery_only = next(item for item in payload["items"] if item["id"] == pending["id"])
        self.assertEqual(gallery_only["image_url"], "https://example.com/images/gallery-only.jpg")

        completed = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "published"
        )
        before = catalog_db.get_product(self.catalog_db_path, completed["id"])
        b_user = next(user for user in catalog_db.list_users(self.catalog_db_path) if user["username"] == "b_editor")
        changed = dict(before)
        changed["image_url"] = "https://example.com/images/professional-replacement.jpg"
        changed["image_gallery_json"] = json.dumps([changed["image_url"]])
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            catalog_db.update_product(connection, completed["id"], changed, b_user["id"])
        after = catalog_db.get_product(self.catalog_db_path, completed["id"])
        self.assertEqual(after["status"], before["status"])
        self.assertEqual(after["current_version_no"], before["current_version_no"])
        self.assertEqual(after["c_published_version_no"], before["c_published_version_no"])
        self.assertEqual(after["revision_flag"], before["revision_flag"])
        self.assertEqual(after["image_version_no"], before["image_version_no"] + 1)
        c_formal_after_image = catalog_db.products_for_c_published_versions(self.catalog_db_path, [after])[0]
        self.assertEqual(c_formal_after_image["image_url"], changed["image_url"])

        response = self.wsgi_request(
            catalog_app,
            f"/api/internal/planning/products?known_ids={completed['id']}",
            authorization="Bearer planning-secret",
        )
        payload = json.loads(response["body"])
        update = next(item for item in payload["image_updates"] if item["id"] == completed["id"])
        self.assertEqual(update["image_url"], changed["image_url"])
        self.assertEqual(update["image_version_no"], after["image_version_no"])

        replacement_form = {
            "image_url": "https://example.com/images/professional-main.jpg",
            "image_gallery_existing__0": "https://example.com/images/snapshot.jpg",
            "image_gallery_manual__0": "https://example.com/images/snapshot.jpg",
        }
        catalog_app.apply_image_upload(
            replacement_form,
            {},
            existing_image_url="https://example.com/images/snapshot.jpg",
        )
        self.assertEqual(
            json.loads(replacement_form["image_gallery_json"]),
            ["https://example.com/images/professional-main.jpg"],
        )
        preserved_form = {}
        catalog_app.apply_image_upload(
            preserved_form,
            {},
            existing_image_url="https://example.com/images/professional-main.jpg",
            existing_image_gallery=json.dumps(
                [
                    "https://example.com/images/professional-main.jpg",
                    "https://example.com/images/detail.jpg",
                ]
            ),
        )
        self.assertEqual(
            json.loads(preserved_form["image_gallery_json"]),
            [
                "https://example.com/images/professional-main.jpg",
                "https://example.com/images/detail.jpg",
            ],
        )

        forbidden = dict(after, launch_price=999)
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            with self.assertRaisesRegex(PermissionError, "上新价格只能由商品企划中心"):
                catalog_db.update_product(connection, completed["id"], forbidden, b_user["id"])

        a_user = next(user for user in catalog_db.list_users(self.catalog_db_path) if user["username"] == "a_editor")
        forbidden_image = dict(after, image_url="https://example.com/images/a-cannot-replace.jpg")
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            with self.assertRaisesRegex(PermissionError, "图片只能由商品部"):
                catalog_db.update_product(connection, completed["id"], forbidden_image, a_user["id"])
        self.assertEqual(
            editable_field_keys_for_user(b_user, {**after, "status": "received"}),
            ("image_url",),
        )

    def test_catalog_planning_source_requires_positive_tax_included_price(self):
        app = CatalogApplication(
            self.catalog_db_path,
            Path(self.temp.name) / "uploads",
            planning_api_token="planning-secret",
        )
        pending = next(
            item for item in catalog_db.list_products(self.catalog_db_path) if item["status"] == "pending"
        )
        with catalog_db.get_connection(self.catalog_db_path) as connection:
            connection.execute(
                "UPDATE products SET image_url = ?, image_gallery_json = ?, tax_included_price = ? WHERE id = ?",
                ("https://example.com/images/cost-gate.jpg", "[]", 0, pending["id"]),
            )
        response = self.wsgi_request(
            app,
            "/api/internal/planning/products",
            authorization="Bearer planning-secret",
        )
        payload = json.loads(response["body"])
        self.assertNotIn(pending["id"], {item["id"] for item in payload["items"]})
        self.assertIn(pending["id"], payload["withdrawn_ids"])
        self.assertTrue(payload["cost_gate"])
        rejected_publication = self.wsgi_request(
            app,
            f"/api/internal/planning/products/{pending['id']}/price-publication",
            method="POST",
            body=json.dumps(
                {
                    "publication_id": "PC-COST-GATE-REJECTED",
                    "source_version_no": pending["current_version_no"],
                    "category": "其他",
                    "launch_channel": "唯品",
                    "launch_price": 599,
                }
            ).encode(),
            content_type="application/json",
            authorization="Bearer planning-secret",
        )
        self.assertTrue(rejected_publication["status"].startswith("400"))
        self.assertIn("有效的含税采购成本", rejected_publication["body"].decode("utf-8"))

        with catalog_db.get_connection(self.catalog_db_path) as connection:
            connection.execute(
                "UPDATE products SET tax_included_price = ? WHERE id = ?",
                (188, pending["id"]),
            )
        response = self.wsgi_request(
            app,
            "/api/internal/planning/products",
            authorization="Bearer planning-secret",
        )
        payload = json.loads(response["body"])
        self.assertIn(pending["id"], {item["id"] for item in payload["items"]})

    def test_sync_requires_valid_catalog_cost_and_cleans_legacy_missing_cost(self):
        invalid = {
            "id": 4101,
            "style_code": "COST-GATE-INVALID",
            "product_name": "缺少含税价款",
            "season_year": "2027秋冬",
            "supplier": "成本门槛供应商",
            "actual_cost": None,
            "tax_included_price": None,
            "image_url": "https://example.com/images/cost-gate-invalid.jpg",
            "status": "pending",
            "lifecycle_status": "active",
            "submitted_to_merchandise": True,
            "source_version_no": 1,
        }
        valid_tax_fallback = dict(
            invalid,
            id=4102,
            style_code="COST-GATE-FALLBACK",
            product_name="只有藏宝阁含税价款",
            tax_included_price=188,
        )
        planning_db.upsert_source_products(self.planning_db_path, [invalid])
        result = planning_db.synchronize_source_products(
            self.planning_db_path,
            [invalid, valid_tax_fallback],
        )
        self.assertEqual(result["synced"], 1)
        self.assertEqual(result["removed"], 1)
        self.assertIsNone(planning_db.get_source_product(self.planning_db_path, invalid["id"]))
        fallback = planning_db.get_source_product(self.planning_db_path, valid_tax_fallback["id"])
        self.assertEqual(fallback["actual_cost"], 188)

        actual_cost_only = dict(
            valid_tax_fallback,
            id=4103,
            style_code="COST-GATE-ACTUAL-ONLY",
            actual_cost=188,
            tax_included_price=None,
        )
        result = planning_db.synchronize_source_products(self.planning_db_path, [actual_cost_only])
        self.assertEqual(result["synced"], 0)
        self.assertIsNone(planning_db.get_source_product(self.planning_db_path, actual_cost_only["id"]))

        for invalid_cost in (0, -1, "not-a-number", "NaN", "Infinity"):
            item = dict(valid_tax_fallback, id=4200 + len(str(invalid_cost)), actual_cost=invalid_cost, tax_included_price=invalid_cost)
            self.assertFalse(planning_db.has_valid_source_cost(item))

    def test_sync_rechecks_every_admission_condition_before_insert(self):
        valid = {
            "id": 4300,
            "style_code": "STRICT-SYNC-VALID",
            "product_name": "严格同步有效款",
            "actual_cost": 188,
            "tax_included_price": 188,
            "image_url": "https://example.com/images/strict-sync.jpg",
            "status": "pending",
            "lifecycle_status": "active",
            "submitted_to_merchandise": True,
            "source_version_no": 1,
        }
        candidates = [
            valid,
            dict(valid, id=4301, style_code="STRICT-SYNC-NOT-SUBMITTED", submitted_to_merchandise=False),
            dict(valid, id=4302, style_code="STRICT-SYNC-NO-IMAGE", image_url="", image_gallery_json="[]"),
            dict(valid, id=4303, style_code="STRICT-SYNC-NO-COST", tax_included_price=None),
            dict(valid, id=4304, style_code="STRICT-SYNC-WRONG-STATUS", status="published"),
            dict(valid, id=4305, style_code="STRICT-SYNC-INACTIVE", lifecycle_status="withdrawn"),
        ]

        result = planning_db.synchronize_source_products(self.planning_db_path, candidates)

        self.assertEqual(result["synced"], 1)
        self.assertEqual(result["rejected"], 5)
        self.assertEqual(
            [item["id"] for item in planning_db.list_source_products(self.planning_db_path)],
            [valid["id"]],
        )

    def test_planning_revision_reuses_same_source_and_is_visible_in_published_filter(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋冬", None, 700, 4)
        source = {
            "id": 901,
            "style_code": "REV-901",
            "product_name": "修订测试毛衣",
            "season_year": "2026秋冬",
            "supplier": "修订供应商",
            "category": "其他",
            "actual_cost": 150,
            "image_url": "https://example.com/images/rev-901.jpg",
            "status": "pending",
            "lifecycle_status": "active",
            "source_version_no": 2,
            "image_version_no": 1,
        }
        planning_db.upsert_source_products(self.planning_db_path, [source])
        first = planning_db.create_pricing_record(self.planning_db_path, source, "商品部企划员")
        planning_db.mark_record_published(self.planning_db_path, first["id"], {"status": "published"})
        with planning_db.get_connection(self.planning_db_path) as connection:
            connection.execute(
                "UPDATE source_products SET lifecycle_status = 'withdrawn' WHERE id = ?",
                (source["id"],),
            )

        history = planning_db.list_published_source_products(self.planning_db_path)
        self.assertEqual([item["id"] for item in history], [source["id"]])
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")
        page = self.wsgi_request(app, "/workbench?status=published", cookie=planner_cookie)["body"].decode("utf-8")
        self.assertIn("发起同款修订", page)
        self.assertIn(f"action='/pricing/{source['id']}/revise'", page)

        response = self.wsgi_request(
            app,
            f"/pricing/{source['id']}/revise",
            method="POST",
            cookie=planner_cookie,
        )
        self.assertTrue(response["status"].startswith("302"))
        records = planning_db.list_pricing_records(self.planning_db_path)
        self.assertEqual(len(records), 2)
        self.assertEqual({record["source_product_id"] for record in records}, {source["id"]})
        self.assertEqual(planning_db.get_source_product(self.planning_db_path, source["id"])["lifecycle_status"], "active")
        sync_result = planning_db.synchronize_source_products(
            self.planning_db_path,
            [],
            withdrawn_ids=[source["id"]],
        )
        self.assertEqual(sync_result, {"synced": 0, "removed": 0, "withdrawn": 0})
        self.assertEqual(planning_db.get_source_product(self.planning_db_path, source["id"])["lifecycle_status"], "active")

    def test_planning_image_proxy_uses_internal_catalog_endpoint(self):
        source = {
            "id": 77,
            "style_code": "IMG-077",
            "product_name": "图片代理测试款",
            "image_url": "/media/planning-test.png",
            "status": "pending",
            "lifecycle_status": "active",
            "source_version_no": 2,
        }
        planning_db.upsert_source_products(self.planning_db_path, [source])
        app = PlanningApplication(self.planning_db_path, "http://catalog.test", "planning-secret")
        cookie = self.login_cookie(app, "planner")
        image_bytes = b"\x89PNG\r\n\x1a\nplanning-image"

        with patch("planning_center.web.urlopen", return_value=io.BytesIO(image_bytes)) as mocked_urlopen:
            response = self.wsgi_request(
                app,
                "/source-products/77/image?v=2",
                cookie=cookie,
            )

        self.assertTrue(response["status"].startswith("200"))
        self.assertEqual(response["body"], image_bytes)
        self.assertEqual(dict(response["headers"])["Content-Type"], "image/png")
        request = mocked_urlopen.call_args.args[0]
        self.assertEqual(request.full_url, "http://catalog.test/api/internal/planning/products/77/image")
        self.assertEqual(request.get_header("Authorization"), "Bearer planning-secret")

    def test_planning_image_proxy_sniffs_opaque_external_jpeg(self):
        source = {
            "id": 78,
            "style_code": "IMG-078",
            "product_name": "无后缀图片代理测试款",
            "image_url": "https://image-host.test/asset?id=professional-002",
            "status": "pending",
            "lifecycle_status": "active",
            "source_version_no": 1,
        }
        planning_db.upsert_source_products(self.planning_db_path, [source])
        app = PlanningApplication(self.planning_db_path, "http://catalog.test", "planning-secret")
        cookie = self.login_cookie(app, "planner")
        image_bytes = b"\xff\xd8\xffopaque-jpeg"

        response_headers = {"Content-Type": "application/octet-stream"}
        image_response = io.BytesIO(image_bytes)
        image_response.headers = response_headers
        with patch("planning_center.web.urlopen", return_value=image_response):
            response = self.wsgi_request(
                app,
                "/source-products/78/image?v=1",
                cookie=cookie,
            )

        self.assertTrue(response["status"].startswith("200"))
        self.assertEqual(response["body"], image_bytes)
        self.assertEqual(dict(response["headers"])["Content-Type"], "image/jpeg")

    def test_pricing_formula_rounds_down_to_9_and_snapshots_rules(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋", None, 600, 4)
        product = {
            "id": 7, "style_code": "M001", "product_name": "测试毛衣", "season_year": "2026秋",
            "supplier": "供应商 A", "category": "其他", "actual_cost": 150, "source_version_no": 3,
        }
        record = planning_db.create_pricing_record(self.planning_db_path, product, "测试企划员")
        self.assertEqual(record["raw_price"], 600)
        self.assertEqual(record["launch_price"], 599)
        self.assertEqual(record["fixed_multiplier"], 4)

    def test_missing_category_rule_blocks_pricing(self):
        product = {
            "id": 8, "style_code": "M008", "product_name": "未配置品类", "season_year": "2026秋",
            "supplier": "供应商 A", "category": "外套", "actual_cost": 150, "source_version_no": 1,
        }
        with self.assertRaisesRegex(ValueError, "尚未落入成本区间倍率规则"):
            planning_db.create_pricing_record(self.planning_db_path, product, "测试企划员")

    def test_idempotent_publication_is_marked_published(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋", None, 600, 4)
        product = {
            "id": 10, "style_code": "M010", "product_name": "幂等测试", "season_year": "2026秋",
            "supplier": "供应商 A", "category": "其他", "actual_cost": 150, "source_version_no": 1,
        }
        record = planning_db.create_pricing_record(self.planning_db_path, product, "测试企划员")
        updated = planning_db.mark_record_published(
            self.planning_db_path,
            record["id"],
            {"status": "already_published"},
        )
        self.assertEqual(updated["status"], "published")

    def test_catalog_launch_price_input_and_validation_require_integer(self):
        app = CatalogApplication(self.catalog_db_path, Path(self.temp.name) / "uploads")
        markup = app.render_input(catalog_db.PRODUCT_FIELD_MAP["launch_price"], {"launch_price": 2539.0})
        self.assertIn('min="1" step="1" inputmode="numeric"', markup)
        self.assertIn('value="2539"', markup)
        self.assertNotIn('value="2539.0"', markup)

        valid_errors = app.validate_product_form(
            {"product_name": "整数价格测试", "style_code": "INTEGER-PRICE", "launch_price": "2539"}
        )
        self.assertEqual(valid_errors, [])
        fractional_errors = app.validate_product_form(
            {"product_name": "小数价格测试", "style_code": "FRACTION-PRICE", "launch_price": "2539.5"}
        )
        self.assertIn("上新价格必须是大于 0 的整数，不保留小数位。", fractional_errors)

    def test_dress_uses_fixed_multiplier_regardless_of_cost(self):
        planning_db.save_category_rule(self.planning_db_path, "", "连衣裙", 4.2)
        for product_id, cost in ((20, 150), (21, 900)):
            product = {
                "id": product_id,
                "style_code": f"D{product_id}",
                "product_name": "测试连衣裙",
                "season_year": "2026秋",
                "supplier": "供应商 A",
                "category": "连衣裙",
                "actual_cost": cost,
                "source_version_no": 1,
            }
            record = planning_db.create_pricing_record(self.planning_db_path, product, "测试企划员")
            self.assertEqual(record["fixed_multiplier"], 4.2)
            self.assertEqual(record["raw_price"], cost * 4.2)

    def test_other_category_cost_ranges_use_inclusive_lower_exclusive_upper(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "", None, 600, 4)
        planning_db.save_category_cost_rule(self.planning_db_path, "", 600, 791, 3.9)
        planning_db.save_category_cost_rule(self.planning_db_path, "", 791, 1001, 3.8)
        planning_db.save_category_cost_rule(self.planning_db_path, "", 1001, None, 3.7)

        cases = ((599.99, 4), (600, 3.9), (790, 3.9), (791, 3.8), (1000, 3.8), (1001, 3.7))
        for cost, expected in cases:
            fixed, coefficient = planning_db.resolve_rules(
                self.planning_db_path,
                "2026秋",
                "其他",
                "供应商 A",
                cost,
            )
            self.assertEqual(fixed, expected)
            self.assertEqual(coefficient, 1)

    def test_every_non_dress_category_uses_the_shared_cost_range_rules(self):
        planning_db.save_category_option(self.planning_db_path, "上衣", "上衣", 20)
        planning_db.save_category_option(self.planning_db_path, "裤装", "裤装", 30)
        planning_db.save_category_cost_rule(self.planning_db_path, "", None, 600, 4)
        planning_db.save_category_cost_rule(self.planning_db_path, "", 600, None, 3.8)

        for category in ("其他", "上衣", "裤装"):
            fixed, coefficient = planning_db.resolve_rules(
                self.planning_db_path,
                "2026秋",
                category,
                "供应商 A",
                620,
            )
            self.assertEqual(fixed, 3.8, category)
            self.assertEqual(coefficient, 1, category)

        planning_db.save_category_rule(self.planning_db_path, "", "连衣裙", 4.2)
        fixed, _ = planning_db.resolve_rules(self.planning_db_path, "2026秋", "连衣裙", "供应商 A", 620)
        self.assertEqual(fixed, 4.2)

    def test_legacy_category_and_cost_rule_labels_are_migrated(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "", None, 600, 4)
        legacy_product = {
            "id": 9901,
            "style_code": "LEGACY-9901",
            "product_name": "历史兜底分类资料",
            "season_year": "2026秋",
            "supplier": "历史供应商",
            "category": planning_db.CATEGORY_FALLBACK_OPTION,
            "actual_cost": 150,
            "status": "pending",
            "lifecycle_status": "active",
            "source_version_no": 1,
        }
        planning_db.upsert_source_products(self.planning_db_path, [legacy_product])
        legacy_record = planning_db.create_pricing_record(self.planning_db_path, legacy_product, "测试企划员")
        with planning_db.get_connection(self.planning_db_path) as connection:
            connection.execute(
                "UPDATE category_cost_rules SET category = ?",
                (planning_db.LEGACY_CATEGORY_FALLBACK_OPTION,),
            )
            connection.execute(
                "UPDATE category_options SET name = ?, note = ? WHERE name = ?",
                (
                    planning_db.LEGACY_CATEGORY_FALLBACK_OPTION,
                    "未命中其他关键词时的默认品类；使用成本区间倍率。",
                    planning_db.CATEGORY_FALLBACK_OPTION,
                ),
            )
            connection.executemany(
                "INSERT INTO category_options (name, keywords, pricing_group, sort_order, note, updated_at) VALUES (?, ?, 'other', ?, '', ?)",
                [
                    (name, keywords, sort_order, planning_db.utc_now())
                    for name, (keywords, sort_order) in planning_db.LEGACY_DEFAULT_CATEGORY_OPTIONS.items()
                ],
            )
            connection.execute(
                "UPDATE source_products SET category = ?, category_suggestion = ? WHERE id = ?",
                (
                    planning_db.LEGACY_CATEGORY_FALLBACK_OPTION,
                    planning_db.LEGACY_CATEGORY_FALLBACK_OPTION,
                    legacy_product["id"],
                ),
            )
            connection.execute(
                "UPDATE pricing_records SET category = ? WHERE id = ?",
                (planning_db.LEGACY_CATEGORY_FALLBACK_OPTION, legacy_record["id"]),
            )

        planning_db.init_db(self.planning_db_path)
        option_names = [item["name"] for item in planning_db.list_category_options(self.planning_db_path)]
        self.assertIn(planning_db.CATEGORY_FALLBACK_OPTION, option_names)
        self.assertNotIn(planning_db.LEGACY_CATEGORY_FALLBACK_OPTION, option_names)
        self.assertNotIn("毛衣", option_names)
        self.assertNotIn("衬衫", option_names)
        self.assertNotIn("外套", option_names)
        self.assertNotIn("半身裙", option_names)
        self.assertNotIn("裤装", option_names)
        self.assertEqual(planning_db.get_source_product(self.planning_db_path, legacy_product["id"])["category"], "其他")
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, legacy_record["id"])["category"], "其他")
        cost_rules = planning_db.list_category_cost_rules(self.planning_db_path)
        self.assertEqual([item["category"] for item in cost_rules], [planning_db.NON_DRESS_PRICING_CATEGORY])
        fixed, _ = planning_db.resolve_rules(self.planning_db_path, "2026秋", "其他", "供应商 A", 150)
        self.assertEqual(fixed, 4)

    def test_other_category_cost_ranges_reject_overlap_and_allow_season_override(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "", None, 600, 4)
        with self.assertRaisesRegex(ValueError, "区间.*重叠"):
            planning_db.save_category_cost_rule(self.planning_db_path, "", 500, 700, 3.9)

        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋", None, 600, 4.1)
        fixed, _ = planning_db.resolve_rules(
            self.planning_db_path,
            "2026秋",
            "外套",
            "供应商 A",
            500,
        )
        self.assertEqual(fixed, 4.1)

    def test_all_pricing_rule_types_can_be_edited_and_deleted(self):
        planning_db.save_category_rule(self.planning_db_path, "", "连衣裙", 4.2, "默认")
        dress = planning_db.list_category_rules(self.planning_db_path)[0]
        planning_db.save_category_rule(
            self.planning_db_path,
            "2026秋冬",
            "连衣裙",
            4.3,
            "秋冬调整",
            dress["id"],
        )
        edited_dress = planning_db.list_category_rules(self.planning_db_path)[0]
        self.assertEqual(edited_dress["season_year"], "2026秋冬")
        self.assertEqual(edited_dress["multiplier"], 4.3)

        planning_db.save_category_cost_rule(self.planning_db_path, "", None, 600, 4, "首档")
        cost_rule = planning_db.list_category_cost_rules(self.planning_db_path)[0]
        planning_db.save_category_cost_rule(
            self.planning_db_path,
            "",
            None,
            650,
            4.1,
            "调整首档",
            cost_rule["id"],
        )
        fixed, _ = planning_db.resolve_rules(self.planning_db_path, "2026秋冬", "其他", "供应商 A", 620)
        self.assertEqual(fixed, 4.1)

        planning_db.save_supplier_coefficient(self.planning_db_path, "", "供应商 A", 1.0, "默认")
        supplier_rule = planning_db.list_supplier_coefficients(self.planning_db_path)[0]
        planning_db.save_supplier_coefficient(
            self.planning_db_path,
            "2026秋冬",
            "供应商 A",
            1.05,
            "秋冬调整",
            supplier_rule["id"],
        )
        _, coefficient = planning_db.resolve_rules(self.planning_db_path, "2026秋冬", "连衣裙", "供应商 A", 300)
        self.assertEqual(coefficient, 1.05)

        planning_db.delete_category_rule(self.planning_db_path, dress["id"])
        planning_db.delete_category_cost_rule(self.planning_db_path, cost_rule["id"])
        planning_db.delete_supplier_coefficient(self.planning_db_path, supplier_rule["id"])
        self.assertEqual(planning_db.list_category_rules(self.planning_db_path), [])
        self.assertEqual(planning_db.list_category_cost_rules(self.planning_db_path), [])
        self.assertEqual(planning_db.list_supplier_coefficients(self.planning_db_path), [])

    def test_rule_page_explains_category_and_cost_logic(self):
        category_names = [item["name"] for item in planning_db.list_category_options(self.planning_db_path)]
        self.assertEqual(category_names, ["连衣裙", "其他"])
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        login = self.wsgi_request(
            app,
            "/login",
            method="POST",
            body=urlencode({"username": "planning_admin", "password": "demo123"}).encode(),
        )
        cookie = dict(login["headers"])["Set-Cookie"].split(";", 1)[0]
        response = self.wsgi_request(app, "/rules", cookie=cookie)
        page = response["body"].decode("utf-8")
        self.assertIn("连衣裙固定倍率", page)
        self.assertIn("非连衣裙品类倍率", page)
        self.assertIn("下限包含，上限不包含", page)
        self.assertIn("可按实际业务新增任意数量的成本区间", page)
        self.assertIn("规则维护账号", page)
        self.assertIn("<h1>规则</h1>", page)
        self.assertIn("品类选项", page)
        self.assertIn("渠道选项", page)
        self.assertNotIn("<h1>定价规则</h1>", page)

    def test_only_admin_can_maintain_pricing_rules(self):
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_login = self.wsgi_request(
            app,
            "/login",
            method="POST",
            body=urlencode({"username": "planner", "password": "demo123"}).encode(),
        )
        planner_cookie = dict(planner_login["headers"])["Set-Cookie"].split(";", 1)[0]
        readonly_page = self.wsgi_request(app, "/rules", cookie=planner_cookie)["body"].decode("utf-8")
        self.assertIn("规则只读账号", readonly_page)
        self.assertNotIn("action='/rules/category-cost'", readonly_page)
        denied = self.wsgi_request(
            app,
            "/rules/category-cost",
            method="POST",
            body=urlencode({"upper_cost": "500", "multiplier": "4"}).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(denied["status"].startswith("403"))

        admin_login = self.wsgi_request(
            app,
            "/login",
            method="POST",
            body=urlencode({"username": "planning_admin", "password": "demo123"}).encode(),
        )
        admin_cookie = dict(admin_login["headers"])["Set-Cookie"].split(";", 1)[0]
        created = self.wsgi_request(
            app,
            "/rules/category-cost",
            method="POST",
            body=urlencode({"upper_cost": "600", "multiplier": "4", "note": "首档"}).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(created["status"].startswith("302"))
        cost_rule = planning_db.list_category_cost_rules(self.planning_db_path)[0]
        edit_page = self.wsgi_request(app, f"/rules?edit_cost={cost_rule['id']}", cookie=admin_cookie)["body"].decode("utf-8")
        self.assertIn("保存修改", edit_page)
        self.assertIn("编辑", edit_page)
        self.assertIn("删除", edit_page)

        updated = self.wsgi_request(
            app,
            "/rules/category-cost",
            method="POST",
            body=urlencode({"rule_id": str(cost_rule["id"]), "upper_cost": "650", "multiplier": "4.1", "note": "调整后"}).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(updated["status"].startswith("302"))
        edited = planning_db.list_category_cost_rules(self.planning_db_path)[0]
        self.assertEqual(edited["upper_cost"], 650)
        self.assertEqual(edited["multiplier"], 4.1)

    def test_category_and_channel_options_are_configurable_by_admin_only(self):
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")
        admin_cookie = self.login_cookie(app, "planning_admin")

        denied = self.wsgi_request(
            app,
            "/rules/category-option",
            method="POST",
            body=urlencode({"name": "西装", "keywords": "西服,西装外套", "sort_order": "35"}).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(denied["status"].startswith("403"))

        created_category = self.wsgi_request(
            app,
            "/rules/category-option",
            method="POST",
            body=urlencode({"name": "西装", "keywords": "西服,西装外套", "sort_order": "35", "note": "测试"}).encode(),
            cookie=admin_cookie,
        )
        created_channel = self.wsgi_request(
            app,
            "/rules/channel-option",
            method="POST",
            body=urlencode({"name": "抖音", "sort_order": "40"}).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(created_category["status"].startswith("302"))
        self.assertTrue(created_channel["status"].startswith("302"))
        category = next(item for item in planning_db.list_category_options(self.planning_db_path) if item["name"] == "西装")
        channel = next(item for item in planning_db.list_channel_options(self.planning_db_path) if item["name"] == "抖音")

        updated = self.wsgi_request(
            app,
            "/rules/category-option",
            method="POST",
            body=urlencode({"option_id": str(category["id"]), "name": "西装", "keywords": "西服,西装外套,套装", "sort_order": "34"}).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(updated["status"].startswith("302"))
        self.assertIn("套装", next(item for item in planning_db.list_category_options(self.planning_db_path) if item["id"] == category["id"])["keywords"])

        deleted = self.wsgi_request(
            app,
            f"/rules/channel-option/{channel['id']}/delete",
            method="POST",
            cookie=admin_cookie,
        )
        self.assertTrue(deleted["status"].startswith("302"))
        self.assertNotIn("抖音", [item["name"] for item in planning_db.list_channel_options(self.planning_db_path)])

    def test_sync_infers_category_from_product_name_with_longest_keyword(self):
        planning_db.save_category_option(self.planning_db_path, "西装", "西服,西装外套", 35)
        products = [
            {"id": 61, "product_name": "通勤西装外套", "actual_cost": 300, "status": "pending", "lifecycle_status": "active", "source_version_no": 1},
            {"id": 62, "product_name": "法式连衣裙", "actual_cost": 300, "status": "pending", "lifecycle_status": "active", "source_version_no": 1},
            {"id": 63, "product_name": "基础测试款", "actual_cost": 300, "status": "pending", "lifecycle_status": "active", "source_version_no": 1},
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        self.assertEqual(planning_db.get_source_product(self.planning_db_path, 61)["category_suggestion"], "西装")
        self.assertEqual(planning_db.get_source_product(self.planning_db_path, 62)["category"], "连衣裙")
        self.assertEqual(planning_db.get_source_product(self.planning_db_path, 63)["category"], "其他")

    def test_initial_review_can_change_category_recalculate_and_select_channel(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋冬", None, 700, 4)
        planning_db.save_category_rule(self.planning_db_path, "2026秋冬", "连衣裙", 4.2)
        product = {
            "id": 64,
            "style_code": "R064",
            "product_name": "基础毛衣",
            "season_year": "2026秋冬",
            "supplier": "测试供应商",
            "category": "其他",
            "actual_cost": 150,
            "status": "pending",
            "lifecycle_status": "active",
            "source_version_no": 1,
        }
        planning_db.upsert_source_products(self.planning_db_path, [product])
        record = planning_db.create_pricing_record(self.planning_db_path, product, "商品部企划员")
        self.assertEqual(record["calculated_price"], 599)
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        cookie = self.login_cookie(app, "planner")

        missing_channel = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/submit-review",
            method="POST",
            body=urlencode({"launch_price": "599", "category": "连衣裙"}).encode(),
            cookie=cookie,
        )
        self.assertTrue(missing_channel["status"].startswith("400"))

        submitted = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/submit-review",
            method="POST",
            body=urlencode({"launch_price": "599", "category": "连衣裙", "channel": "唯品"}).encode(),
            cookie=cookie,
        )
        self.assertTrue(submitted["status"].startswith("302"))
        updated = planning_db.get_pricing_record(self.planning_db_path, record["id"])
        self.assertEqual(updated["category"], "连衣裙")
        self.assertEqual(updated["channel"], "唯品")
        self.assertEqual(updated["fixed_multiplier"], 4.2)
        self.assertEqual(updated["calculated_price"], 629)
        self.assertEqual(updated["launch_price"], 629)
        self.assertEqual(planning_db.get_source_product(self.planning_db_path, 64)["category"], "连衣裙")

    def test_formal_database_can_bootstrap_first_admin(self):
        formal_path = Path(self.temp.name) / "formal-planning.db"
        planning_db.init_db(
            formal_path,
            seed_demo=False,
            bootstrap_admin={"username": "owner", "password": "ChangeMe123", "display_name": "正式管理员"},
        )
        user = planning_db.authenticate_user(formal_path, "owner", "ChangeMe123")
        self.assertEqual(user["role"], "admin")

    def test_account_management_and_password_workflows(self):
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")
        admin_cookie = self.login_cookie(app, "planning_admin")

        denied = self.wsgi_request(app, "/accounts", cookie=planner_cookie)
        self.assertTrue(denied["status"].startswith("403"))
        self.assertIn("只有企划管理员可以管理账号", denied["body"].decode("utf-8"))

        account_page = self.wsgi_request(app, "/accounts", cookie=admin_cookie)
        account_markup = account_page["body"].decode("utf-8")
        self.assertTrue(account_page["status"].startswith("200"))
        self.assertIn("账号管理", account_markup)
        self.assertIn("新增账号", account_markup)
        self.assertIn("修改我的密码", account_markup)

        mismatch = self.wsgi_request(
            app,
            "/accounts",
            method="POST",
            cookie=admin_cookie,
            body=urlencode(
                {
                    "username": "merch_f",
                    "display_name": "商品部同事 F",
                    "role": "planner",
                    "password": "StartPass88",
                    "confirm_password": "Different88",
                }
            ).encode(),
        )
        self.assertTrue(mismatch["status"].startswith("400"))
        self.assertIn("两次输入的初始密码不一致", mismatch["body"].decode("utf-8"))

        weak_password = self.wsgi_request(
            app,
            "/accounts",
            method="POST",
            cookie=admin_cookie,
            body=urlencode(
                {
                    "username": "merch_f",
                    "display_name": "商品部同事 F",
                    "role": "planner",
                    "password": "short",
                    "confirm_password": "short",
                }
            ).encode(),
        )
        self.assertTrue(weak_password["status"].startswith("400"))
        self.assertIn("密码至少需要 8 位", weak_password["body"].decode("utf-8"))

        created_response = self.wsgi_request(
            app,
            "/accounts",
            method="POST",
            cookie=admin_cookie,
            body=urlencode(
                {
                    "username": "merch_f",
                    "display_name": "商品部同事 F",
                    "role": "planner",
                    "password": "StartPass88",
                    "confirm_password": "StartPass88",
                }
            ).encode(),
        )
        self.assertTrue(created_response["status"].startswith("302"))
        managed_user = next(item for item in planning_db.list_users(self.planning_db_path) if item["username"] == "merch_f")
        self.assertEqual(managed_user["role"], "planner")
        self.assertIsNotNone(planning_db.authenticate_user(self.planning_db_path, "merch_f", "StartPass88"))

        duplicate = self.wsgi_request(
            app,
            "/accounts",
            method="POST",
            cookie=admin_cookie,
            body=urlencode(
                {
                    "username": "MERCH_F",
                    "display_name": "重复账号",
                    "role": "planner",
                    "password": "AnotherPass88",
                    "confirm_password": "AnotherPass88",
                }
            ).encode(),
        )
        self.assertTrue(duplicate["status"].startswith("400"))
        self.assertIn("该登录账号已存在", duplicate["body"].decode("utf-8"))

        login = self.wsgi_request(
            app,
            "/login",
            method="POST",
            body=urlencode({"username": "merch_f", "password": "StartPass88"}).encode(),
        )
        managed_cookie = dict(login["headers"])["Set-Cookie"].split(";", 1)[0]
        wrong_current_password = self.wsgi_request(
            app,
            "/profile/password",
            method="POST",
            cookie=managed_cookie,
            body=urlencode(
                {
                    "current_password": "WrongPass88",
                    "new_password": "PersonalPass99",
                    "confirm_password": "PersonalPass99",
                }
            ).encode(),
        )
        self.assertTrue(wrong_current_password["status"].startswith("400"))
        self.assertIn("当前密码不正确", wrong_current_password["body"].decode("utf-8"))

        changed = self.wsgi_request(
            app,
            "/profile/password",
            method="POST",
            cookie=managed_cookie,
            body=urlencode(
                {
                    "current_password": "StartPass88",
                    "new_password": "PersonalPass99",
                    "confirm_password": "PersonalPass99",
                }
            ).encode(),
        )
        self.assertTrue(changed["status"].startswith("302"))
        self.assertIsNone(planning_db.authenticate_user(self.planning_db_path, "merch_f", "StartPass88"))
        self.assertIsNotNone(planning_db.authenticate_user(self.planning_db_path, "merch_f", "PersonalPass99"))
        self.assertTrue(self.wsgi_request(app, "/dashboard", cookie=managed_cookie)["status"].startswith("200"))

        reset = self.wsgi_request(
            app,
            f"/accounts/{managed_user['id']}/reset-password",
            method="POST",
            cookie=admin_cookie,
            body=urlencode({"new_password": "ResetPass101", "confirm_password": "ResetPass101"}).encode(),
        )
        self.assertTrue(reset["status"].startswith("302"))
        self.assertIsNone(planning_db.authenticate_user(self.planning_db_path, "merch_f", "PersonalPass99"))
        self.assertIsNotNone(planning_db.authenticate_user(self.planning_db_path, "merch_f", "ResetPass101"))
        invalidated_session = self.wsgi_request(app, "/dashboard", cookie=managed_cookie)
        self.assertEqual(dict(invalidated_session["headers"])["Location"], "/login")

        disabled = self.wsgi_request(
            app,
            f"/accounts/{managed_user['id']}/toggle",
            method="POST",
            cookie=admin_cookie,
        )
        self.assertTrue(disabled["status"].startswith("302"))
        self.assertIsNone(planning_db.authenticate_user(self.planning_db_path, "merch_f", "ResetPass101"))
        self.assertFalse(planning_db.get_user(self.planning_db_path, managed_user["id"])["is_active"])

        enabled = self.wsgi_request(
            app,
            f"/accounts/{managed_user['id']}/toggle",
            method="POST",
            cookie=admin_cookie,
        )
        self.assertTrue(enabled["status"].startswith("302"))
        self.assertIsNotNone(planning_db.authenticate_user(self.planning_db_path, "merch_f", "ResetPass101"))

        self_toggle = self.wsgi_request(
            app,
            f"/accounts/{planning_db.authenticate_user(self.planning_db_path, 'planning_admin', 'demo123')['id']}/toggle",
            method="POST",
            cookie=admin_cookie,
        )
        self.assertTrue(self_toggle["status"].startswith("400"))
        self.assertIn("不能停用当前登录", self_toggle["body"].decode("utf-8"))

    def test_last_active_planning_admin_cannot_be_disabled(self):
        admin = planning_db.authenticate_user(self.planning_db_path, "planning_admin", "demo123")
        with self.assertRaisesRegex(ValueError, "最后一个有效的企划管理员"):
            planning_db.set_user_active(self.planning_db_path, admin["id"], False)
        self.assertTrue(planning_db.get_user(self.planning_db_path, admin["id"])["is_active"])

    def test_planning_workbench_login_and_sync(self):
        app = PlanningApplication(self.planning_db_path, "http://catalog.test", "token")
        login = self.wsgi_request(app, "/login", method="POST", body=urlencode({"username": "planner", "password": "demo123"}).encode())
        self.assertTrue(login["status"].startswith("302"))
        cookie = dict(login["headers"])["Set-Cookie"].split(";", 1)[0]
        source = {"id": 9, "style_code": "M009", "product_name": "测试款", "season_year": "2026秋", "supplier": "供应商", "category": "其他", "actual_cost": 150, "tax_included_price": 150, "image_url": "https://example.com/images/m009.jpg", "status": "pending", "lifecycle_status": "active", "submitted_to_merchandise": True, "source_version_no": 1, "updated_at": "", "creator_name": "跟单员"}
        withdrawn_source = dict(source, id=10, style_code="M010", product_name="误同步的已完成款")
        planning_db.upsert_source_products(self.planning_db_path, [withdrawn_source])
        with patch.object(app, "fetch_catalog_products", return_value={"source": "cangbaoge", "items": [source]}):
            rejected = self.wsgi_request(app, "/sync", method="POST", cookie=cookie)
        self.assertTrue(rejected["status"].startswith("400"))
        self.assertIn("未完成流程、图片和含税价准入校验", rejected["body"].decode("utf-8"))
        self.assertIsNone(planning_db.get_source_product(self.planning_db_path, source["id"]))
        catalog_payload = {
            "source": "cangbaoge",
            "items": [source],
            "withdrawn_ids": [10],
            "image_updates": [],
            "workflow_gate": True,
            "image_gate": True,
            "cost_gate": True,
            "eligibility_gate_version": 1,
        }
        with patch.object(app, "fetch_catalog_products", return_value=catalog_payload):
            response = self.wsgi_request(app, "/sync", method="POST", cookie=cookie)
        self.assertTrue(response["status"].startswith("302"))
        self.assertEqual(planning_db.get_source_product(self.planning_db_path, 9)["style_code"], "M009")
        self.assertIsNone(planning_db.get_source_product(self.planning_db_path, 10))

        with patch.object(app, "fetch_catalog_products", return_value=catalog_payload):
            workbench = self.wsgi_request(app, "/workbench", cookie=cookie)["body"].decode("utf-8")
        sync_message = "已自动同步 1 条藏宝阁“待商品部填写”资料。"
        self.assertEqual(workbench.count(sync_message), 1)
        self.assertNotIn("当前结果", workbench)
        self.assertLess(workbench.index(sync_message), workbench.index("<form class='workbench-filter'"))

    def test_sync_withdraws_completed_sources_and_preserves_pricing_audit(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋冬", None, 700, 4)
        pending = {
            "id": 201,
            "style_code": "SYNC-201",
            "product_name": "待商品部填写毛衣",
            "season_year": "2026秋冬",
            "supplier": "同步供应商",
            "category": "其他",
            "actual_cost": 150,
            "tax_included_price": 150,
            "status": "pending",
            "lifecycle_status": "active",
            "submitted_to_merchandise": True,
            "image_url": "https://example.com/images/sync-201.jpg",
            "source_version_no": 1,
        }
        completed_unworked = dict(pending, id=202, style_code="SYNC-202", product_name="误同步已完成款")
        completed_worked = dict(pending, id=203, style_code="SYNC-203", product_name="已测算后完成款")
        planning_db.upsert_source_products(
            self.planning_db_path,
            [pending, completed_unworked, completed_worked],
        )
        pricing_record = planning_db.create_pricing_record(
            self.planning_db_path,
            completed_worked,
            "测试企划员",
        )
        with planning_db.get_connection(self.planning_db_path) as connection:
            connection.execute(
                "UPDATE source_products SET status = 'published' WHERE id IN (202, 203)"
            )

        result = planning_db.synchronize_source_products(
            self.planning_db_path,
            [pending],
            withdrawn_ids=[202, 203],
        )

        self.assertEqual(result, {"synced": 1, "removed": 1, "withdrawn": 1})
        self.assertIsNone(planning_db.get_source_product(self.planning_db_path, 202))
        self.assertEqual(planning_db.get_source_product(self.planning_db_path, 203)["lifecycle_status"], "withdrawn")
        self.assertIsNotNone(planning_db.get_pricing_record(self.planning_db_path, pricing_record["id"]))
        self.assertEqual([item["id"] for item in planning_db.list_source_products(self.planning_db_path)], [201])

        second_result = planning_db.synchronize_source_products(self.planning_db_path, [pending])
        self.assertEqual(second_result, {"synced": 1, "removed": 0, "withdrawn": 0})

    def test_sync_does_not_withdraw_pending_source_only_because_response_omits_it(self):
        source = {
            "id": 204,
            "style_code": "SYNC-204",
            "product_name": "仍待商品部填写的款式",
            "status": "pending",
            "lifecycle_status": "active",
            "submitted_to_merchandise": True,
            "source_version_no": 1,
        }
        planning_db.upsert_source_products(self.planning_db_path, [source])

        result = planning_db.synchronize_source_products(self.planning_db_path, [])

        self.assertEqual(result, {"synced": 0, "removed": 0, "withdrawn": 0})
        self.assertIsNotNone(planning_db.get_source_product(self.planning_db_path, 204))

    def test_workbench_paginates_50_items_and_merges_filter_toolbar(self):
        products = [
            {
                "id": product_id,
                "style_code": f"PAGE-{product_id}",
                "style_color": f"PAGE-{product_id}-黑",
                "product_name": f"分页测试毛衣 {product_id}",
                "season_year": "2026秋冬",
                "supplier": "分页测试供应商",
                "actual_cost": 100,
                "status": "pending",
                "source_version_no": 1,
            }
            for product_id in range(1001, 1052)
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")

        page_one = self.wsgi_request(
            app,
            "/workbench?season_year=2026%E7%A7%8B%E5%86%AC&status=waiting&page=1",
            cookie=planner_cookie,
        )["body"].decode("utf-8")
        self.assertEqual(page_one.count("id='pricing-row-"), 50)
        self.assertEqual(page_one.count("name='suggest_ids'"), 50)
        self.assertNotIn("当前结果", page_one)
        self.assertIn("商品资料已加载，可按条件筛选。", page_one)
        self.assertIn("每页 50 款 · 第 1 / 2 页 · 共 51 款", page_one)
        self.assertEqual(page_one.count("aria-label='列表顶部分页'"), 1)
        self.assertEqual(page_one.count("aria-label='列表底部分页'"), 1)
        self.assertIn("<span>第 1 / 2 页</span>", page_one)
        self.assertLess(page_one.index("aria-label='列表顶部分页'"), page_one.index("<form id='pricing-batch-form'"))
        self.assertIn("href='/workbench?season_year=2026%E7%A7%8B%E5%86%AC&amp;status=waiting&amp;page=2'", page_one)
        self.assertLess(page_one.index("workbench-toolbar-summary"), page_one.index("workbench-filter"))
        self.assertNotIn("<section class='filter-bar'>", page_one)

        page_two = self.wsgi_request(
            app,
            "/workbench?season_year=2026%E7%A7%8B%E5%86%AC&status=waiting&page=2",
            cookie=planner_cookie,
        )["body"].decode("utf-8")
        self.assertEqual(page_two.count("id='pricing-row-"), 1)
        self.assertEqual(page_two.count("name='suggest_ids'"), 1)
        self.assertIn("每页 50 款 · 第 2 / 2 页 · 共 51 款", page_two)
        self.assertIn("id='pricing-row-1001'", page_two)

        clamped_page = self.wsgi_request(
            app,
            "/workbench?season_year=2026%E7%A7%8B%E5%86%AC&status=waiting&page=999",
            cookie=planner_cookie,
        )["body"].decode("utf-8")
        self.assertIn("每页 50 款 · 第 2 / 2 页 · 共 51 款", clamped_page)

    def test_cross_page_selection_prices_all_waiting_items_without_posting_every_id(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2027春夏", None, 600, 4)
        products = [
            {
                "id": 3000 + index,
                "style_code": f"ALL-{index:03d}",
                "style_color": f"ALL-{index:03d}-黑",
                "product_name": f"跨页批量毛衣 {index}",
                "season_year": "2027春夏",
                "supplier": "跨页测试供应商",
                "actual_cost": 100,
                "status": "pending",
                "source_version_no": 1,
            }
            for index in range(101)
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")

        page = self.wsgi_request(
            app,
            "/workbench?season_year=2027%E6%98%A5%E5%A4%8F&status=waiting",
            cookie=planner_cookie,
        )["body"].decode("utf-8")
        self.assertIn("选择全部筛选资料（101）", page)
        self.assertIn("id='pricing-selection-scope'", page)
        self.assertIn("const filteredActionCounts = {\"suggest\": 101", page)
        self.assertEqual(page.count("name='suggest_ids'"), 50)

        response = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                {
                    "batch_action": "suggest",
                    "selection_scope": "filtered",
                    "season_year": "2027春夏",
                }
            ).encode("utf-8"),
            cookie=planner_cookie,
        )
        self.assertTrue(response["status"].startswith("302"))
        self.assertIn("status=suggested", dict(response["headers"])["Location"])
        records = planning_db.list_pricing_records(self.planning_db_path, season_year="2027春夏")
        self.assertEqual(len(records), 101)
        self.assertEqual({record["calculated_price"] for record in records}, {399})

    def test_filtered_batch_suggest_skips_waiting_items_without_valid_cost(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2027秋冬", None, 600, 4)
        products = [
            {
                "id": 3090,
                "style_code": "FILTERED-COST-OK",
                "product_name": "筛选批量有效成本款",
                "season_year": "2027秋冬",
                "supplier": "筛选测试供应商",
                "category": "其他",
                "actual_cost": 100,
                "status": "pending",
                "source_version_no": 1,
            },
            {
                "id": 3091,
                "style_code": "FILTERED-COST-MISSING",
                "product_name": "筛选批量缺少成本款",
                "season_year": "2027秋冬",
                "supplier": "筛选测试供应商",
                "category": "其他",
                "actual_cost": None,
                "status": "pending",
                "source_version_no": 1,
            },
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")

        page = self.wsgi_request(
            app,
            "/workbench?season_year=2027%E7%A7%8B%E5%86%AC&status=waiting",
            cookie=planner_cookie,
        )["body"].decode("utf-8")
        self.assertIn("选择全部筛选资料（1）", page)
        self.assertIn('"suggest": 1', page)

        response = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                {
                    "batch_action": "suggest",
                    "selection_scope": "filtered",
                    "season_year": "2027秋冬",
                    "status": "waiting",
                }
            ).encode("utf-8"),
            cookie=planner_cookie,
        )
        self.assertTrue(response["status"].startswith("302"))
        records = planning_db.list_pricing_records(self.planning_db_path, season_year="2027秋冬")
        self.assertEqual([record["source_product_id"] for record in records], [3090])

    def test_batch_suggest_reinfers_stale_category_after_rule_options_change(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2027春夏", None, 600, 4)
        product = {
            "id": 3099,
            "style_code": "STALE-CATEGORY-3099",
            "style_color": "STALE-CATEGORY-3099-黑",
            "product_name": "历史同步毛衣资料",
            "season_year": "2027春夏",
            "supplier": "历史品类供应商",
            "category": "其他",
            "actual_cost": 100,
            "status": "pending",
            "source_version_no": 1,
        }
        planning_db.upsert_source_products(self.planning_db_path, [product])
        planning_db.save_category_option(self.planning_db_path, "历史品类", "历史同步毛衣", 80)
        product["category"] = "历史品类"
        planning_db.upsert_source_products(self.planning_db_path, [product])
        old_option = next(item for item in planning_db.list_category_options(self.planning_db_path) if item["name"] == "历史品类")
        planning_db.delete_category_option(self.planning_db_path, old_option["id"])
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")

        response = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode({"batch_action": "suggest", "selection_scope": "selected", "suggest_ids": "3099"}).encode("utf-8"),
            cookie=planner_cookie,
        )

        self.assertTrue(response["status"].startswith("302"))
        record = planning_db.list_pricing_records(self.planning_db_path, season_year="2027春夏")[0]
        self.assertEqual(record["category"], "其他")
        self.assertEqual(record["calculated_price"], 399)

    def test_filtered_batch_initial_review_rematches_deleted_category(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2027春夏", None, 600, 4)
        planning_db.save_category_option(self.planning_db_path, "临时旧品类", "临时测试", 80)
        product = {
            "id": 3100,
            "style_code": "FILTERED-REVIEW-STALE-CATEGORY",
            "product_name": "临时测试款",
            "season_year": "2027春夏",
            "supplier": "跨页初审供应商",
            "category": "临时旧品类",
            "actual_cost": 100,
            "status": "pending",
            "source_version_no": 1,
        }
        planning_db.upsert_source_products(self.planning_db_path, [product])
        record = planning_db.create_pricing_record(self.planning_db_path, product, "商品部企划员")
        with planning_db.get_connection(self.planning_db_path) as connection:
            connection.execute(
                "UPDATE pricing_records SET channel = '天猫' WHERE id = ?",
                (record["id"],),
            )
        old_option = next(
            option
            for option in planning_db.list_category_options(self.planning_db_path)
            if option["name"] == "临时旧品类"
        )
        planning_db.delete_category_option(self.planning_db_path, old_option["id"])
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        response = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                {
                    "batch_action": "submit-review",
                    "selection_scope": "filtered",
                    "season_year": "2027春夏",
                    "status": "suggested",
                }
            ).encode("utf-8"),
            cookie=self.login_cookie(app, "planner"),
        )
        self.assertTrue(response["status"].startswith("302"))
        updated = planning_db.get_pricing_record(self.planning_db_path, record["id"])
        self.assertEqual(updated["status"], "review_pending")
        self.assertEqual(updated["category"], "其他")

    def test_filtered_batch_initial_review_reports_unsaved_channels_atomically(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2027春夏", None, 600, 4)
        products = [
            {
                "id": product_id,
                "style_code": f"FILTERED-CHANNEL-{product_id}",
                "product_name": f"待填写渠道款 {product_id}",
                "season_year": "2027春夏",
                "supplier": "跨页初审供应商",
                "category": "其他",
                "actual_cost": 100,
                "status": "pending",
                "source_version_no": 1,
            }
            for product_id in (3110, 3111)
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        planning_db.create_pricing_records(self.planning_db_path, products, "商品部企划员")
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")

        response = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                {
                    "batch_action": "submit-review",
                    "selection_scope": "filtered",
                    "season_year": "2027春夏",
                    "status": "suggested",
                }
            ).encode("utf-8"),
            cookie=self.login_cookie(app, "planner"),
        )

        self.assertTrue(response["status"].startswith("400"))
        self.assertIn("有 2 款尚未填写渠道划分", response["body"].decode("utf-8"))
        self.assertEqual(
            {record["status"] for record in planning_db.list_pricing_records(self.planning_db_path)},
            {"suggested"},
        )

    def test_filtered_selection_runs_each_later_stage_across_pages_and_exports_history(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2028春夏", None, 600, 4)
        products = [
            {
                "id": 3200 + index,
                "style_code": f"FLOW-{index:03d}",
                "style_color": f"FLOW-{index:03d}-黑",
                "product_name": f"全流程跨页毛衣 {index}",
                "season_year": "2028春夏",
                "supplier": "全流程测试供应商",
                "category": "其他",
                "actual_cost": 100,
                "status": "pending",
                "source_version_no": 1,
            }
            for index in range(101)
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        planning_db.create_pricing_records(self.planning_db_path, products, "商品部企划员")
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")
        admin_cookie = self.login_cookie(app, "planning_admin")

        suggested_page = self.wsgi_request(
            app,
            "/workbench?season_year=2028%E6%98%A5%E5%A4%8F&status=suggested",
            cookie=planner_cookie,
        )["body"].decode("utf-8")
        self.assertIn("选择全部筛选资料（101）", suggested_page)
        self.assertIn('"submit-review": 101', suggested_page)
        self.assertEqual(suggested_page.count("name='submit_review_ids'"), 50)
        suggested_export = self.wsgi_request(
            app,
            "/pricing/export.xlsx?season_year=2028%E6%98%A5%E5%A4%8F&status=suggested&selection_scope=filtered",
            cookie=planner_cookie,
        )
        suggested_sheet = load_workbook(io.BytesIO(suggested_export["body"])).active
        self.assertEqual(suggested_sheet.max_row, 102)
        self.assertFalse(suggested_sheet["O2"].protection.locked)
        for row_number in range(2, suggested_sheet.max_row + 1):
            suggested_sheet.cell(row_number, 17).value = "天猫"
        edited_workbook = io.BytesIO()
        suggested_sheet.parent.save(edited_workbook)
        import_body, import_content_type = self.workbook_multipart(edited_workbook.getvalue())
        import_response = self.wsgi_request(
            app,
            "/pricing/import",
            method="POST",
            body=import_body,
            content_type=import_content_type,
            cookie=planner_cookie,
        )
        self.assertTrue(import_response["status"].startswith("302"))
        self.assertEqual(
            {record["channel"] for record in planning_db.list_pricing_records(self.planning_db_path)},
            {"天猫"},
        )
        initial_response = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                {
                    "batch_action": "submit-review",
                    "selection_scope": "filtered",
                    "season_year": "2028春夏",
                    "status": "suggested",
                }
            ).encode("utf-8"),
            cookie=planner_cookie,
        )
        self.assertTrue(initial_response["status"].startswith("302"))
        self.assertEqual(
            sum(record["status"] == "review_pending" for record in planning_db.list_pricing_records(self.planning_db_path)),
            101,
        )

        review_page = self.wsgi_request(
            app,
            "/workbench?season_year=2028%E6%98%A5%E5%A4%8F&status=review_pending",
            cookie=admin_cookie,
        )["body"].decode("utf-8")
        self.assertIn("选择全部筛选资料（101）", review_page)
        self.assertIn('"approve": 101', review_page)
        self.assertEqual(review_page.count("name='approve_ids'"), 50)
        approval_response = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                {
                    "batch_action": "approve",
                    "selection_scope": "filtered",
                    "season_year": "2028春夏",
                    "status": "review_pending",
                }
            ).encode("utf-8"),
            cookie=admin_cookie,
        )
        self.assertTrue(approval_response["status"].startswith("302"))
        self.assertEqual(
            sum(record["status"] == "confirmed" for record in planning_db.list_pricing_records(self.planning_db_path)),
            101,
        )

        confirmed_page = self.wsgi_request(
            app,
            "/workbench?season_year=2028%E6%98%A5%E5%A4%8F&status=confirmed",
            cookie=planner_cookie,
        )["body"].decode("utf-8")
        self.assertIn("选择全部筛选资料（101）", confirmed_page)
        self.assertIn('"publish": 101', confirmed_page)
        self.assertEqual(confirmed_page.count("name='publish_ids'"), 50)
        app.catalog_api_token = "planning-secret"
        publication_result = json.dumps({"status": "published"}).encode("utf-8")
        with patch("planning_center.web.urlopen", side_effect=lambda *args, **kwargs: io.BytesIO(publication_result)) as mocked_urlopen:
            publish_response = self.wsgi_request(
                app,
                "/pricing/batch",
                method="POST",
                body=urlencode(
                    {
                        "batch_action": "publish",
                        "selection_scope": "filtered",
                        "season_year": "2028春夏",
                        "status": "confirmed",
                    }
                ).encode("utf-8"),
                cookie=planner_cookie,
            )
        self.assertTrue(publish_response["status"].startswith("302"))
        self.assertEqual(mocked_urlopen.call_count, 101)
        self.assertEqual(
            sum(record["status"] == "published" for record in planning_db.list_pricing_records(self.planning_db_path)),
            101,
        )

        with patch.object(app, "fetch_catalog_products", return_value=[]):
            published_page = self.wsgi_request(
                app,
                "/workbench?season_year=2028%E6%98%A5%E5%A4%8F&status=published",
                cookie=planner_cookie,
            )["body"].decode("utf-8")
        self.assertIn("选择全部筛选资料（101）", published_page)
        self.assertEqual(published_page.count("name='export_ids'"), 50)
        published_export = self.wsgi_request(
            app,
            "/pricing/export.xlsx?season_year=2028%E6%98%A5%E5%A4%8F&status=published&selection_scope=filtered",
            cookie=planner_cookie,
        )
        published_sheet = load_workbook(io.BytesIO(published_export["body"])).active
        self.assertEqual(published_sheet.max_row, 102)
        self.assertEqual(published_sheet["R2"].value, "已回传")
        self.assertTrue(published_sheet["O2"].protection.locked)
        self.assertTrue(published_sheet.protection.selectLockedCells)

    def test_initial_review_excel_export_and_import_save_drafts_without_advancing(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2027秋冬", None, 600, 4)
        planning_db.save_category_rule(self.planning_db_path, "2027秋冬", "连衣裙", 4.2)
        products = [
            {
                "id": 4101,
                "style_code": "XLSX-001",
                "style_color": "XLSX-001-黑",
                "product_name": "Excel 测试毛衣",
                "season_year": "2027秋冬",
                "supplier": "Excel 测试供应商",
                "category": "其他",
                "actual_cost": 150,
                "status": "pending",
                "source_version_no": 3,
            },
            {
                "id": 4102,
                "style_code": "XLSX-002",
                "style_color": "XLSX-002-白",
                "product_name": "Excel 测试开衫",
                "season_year": "2027秋冬",
                "supplier": "Excel 测试供应商",
                "category": "其他",
                "actual_cost": 120,
                "status": "pending",
                "source_version_no": 1,
            },
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        planning_db.create_pricing_records(self.planning_db_path, products, "商品部企划员")
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")

        export_response = self.wsgi_request(
            app,
            "/pricing/export.xlsx?season_year=2027%E7%A7%8B%E5%86%AC",
            cookie=planner_cookie,
        )
        self.assertTrue(export_response["status"].startswith("200"))
        self.assertIn("planning-initial-review.xlsx", dict(export_response["headers"])["Content-Disposition"])
        workbook = load_workbook(io.BytesIO(export_response["body"]))
        self.assertEqual(workbook.sheetnames, ["待初审资料", "填写说明"])
        sheet = workbook["待初审资料"]
        self.assertEqual(sheet.freeze_panes, "E2")
        self.assertEqual(sheet.auto_filter.ref, "A1:R3")
        self.assertEqual(sheet["A1"].value, "定价记录ID")
        self.assertEqual(sheet["O1"].value, "初审品类")
        self.assertEqual(sheet["P1"].value, "初审上新价")
        self.assertEqual(sheet["Q1"].value, "渠道划分")
        self.assertFalse(sheet["O2"].protection.locked)
        self.assertFalse(sheet["P2"].protection.locked)
        self.assertFalse(sheet["Q2"].protection.locked)
        self.assertTrue(sheet["N2"].protection.locked)
        self.assertTrue(sheet["R2"].protection.locked)
        # Initial-review exports must remain editable in both WPS and
        # LibreOffice.  The yellow columns identify the only fields imported
        # back into the system; the gray source columns are informational.
        self.assertFalse(sheet.protection.sheet)
        self.assertEqual(sheet.sheet_view.selection[0].activeCell, "O2")
        self.assertIn("可修改字段：初审品类、初审上新价、渠道划分", workbook["填写说明"]["C2"].value)
        self.assertIn("兼容 WPS 和 LibreOffice", workbook["填写说明"]["C2"].value)
        instructions = [
            row[0].value
            for row in workbook["填写说明"].iter_rows(min_row=2, min_col=3, max_col=3)
            if row[0].value
        ]
        self.assertTrue(any("不要整行粘贴" in value for value in instructions))
        self.assertEqual(len(sheet.data_validations.dataValidation), 2)

        row_by_style = {sheet.cell(row, 6).value: row for row in range(2, sheet.max_row + 1)}
        target_row = row_by_style["XLSX-001"]
        sheet.cell(target_row, 15).value = "连衣裙"
        sheet.cell(target_row, 16).value = 639
        sheet.cell(target_row, 17).value = "天猫"
        second_row = row_by_style["XLSX-002"]
        sheet.cell(second_row, 17).value = "唯品"
        buffer = io.BytesIO()
        workbook.save(buffer)
        body, content_type = self.workbook_multipart(buffer.getvalue())
        import_response = self.wsgi_request(
            app,
            "/pricing/import",
            method="POST",
            body=body,
            content_type=content_type,
            cookie=planner_cookie,
        )
        self.assertTrue(import_response["status"].startswith("302"))
        self.assertIn("status=suggested", dict(import_response["headers"])["Location"])

        records = {
            record["style_code"]: record
            for record in planning_db.list_pricing_records(self.planning_db_path, season_year="2027秋冬")
        }
        self.assertEqual(records["XLSX-001"]["category"], "连衣裙")
        self.assertEqual(records["XLSX-001"]["calculated_price"], 629)
        self.assertEqual(records["XLSX-001"]["launch_price"], 639)
        self.assertEqual(records["XLSX-001"]["channel"], "天猫")
        self.assertEqual(records["XLSX-001"]["status"], "suggested")
        self.assertEqual(records["XLSX-002"]["channel"], "唯品")
        self.assertEqual(records["XLSX-002"]["status"], "suggested")

    def test_initial_review_excel_import_rejects_stale_source_version_atomically(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2027春", None, 600, 4)
        products = [
            {
                "id": 4201 + index,
                "style_code": f"STALE-{index}",
                "product_name": f"版本测试毛衣 {index}",
                "season_year": "2027春",
                "supplier": "版本测试供应商",
            "category": "其他",
                "actual_cost": 100,
                "status": "pending",
                "source_version_no": 1,
            }
            for index in range(2)
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        planning_db.create_pricing_records(self.planning_db_path, products, "商品部企划员")
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")
        exported = self.wsgi_request(app, "/pricing/export.xlsx", cookie=planner_cookie)
        workbook = load_workbook(io.BytesIO(exported["body"]))
        sheet = workbook["待初审资料"]
        for row_number in range(2, 4):
            sheet.cell(row_number, 16).value = 459
            sheet.cell(row_number, 17).value = "天猫"
        with planning_db.get_connection(self.planning_db_path) as connection:
            connection.execute("UPDATE source_products SET source_version_no = 2 WHERE id = 4202")
        buffer = io.BytesIO()
        workbook.save(buffer)
        body, content_type = self.workbook_multipart(buffer.getvalue())

        response = self.wsgi_request(
            app,
            "/pricing/import",
            method="POST",
            body=body,
            content_type=content_type,
            cookie=planner_cookie,
        )
        self.assertTrue(response["status"].startswith("400"))
        self.assertIn("来源版本已变化", response["body"].decode("utf-8"))
        records = planning_db.list_pricing_records(self.planning_db_path, season_year="2027春")
        self.assertEqual({record["launch_price"] for record in records}, {399})
        self.assertEqual({record["channel"] for record in records}, {""})

    def test_dashboard_uses_new_arrival_review_card_title(self):
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        cookie = self.login_cookie(app, "planner")

        dashboard = self.wsgi_request(app, "/dashboard", cookie=cookie)["body"].decode("utf-8")

        self.assertIn("<h2>上新审核</h2>", dashboard)
        self.assertNotIn("<h2>上新定价</h2>", dashboard)

    def test_pricing_workbench_uses_one_card_and_requires_admin_review(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋冬", None, 700, 4)
        planning_db.save_category_rule(self.planning_db_path, "2026秋冬", "连衣裙", 4.2)
        source = {
            "id": 31,
            "style_code": "M031",
            "style_color": "M031-黑",
            "image_url": "https://example.com/m031.jpg",
            "product_name": "审核测试款",
            "season_year": "2026秋冬",
            "supplier": "供应商审核",
            "category": "其他",
            "actual_cost": 150,
            "tax_included_price": 150,
            "status": "pending",
            "lifecycle_status": "active",
            "source_version_no": 1,
            "updated_at": "",
            "creator_name": "跟单员",
        }
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_login = self.wsgi_request(
            app,
            "/login",
            method="POST",
            body=urlencode({"username": "planner", "password": "demo123"}).encode(),
        )
        planner_cookie = dict(planner_login["headers"])["Set-Cookie"].split(";", 1)[0]
        source["submitted_to_merchandise"] = True
        with patch.object(
            app,
            "fetch_catalog_products",
            return_value={
                "source": "cangbaoge",
                "items": [source],
                "withdrawn_ids": [],
                "image_updates": [],
                "workflow_gate": True,
                "image_gate": True,
                "cost_gate": True,
                "eligibility_gate_version": 1,
            },
        ):
            self.wsgi_request(app, "/sync", method="POST", cookie=planner_cookie)
        suggest = self.wsgi_request(
            app,
            "/pricing/suggest",
            method="POST",
            body=urlencode({"product_id": "31", "category": "其他"}).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(suggest["status"].startswith("302"))
        self.assertTrue(dict(suggest["headers"])["Location"].endswith("#pricing-row-31"))
        record = planning_db.list_pricing_records(self.planning_db_path)[0]
        self.assertEqual(record["status"], "suggested")
        workbench = self.wsgi_request(app, "/workbench", cookie=planner_cookie)["body"].decode("utf-8")
        self.assertEqual(workbench.count("<section class='panel pricing-board'>"), 1)
        self.assertEqual(workbench.count("data-resizable-columns='pricing-v1'"), 1)
        self.assertEqual(workbench.count("class='column-resize-handle'"), 14)
        self.assertEqual(workbench.count("role='separator'"), 14)
        self.assertIn("['ArrowLeft', 'ArrowRight']", workbench)
        self.assertIn("id='pricing-reset-columns'", workbench)
        self.assertIn("planning-workbench-column-widths-v1", workbench)
        self.assertIn("table-layout:fixed", workbench)
        self.assertNotIn("<article class='product-card'>", workbench)
        self.assertIn("系统按商品名称自动判定", workbench)
        self.assertNotIn("name='category' value", workbench)
        headers = [
            ("image", "图片"),
            ("season", "年份季节"),
            ("style", "款号"),
            ("color", "款色"),
            ("product", "商品名称"),
            ("supplier", "供应商"),
            ("cost", "含税成本"),
            ("source-status", "来源状态"),
        ]
        header_positions = [workbench.index(f"data-column-key='{key}'>{header}") for key, header in headers]
        self.assertEqual(header_positions, sorted(header_positions))
        self.assertIn("data-column-key='category'>品类", workbench)
        self.assertIn("data-column-key='rule'>规则计算", workbench)
        self.assertIn("<span class='rule-expression'>4 × 1</span><span class='rule-raw-price'>= 600.0 原始</span>", workbench)
        self.assertIn("data-column-key='price'>测算上新价", workbench)
        self.assertIn("data-column-key='channel'>渠道划分", workbench)
        self.assertIn("data-column-key='workflow'>流程状态与操作", workbench)
        self.assertNotIn("<th>定价状态</th>", workbench)
        self.assertNotIn("<th>初审 / 复核 / 回传</th>", workbench)
        self.assertIn("上新审核工作台", workbench)
        self.assertIn("初审与复核", workbench)
        self.assertNotIn("定价初审与复核", workbench)
        self.assertIn("测算上新价", workbench)
        self.assertIn("初审上新价", workbench)
        self.assertIn("待初审 Excel 请仅修改黄色列：初审品类、初审上新价、渠道划分（兼容 WPS / LibreOffice）", workbench)
        self.assertIn("category-edit-button", workbench)
        self.assertIn(">修改</button>", workbench)
        self.assertIn("class='initial-review-category'", workbench)
        self.assertIn("aria-expanded='false'", workbench)
        self.assertIn("min='1' step='1' inputmode='numeric'", workbench)
        self.assertIn("待初审", workbench)
        self.assertIn("确认并提交复核", workbench)
        self.assertIn("id='pricing-row-31'", workbench)
        self.assertIn("src='/source-products/31/image?v=1'", workbench)
        self.assertIn("loading='lazy'", workbench)
        self.assertIn("class='product-image-zoom'", workbench)
        self.assertIn("data-image-src='/source-products/31/image?v=1'", workbench)
        self.assertIn("class='pricing-source-image'", workbench)
        self.assertIn("data-fallback-src='https://example.com/m031.jpg'", workbench)
        self.assertIn("referrerpolicy='no-referrer'", workbench)
        self.assertIn("aria-label='查看 M031-黑 大图'", workbench)
        self.assertIn("aria-haspopup='dialog'", workbench)
        self.assertEqual(workbench.count("id='product-image-dialog'"), 1)
        self.assertIn("imageDialog.showModal()", workbench)
        self.assertIn("imageDialogClose?.addEventListener('click'", workbench)
        self.assertIn("if (event.target === imageDialog) imageDialog.close()", workbench)
        self.assertIn("if (event.key === 'Escape' && imageDialog?.open) imageDialog.close()", workbench)
        self.assertIn("image.dataset.fallbackTried = '1'", workbench)
        self.assertIn("cursor:zoom-in", workbench)
        self.assertIn(".product-image-dialog::backdrop", workbench)
        self.assertIn("planning-workbench-scroll", workbench)
        self.assertIn("tableWrap.scrollLeft", workbench)
        self.assertIn("history.replaceState", workbench)
        self.assertIn("setTimeout(restoreScroll, 150)", workbench)
        self.assertNotIn("PRICING RECORDS", workbench)

        rejected_fractional = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/submit-review",
            method="POST",
            body=urlencode({"launch_price": "589.5", "category": "其他", "channel": "天猫"}).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(rejected_fractional["status"].startswith("400"))
        self.assertIn("必须是大于 0 的整数", rejected_fractional["body"].decode("utf-8"))
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, record["id"])["status"], "suggested")

        submitted = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/submit-review",
            method="POST",
            body=urlencode({"launch_price": "2539", "category": "其他", "channel": "天猫"}).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(submitted["status"].startswith("302"))
        self.assertTrue(dict(submitted["headers"])["Location"].endswith("#pricing-row-31"))
        submitted_record = planning_db.get_pricing_record(self.planning_db_path, record["id"])
        self.assertEqual(submitted_record["status"], "review_pending")
        self.assertEqual(submitted_record["calculated_price"], 599)
        self.assertEqual(submitted_record["launch_price"], 2539)
        self.assertEqual(submitted_record["channel"], "天猫")
        denied = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/approve",
            method="POST",
            body=urlencode({"launch_price": "579"}).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(denied["status"].startswith("403"))

        admin_login = self.wsgi_request(
            app,
            "/login",
            method="POST",
            body=urlencode({"username": "planning_admin", "password": "demo123"}).encode(),
        )
        admin_cookie = dict(admin_login["headers"])["Set-Cookie"].split(";", 1)[0]
        review_page = self.wsgi_request(app, "/workbench?status=review_pending", cookie=admin_cookie)["body"].decode("utf-8")
        self.assertEqual(review_page.count("复核上新价"), 1)
        self.assertEqual(review_page.count("复核渠道"), 1)
        self.assertEqual(review_page.count("<input name='launch_price'"), 1)
        self.assertEqual(review_page.count("<select name='channel'"), 1)
        self.assertIn("min='1' step='1' inputmode='numeric'", review_page)
        self.assertIn("data-saved-value='2539'", review_page)
        self.assertIn("data-saved-value='天猫'", review_page)
        self.assertIn("修改保存", review_page)
        self.assertIn("复核通过", review_page)
        self.assertIn("grid-template-columns:max-content max-content", review_page)
        self.assertIn(".review-approve-button{grid-column:2;justify-self:end}", review_page)
        rejected_review_fractional = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/approve",
            method="POST",
            body=urlencode({"launch_price": "2539.5"}).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(rejected_review_fractional["status"].startswith("400"))
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, record["id"])["status"], "review_pending")
        rejected_unsaved_review = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/approve",
            method="POST",
            body=urlencode({"launch_price": "579", "channel": "天猫"}).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(rejected_unsaved_review["status"].startswith("400"))
        self.assertIn("请先点击“修改保存”", rejected_unsaved_review["body"].decode("utf-8"))
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, record["id"])["status"], "review_pending")
        rejected_unsaved_channel = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/approve",
            method="POST",
            body=urlencode({"launch_price": "2539", "channel": "唯品"}).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(rejected_unsaved_channel["status"].startswith("400"))
        self.assertIn("上新价或渠道已修改", rejected_unsaved_channel["body"].decode("utf-8"))
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, record["id"])["channel"], "天猫")
        saved_review = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/review-save",
            method="POST",
            body=urlencode({"launch_price": "579", "channel": "唯品"}).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(saved_review["status"].startswith("302"))
        saved_record = planning_db.get_pricing_record(self.planning_db_path, record["id"])
        self.assertEqual(saved_record["status"], "review_pending")
        self.assertEqual(saved_record["launch_price"], 579)
        self.assertEqual(saved_record["channel"], "唯品")
        approved = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/approve",
            method="POST",
            body=urlencode({"launch_price": "579", "channel": "唯品"}).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(approved["status"].startswith("302"))
        approved_record = planning_db.get_pricing_record(self.planning_db_path, record["id"])
        self.assertEqual(approved_record["status"], "confirmed")
        self.assertEqual(approved_record["launch_price"], 579)
        self.assertEqual(approved_record["channel"], "唯品")

        app.catalog_api_token = "token"
        admin_confirmed_page = self.wsgi_request(
            app,
            "/workbench?status=confirmed",
            cookie=admin_cookie,
        )["body"].decode("utf-8")
        self.assertIn("复核已通过，待商品部回传", admin_confirmed_page)
        self.assertNotIn(f"action='/pricing/{record['id']}/publish'", admin_confirmed_page)
        self.assertNotIn("action='/sync'", admin_confirmed_page)
        denied_publish = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/publish",
            method="POST",
            cookie=admin_cookie,
        )
        self.assertTrue(denied_publish["status"].startswith("403"))
        denied_sync = self.wsgi_request(app, "/sync", method="POST", cookie=admin_cookie)
        self.assertTrue(denied_sync["status"].startswith("403"))

        with patch.object(app, "fetch_catalog_products", return_value=[]):
            planner_confirmed_page = self.wsgi_request(
                app,
                "/workbench?status=confirmed",
                cookie=planner_cookie,
            )["body"].decode("utf-8")
        self.assertIn(f"action='/pricing/{record['id']}/publish'", planner_confirmed_page)
        self.assertIn("回传藏宝阁", planner_confirmed_page)
        self.assertIn("action='/sync'", planner_confirmed_page)
        publication_result = io.BytesIO(json.dumps({"status": "published"}).encode("utf-8"))
        with patch("planning_center.web.urlopen", return_value=publication_result) as mocked_urlopen:
            published = self.wsgi_request(
                app,
                f"/pricing/{record['id']}/publish",
                method="POST",
                cookie=planner_cookie,
            )
        self.assertTrue(published["status"].startswith("302"))
        publication_payload = json.loads(mocked_urlopen.call_args.args[0].data.decode("utf-8"))
        self.assertEqual(publication_payload["operator_name"], "商品部企划员")
        self.assertEqual(publication_payload["launch_channel"], "唯品")
        product_field_keys = {field.key for field in catalog_db.PRODUCT_FIELDS}
        self.assertEqual(
            set(publication_payload) & product_field_keys,
            catalog_db.PLANNING_PRODUCT_MUTABLE_FIELD_KEYS,
        )
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, record["id"])["status"], "published")

    def test_confirmed_stage_is_export_only_and_can_reopen_for_new_review(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋冬", None, 700, 4)
        product = {
            "id": 61,
            "style_code": "PREPUBLISH-061",
            "style_color": "PREPUBLISH-061-黑",
            "image_url": "https://example.com/prepublish-061.jpg",
            "product_name": "回传前修改测试款",
            "season_year": "2026秋冬",
            "supplier": "回传前修改供应商",
            "category": "其他",
            "actual_cost": 150,
            "status": "pending",
            "lifecycle_status": "active",
            "source_version_no": 1,
        }
        planning_db.upsert_source_products(self.planning_db_path, [product])
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")
        admin_cookie = self.login_cookie(app, "planning_admin")
        record = planning_db.create_pricing_record(self.planning_db_path, product, "商品部企划员")
        submitted = planning_db.submit_pricing_for_review(
            self.planning_db_path,
            record["id"],
            599,
            "商品部企划员",
            "其他",
            "天猫",
        )
        planning_db.approve_pricing_record(
            self.planning_db_path,
            submitted["id"],
            599,
            "天猫",
            "企划管理员",
        )

        with patch.object(app, "fetch_catalog_products", return_value=[]):
            confirmed_page = self.wsgi_request(
                app,
                "/workbench?status=confirmed",
                cookie=planner_cookie,
            )["body"].decode("utf-8")
        self.assertIn(f"action='/pricing/{record['id']}/reopen'", confirmed_page)
        self.assertIn(f"action='/pricing/{record['id']}/publish'", confirmed_page)
        self.assertIn("复核通过阶段仅支持导出二次检查", confirmed_page)
        self.assertNotIn("action='/pricing/import'", confirmed_page)

        exported = self.wsgi_request(
            app,
            "/pricing/export.xlsx?status=confirmed&selection_scope=filtered",
            cookie=planner_cookie,
        )
        self.assertTrue(exported["status"].startswith("200"))
        workbook = load_workbook(io.BytesIO(exported["body"]))
        sheet = workbook.active
        self.assertEqual(workbook.sheetnames, ["上新审核资料", "填写说明"])
        self.assertEqual(sheet["R2"].value, "复核通过，待回传")
        self.assertTrue(sheet["O2"].protection.locked)
        self.assertTrue(sheet["P2"].protection.locked)
        self.assertTrue(sheet["Q2"].protection.locked)
        self.assertTrue(sheet.protection.sheet)

        # A confirmed workbook cannot be imported, even if a caller bypasses
        # the hidden UI and posts a valid .xlsx directly.
        body, content_type = self.workbook_multipart(exported["body"])
        rejected_import = self.wsgi_request(
            app,
            "/pricing/import",
            method="POST",
            body=body,
            content_type=content_type,
            cookie=planner_cookie,
        )
        self.assertTrue(rejected_import["status"].startswith("400"))
        self.assertIn("只允许导出检查，不能导入 Excel", rejected_import["body"].decode("utf-8"))

        denied_reopen = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/reopen",
            method="POST",
            cookie=admin_cookie,
        )
        self.assertTrue(denied_reopen["status"].startswith("403"))
        reopened = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/reopen",
            method="POST",
            cookie=planner_cookie,
        )
        self.assertTrue(reopened["status"].startswith("302"))
        self.assertIn("status=suggested", dict(reopened["headers"])["Location"])
        reopened_record = planning_db.get_pricing_record(self.planning_db_path, record["id"])
        self.assertEqual(reopened_record["status"], "suggested")
        self.assertIsNone(reopened_record["confirmed_at"])
        self.assertEqual(len(planning_db.list_pricing_records(self.planning_db_path)), 1)

    def test_batch_initial_review_approval_and_publication(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋冬", None, 700, 4)
        products = [
            {
                "id": product_id,
                "style_code": f"B{product_id}",
                "style_color": f"B{product_id}-黑",
                "product_name": f"批量测试款 {product_id}",
                "season_year": "2026秋冬",
                "supplier": "批量供应商",
                "category": "其他",
                "actual_cost": 150,
                "status": "pending",
                "source_version_no": 1,
            }
            for product_id in (71, 72)
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        planner_cookie = self.login_cookie(app, "planner")
        admin_cookie = self.login_cookie(app, "planning_admin")

        waiting_page = self.wsgi_request(app, "/workbench?status=waiting", cookie=planner_cookie)["body"].decode("utf-8")
        self.assertIn("批量测算上新价", waiting_page)
        self.assertNotIn("批量生成测算上新价", waiting_page)
        self.assertEqual(waiting_page.count("name='suggest_ids'"), 2)
        self.assertEqual(waiting_page.count(">生成测算上新价</button>"), 2)
        no_suggest_selection = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode({"batch_action": "suggest"}).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(no_suggest_selection["status"].startswith("400"))
        denied_suggest = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode([("batch_action", "suggest"), ("suggest_ids", "71")]).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(denied_suggest["status"].startswith("403"))
        batch_suggest = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                [
                    ("batch_action", "suggest"),
                    ("suggest_ids", "71"),
                    ("suggest_ids", "72"),
                ]
            ).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(batch_suggest["status"].startswith("302"))
        self.assertIn("status=suggested", dict(batch_suggest["headers"])["Location"])
        records_by_source = {
            int(record["source_product_id"]): record
            for record in planning_db.list_pricing_records(self.planning_db_path)
        }
        first, second = records_by_source[71], records_by_source[72]
        self.assertEqual(first["calculated_price"], 599)
        self.assertEqual(second["calculated_price"], 599)

        initial_page = self.wsgi_request(app, "/workbench?status=suggested", cookie=planner_cookie)["body"].decode("utf-8")
        self.assertIn("pricing-select-all", initial_page)
        self.assertIn("批量测算上新价", initial_page)
        self.assertIn("批量初审提交", initial_page)
        self.assertIn("批量回传藏宝阁", initial_page)
        self.assertIn(".pricing-select-cell{position:sticky;left:0", initial_page)
        self.assertIn(".pricing-table thead .pricing-select-cell{z-index:2", initial_page)
        self.assertIn(".pricing-table .pricing-image-cell{position:sticky;left:44px", initial_page)
        self.assertIn(".pricing-table thead th:nth-child(2){position:sticky;left:44px;z-index:3", initial_page)
        self.assertIn(".pricing-table .pricing-workflow-cell{vertical-align:middle}", initial_page)
        self.assertIn(".pricing-table .workflow-actions>form:only-child{margin-bottom:0}", initial_page)
        self.assertEqual(initial_page.count("name='submit_review_ids'"), 2)
        no_selection = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode({"batch_action": "submit-review"}).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(no_selection["status"].startswith("400"))
        batch_initial = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                [
                    ("batch_action", "submit-review"),
                    ("submit_review_ids", str(first["id"])),
                    ("submit_review_ids", str(second["id"])),
                    (f"launch_price_{first['id']}", "609"),
                    (f"launch_price_{second['id']}", "619"),
                    (f"category_{first['id']}", "其他"),
                    (f"category_{second['id']}", "其他"),
                    (f"channel_{first['id']}", "天猫"),
                    (f"channel_{second['id']}", "唯品"),
                ]
            ).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(batch_initial["status"].startswith("302"))
        self.assertIn("status=review_pending", dict(batch_initial["headers"])["Location"])
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, first["id"])["launch_price"], 609)
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, second["id"])["launch_price"], 619)
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, first["id"])["channel"], "天猫")
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, second["id"])["channel"], "唯品")

        denied_approval = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode([("batch_action", "approve"), ("approve_ids", str(first["id"])), (f"review_price_{first['id']}", "609")]).encode(),
            cookie=planner_cookie,
        )
        self.assertTrue(denied_approval["status"].startswith("403"))
        review_page = self.wsgi_request(app, "/workbench?status=review_pending", cookie=admin_cookie)["body"].decode("utf-8")
        self.assertIn("批量复核通过", review_page)
        self.assertEqual(review_page.count("name='approve_ids'"), 2)
        missing_price = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode([("batch_action", "approve"), ("approve_ids", str(first["id"]))]).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(missing_price["status"].startswith("400"))
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, first["id"])["status"], "review_pending")
        unsaved_price = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                [
                    ("batch_action", "approve"),
                    ("approve_ids", str(first["id"])),
                    (f"review_price_{first['id']}", "629"),
                    (f"review_channel_{first['id']}", "天猫"),
                ]
            ).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(unsaved_price["status"].startswith("400"))
        self.assertIn("请先点击“修改保存”", unsaved_price["body"].decode("utf-8"))
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, first["id"])["status"], "review_pending")
        unsaved_channel = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                [
                    ("batch_action", "approve"),
                    ("approve_ids", str(first["id"])),
                    (f"review_price_{first['id']}", "609"),
                    (f"review_channel_{first['id']}", "唯品"),
                ]
            ).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(unsaved_channel["status"].startswith("400"))
        self.assertIn("上新价或渠道已修改", unsaved_channel["body"].decode("utf-8"))
        batch_approval = self.wsgi_request(
            app,
            "/pricing/batch",
            method="POST",
            body=urlencode(
                [
                    ("batch_action", "approve"),
                    ("approve_ids", str(first["id"])),
                    ("approve_ids", str(second["id"])),
                    (f"review_price_{first['id']}", "609"),
                    (f"review_price_{second['id']}", "619"),
                    (f"review_channel_{first['id']}", "天猫"),
                    (f"review_channel_{second['id']}", "唯品"),
                ]
            ).encode(),
            cookie=admin_cookie,
        )
        self.assertTrue(batch_approval["status"].startswith("302"))
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, first["id"])["status"], "confirmed")
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, second["id"])["status"], "confirmed")

        app.catalog_api_token = "token"
        with patch.object(app, "fetch_catalog_products", return_value=[]):
            publish_page = self.wsgi_request(app, "/workbench?status=confirmed", cookie=planner_cookie)["body"].decode("utf-8")
        self.assertEqual(publish_page.count("name='publish_ids'"), 2)
        publication_responses = [
            io.BytesIO(json.dumps({"status": "published"}).encode("utf-8")),
            io.BytesIO(json.dumps({"status": "published"}).encode("utf-8")),
        ]
        with patch("planning_center.web.urlopen", side_effect=publication_responses) as mocked_urlopen:
            batch_publish = self.wsgi_request(
                app,
                "/pricing/batch",
                method="POST",
                body=urlencode(
                    [
                        ("batch_action", "publish"),
                        ("publish_ids", str(first["id"])),
                        ("publish_ids", str(second["id"])),
                    ]
                ).encode(),
                cookie=planner_cookie,
            )
        self.assertTrue(batch_publish["status"].startswith("302"))
        self.assertEqual(mocked_urlopen.call_count, 2)
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, first["id"])["status"], "published")
        self.assertEqual(planning_db.get_pricing_record(self.planning_db_path, second["id"])["status"], "published")

    def test_initial_review_defaults_to_calculated_price_when_price_is_omitted(self):
        planning_db.save_category_cost_rule(self.planning_db_path, "2026秋冬", None, 700, 4)
        product = {
            "id": 35,
            "style_code": "M035",
            "product_name": "默认测算价",
            "season_year": "2026秋冬",
            "supplier": "供应商",
            "category": "其他",
            "actual_cost": 150,
            "source_version_no": 1,
        }
        planning_db.upsert_source_products(self.planning_db_path, [product])
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        user = planning_db.authenticate_user(self.planning_db_path, "planner", "demo123")
        record = planning_db.create_pricing_record(self.planning_db_path, product, user["display_name"])
        response = self.wsgi_request(
            app,
            f"/pricing/{record['id']}/submit-review",
            method="POST",
            body=urlencode({"category": "其他", "channel": "同款"}).encode(),
            cookie=self.login_cookie(app, "planner"),
        )
        self.assertTrue(response["status"].startswith("302"))
        submitted = planning_db.get_pricing_record(self.planning_db_path, record["id"])
        self.assertEqual(submitted["calculated_price"], 599)
        self.assertEqual(submitted["launch_price"], 599)

    def test_pricing_workbench_keeps_multiple_styles_in_one_board(self):
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        products = [
            {
                "id": 41,
                "style_code": "M041",
                "style_color": "M041-黑",
                "product_name": "多款测试一",
                "season_year": "2026秋冬",
                "supplier": "供应商一",
                "actual_cost": 150,
                "status": "pending",
                "source_version_no": 1,
            },
            {
                "id": 42,
                "style_code": "M042",
                "style_color": "M042-白",
                "product_name": "多款测试二",
                "season_year": "2026秋冬",
                "supplier": "供应商二",
                "actual_cost": 620,
                "status": "pending",
                "source_version_no": 1,
            },
        ]
        planning_db.upsert_source_products(self.planning_db_path, products)
        user = planning_db.authenticate_user(self.planning_db_path, "planner", "demo123")
        workbench = app.render_workbench(user, {})
        self.assertEqual(workbench.count("<section class='panel pricing-board'>"), 1)
        self.assertEqual(workbench.count("<tr"), 3)
        self.assertIn("M041", workbench)
        self.assertIn("M042", workbench)

    def test_category_planning_phase_two_entry_is_visible(self):
        app = PlanningApplication(self.planning_db_path, "http://catalog.test")
        login = self.wsgi_request(
            app,
            "/login",
            method="POST",
            body=urlencode({"username": "planner", "password": "demo123"}).encode(),
        )
        cookie = dict(login["headers"])["Set-Cookie"].split(";", 1)[0]

        dashboard = self.wsgi_request(app, "/dashboard", cookie=cookie)
        dashboard_html = dashboard["body"].decode("utf-8")
        self.assertIn("品类企划", dashboard_html)
        self.assertIn("/category-planning", dashboard_html)

        category_planning = self.wsgi_request(app, "/category-planning", cookie=cookie)
        category_html = category_planning["body"].decode("utf-8")
        self.assertTrue(category_planning["status"].startswith("200"))
        self.assertIn("年份季节", category_html)
        self.assertIn("品类组合", category_html)
        self.assertIn("SKU 数", category_html)
        self.assertIn("第二阶段", category_html)


if __name__ == "__main__":
    unittest.main()
