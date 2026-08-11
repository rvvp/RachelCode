from __future__ import annotations

import io
import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import patch
from urllib.parse import parse_qs, urlencode, urlsplit
from wsgiref.util import setup_testing_defaults

from openpyxl import Workbook, load_workbook

from replenishment_center import db
from replenishment_center.browser_capture import browser_paths, sha256_file
from replenishment_center.browser_import import normalize_browser_reports
from replenishment_center.excel import data_template_bytes, import_data_workbook, plan_workbook_bytes
from replenishment_center.engine import build_replenishment_items
from replenishment_center.scheduler import SHANGHAI, run_due_jobs
from replenishment_center.tmall import calculate_sign as calculate_tmall_sign
from replenishment_center.tmall import test_configured_connection as test_tmall_connection
from replenishment_center.vipshop import (
    VipshopConfig,
    build_authorization_url,
    calculate_sign,
    collect_store_data,
    exchange_authorization_code,
)
from replenishment_center.vipshop import test_configured_connection as test_vipshop_connection
from replenishment_center.web import ReplenishmentApplication


class ReplenishmentTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "replenishment.db"
        db.init_db(self.db_path)
        self.app = ReplenishmentApplication(self.db_path)

    def tearDown(self):
        self.temp_dir.cleanup()

    def request(self, path="/", method="GET", body=b"", cookie="", content_type="application/x-www-form-urlencoded"):
        environ = {}
        setup_testing_defaults(environ)
        if "?" in path:
            path, query = path.split("?", 1)
            environ["QUERY_STRING"] = query
        environ["PATH_INFO"] = path
        environ["REQUEST_METHOD"] = method
        environ["CONTENT_LENGTH"] = str(len(body))
        environ["CONTENT_TYPE"] = content_type
        environ["wsgi.input"] = io.BytesIO(body)
        if cookie:
            environ["HTTP_COOKIE"] = cookie
        captured = {}

        def start_response(status, headers):
            captured["status"] = status
            captured["headers"] = headers

        captured["body"] = b"".join(self.app(environ, start_response))
        return captured

    def login(self, username="merch"):
        response = self.request(
            "/login",
            method="POST",
            body=urlencode({"username": username, "password": "demo123"}).encode(),
        )
        self.assertEqual(response["status"], "302 Found")
        return dict(response["headers"])["Set-Cookie"].split(";", 1)[0]

    def test_default_store_schedule_and_dashboard_risks(self):
        settings = db.get_settings(self.db_path)
        self.assertEqual(settings["store_name"], "马天奴唯品会")
        self.assertEqual(settings["schedule_weekdays"], [2, 5])
        self.assertEqual(settings["schedule_time"], "10:00")
        self.assertEqual(
            (
                settings["min_sales_7"],
                settings["min_sales_14"],
                settings["min_consecutive_sales_days"],
                settings["max_coverage_days"],
            ),
            (5, 10, 3, 14),
        )
        dashboard = db.dashboard_data(self.db_path)
        self.assertEqual(dashboard["goods_count"], len(dashboard["groups"]))
        self.assertGreaterEqual(dashboard["critical_styles"], 1)
        self.assertGreaterEqual(dashboard["warning_styles"], 1)
        self.assertGreaterEqual(dashboard["broken_core"], 1)

    def test_suggestions_are_size_level_and_pack_rounded(self):
        plan = db.list_plans(self.db_path)[0]
        items = db.get_plan_items(self.db_path, plan["id"])
        self.assertTrue(items)
        self.assertTrue(all(item["suggested_qty"] % item["pack_size"] == 0 for item in items))
        groups = db.grouped_plan_items(items)
        self.assertEqual(sum(group["confirmed_qty"] for group in groups), sum(item["confirmed_qty"] for item in items))
        self.assertTrue(any(group["broken_core_count"] for group in groups))

    def test_goods_match_either_sales_condition_and_the_shared_coverage_limit(self):
        as_of = datetime(2026, 7, 22).date()
        goods = [
            # 条件1独有：7天5件、14天10件，但最多只连续销售2天。
            (1, "G-C1", "红色", [0, 1, 3, 4, 6, 7, 9, 10, 12, 13], 0),
            # 条件2独有：连续3天有销量，但累计销量未达到条件1。
            (2, "G-C2", "黄色", [0, 1, 2], 0),
            # 同时命中条件1和条件2。
            (3, "G-BOTH", "蓝色", list(range(10)), 0),
            # 两个销量条件均未命中。
            (4, "G-NONE", "白色", [0, 2, 4, 6], 0),
            # 销量条件命中，但库存支撑超过共同门槛。
            (5, "G-STOCK", "绿色", list(range(10)), 100),
        ]
        skus = [
            {
                "id": sku_id, "style_code": "STYLE-1", "outer_sku_id": goods_code,
                "style_name": "测试款", "color_name": color, "size_name": "M", "category": "",
                "supplier": "", "core_size": 1, "default_size_share": 1, "demand_factor": 1,
                "pack_size": 1, "moq": 0,
            }
            for sku_id, goods_code, color, _, _ in goods
        ]
        sales = [
            {"sku_id": sku_id, "sale_date": (as_of - timedelta(days=offset)).isoformat(), "net_units": 1}
            for sku_id, _, _, sale_offsets, _ in goods
            for offset in sale_offsets
        ]
        inventory = {
            sku_id: {"on_hand": stock, "locked": 0, "defective": 0, "inbound": 0, "inbound_date": ""}
            for sku_id, _, _, _, stock in goods
        }
        items = build_replenishment_items(
            skus, sales, inventory, as_of=as_of, target_days=45, safety_days=7,
            min_sales_7=5, min_sales_14=10, min_consecutive_sales_days=3, max_coverage_days=14,
        )
        self.assertEqual({item["outer_sku_id"] for item in items}, {"G-C1", "G-C2", "G-BOTH"})
        reasons = {item["outer_sku_id"]: item["selection_reason"] for item in items}
        self.assertEqual(reasons["G-C1"], "condition_1")
        self.assertEqual(reasons["G-C2"], "condition_2")
        self.assertEqual(reasons["G-BOTH"], "condition_1,condition_2")
        streaks = {item["outer_sku_id"]: item["consecutive_sales_days"] for item in items}
        self.assertEqual(streaks, {"G-C1": 2, "G-C2": 3, "G-BOTH": 10})
        styles = db.grouped_plan_styles(items)
        self.assertEqual(len(styles), 1)
        self.assertEqual(styles[0]["goods_count"], 3)

    def test_consecutive_sales_are_aggregated_across_sizes_of_one_goods(self):
        as_of = datetime(2026, 7, 22).date()
        skus = [
            {
                "id": sku_id, "style_code": "STYLE-2", "outer_sku_id": "G-SIZES",
                "style_name": "跨尺码测试款", "color_name": "黑色", "size_name": size_name, "category": "",
                "supplier": "", "core_size": 1, "default_size_share": 0.5, "demand_factor": 1,
                "pack_size": 1, "moq": 0,
            }
            for sku_id, size_name in ((1, "S"), (2, "M"))
        ]
        sales = [
            {"sku_id": 1, "sale_date": as_of.isoformat(), "net_units": 1},
            {"sku_id": 2, "sale_date": (as_of - timedelta(days=1)).isoformat(), "net_units": 1},
            {"sku_id": 1, "sale_date": (as_of - timedelta(days=2)).isoformat(), "net_units": 1},
        ]
        inventory = {
            sku_id: {"on_hand": 0, "locked": 0, "defective": 0, "inbound": 0, "inbound_date": ""}
            for sku_id in (1, 2)
        }
        items = build_replenishment_items(
            skus, sales, inventory, as_of=as_of, target_days=45, safety_days=7,
            min_sales_7=5, min_sales_14=10, min_consecutive_sales_days=3, max_coverage_days=14,
        )
        self.assertEqual(len(items), 2)
        self.assertTrue(all(item["selection_reason"] == "condition_2" for item in items))
        self.assertTrue(all(item["consecutive_sales_days"] == 3 for item in items))

    def test_merchandise_adjustment_requires_reason_and_flows_to_followup(self):
        plan = db.list_plans(self.db_path)[0]
        items = db.get_plan_items(self.db_path, plan["id"])
        user = db.authenticate(self.db_path, "merch", "demo123")
        first = items[0]
        quantities = {item["id"]: item["suggested_qty"] for item in items}
        quantities[first["id"]] += first["pack_size"]
        with self.assertRaisesRegex(ValueError, "调整原因"):
            db.save_merchandise_adjustments(self.db_path, plan["id"], quantities, {}, user["id"])
        db.save_merchandise_adjustments(
            self.db_path,
            plan["id"],
            quantities,
            {first["id"]: "活动备货"},
            user["id"],
        )
        db.submit_to_followup(self.db_path, plan["id"], user["id"])
        updated = db.get_plan(self.db_path, plan["id"])
        self.assertEqual(updated["status"], "followup_pending")
        self.assertTrue(all(item["followup_qty"] is not None for item in db.get_plan_items(self.db_path, plan["id"])))

    def test_followup_can_complete_plan(self):
        plan = db.list_plans(self.db_path)[0]
        merchandise = db.authenticate(self.db_path, "merch", "demo123")
        db.submit_to_followup(self.db_path, plan["id"], merchandise["id"])
        followup = db.authenticate(self.db_path, "followup", "demo123")
        items = db.get_plan_items(self.db_path, plan["id"])
        payloads = {
            item["id"]: {
                "followup_qty": item["confirmed_qty"],
                "expected_order_date": "2026-07-23",
                "expected_arrival_date": "2026-08-03",
                "followup_status": "ordered",
                "followup_note": "PO-TEST",
            }
            for item in items
        }
        db.save_followup_response(self.db_path, plan["id"], payloads, followup["id"], complete=True)
        self.assertEqual(db.get_plan(self.db_path, plan["id"])["status"], "completed")

    def test_schedule_runs_only_after_configured_time_once(self):
        before = datetime(2026, 7, 28, 9, 59, tzinfo=SHANGHAI)
        due = datetime(2026, 7, 28, 10, 0, tzinfo=SHANGHAI)
        self.assertIsNone(run_due_jobs(self.db_path, before))
        plan_id = run_due_jobs(self.db_path, due)
        self.assertIsInstance(plan_id, int)
        self.assertIsNone(run_due_jobs(self.db_path, due))
        self.assertIn("2026-07-28", db.get_settings(self.db_path)["last_schedule_key"])

    def test_browser_schedule_does_not_generate_a_plan_from_stale_data(self):
        due = datetime(2026, 7, 28, 10, 0, tzinfo=SHANGHAI)
        with db.get_connection(self.db_path) as connection:
            connection.execute(
                "UPDATE settings SET data_source_mode = 'browser', last_schedule_key = '' WHERE id = 1"
            )
        plan_count = len(db.list_plans(self.db_path))
        with patch(
            "replenishment_center.scheduler.launch_dedicated_browser",
            return_value={
                "loginRequired": True,
                "sessionStatus": "login_required",
                "current": {"url": "https://vis.vip.com/login.php", "title": "登录"},
                "message": "唯品后台需要重新登录。",
            },
        ):
            self.assertIsNone(run_due_jobs(self.db_path, due))
        self.assertEqual(len(db.list_plans(self.db_path)), plan_count)
        sync = db.latest_sync(self.db_path)
        self.assertEqual((sync["source"], sync["status"]), ("vipshop_browser_schedule", "failed"))
        self.assertIn("登录已失效", sync["message"])
        merchandise = db.authenticate(self.db_path, "merch", "demo123")
        notifications = db.unread_notifications(self.db_path, merchandise["id"])
        self.assertTrue(any(item["title"] == "唯品数据更新未完成" for item in notifications))
        self.assertIn("2026-07-28", db.get_settings(self.db_path)["last_schedule_key"])

    def test_browser_schedule_exports_previous_14_complete_days_and_imports_plan(self):
        due = datetime(2026, 7, 28, 10, 0, tzinfo=SHANGHAI)
        with db.get_connection(self.db_path) as connection:
            connection.execute(
                "UPDATE settings SET data_source_mode = 'browser', schedule_time = '10:00', last_schedule_key = '' WHERE id = 1"
            )
        session = {
            "loginRequired": False,
            "sessionStatus": "session_present",
            "current": {
                "url": "https://compass.vip.com/frontend/index.html#/product/details",
                "title": "魔方罗盘 - 商品明细",
            },
            "message": "唯品后台会话可用。",
        }
        with patch("replenishment_center.scheduler.launch_dedicated_browser", return_value=session), patch(
            "replenishment_center.scheduler.run_browser_worker", return_value=session
        ) as export_worker, patch(
            "replenishment_center.scheduler.db.create_browser_capture_job", side_effect=[101, 102]
        ), patch(
            "replenishment_center.scheduler.db.process_browser_capture_jobs", return_value=[101, 102]
        ), patch(
            "replenishment_center.scheduler.db.import_latest_browser_reports",
            return_value={"plan_id": 88, "job_ids": [101, 102]},
        ):
            self.assertEqual(run_due_jobs(self.db_path, due), 88)
        call = export_worker.call_args.kwargs
        self.assertEqual(call["action"], "export_product_detail")
        self.assertEqual((call["start_date"], call["end_date"]), ("2026-07-14", "2026-07-27"))
        self.assertEqual(call["brand"], "马天奴")
        self.assertEqual(db.get_settings(self.db_path)["last_schedule_key"], "2026-07-28@10:00")

    def test_schedule_settings_can_be_changed(self):
        user = db.authenticate(self.db_path, "merch", "demo123")
        db.update_settings(
            self.db_path,
            {
                "schedule_weekdays": [1, 3, 6],
                "schedule_time": "09:15",
                "target_days": 60,
                "safety_days": 10,
                "auto_generate": True,
                "min_sales_7": 6,
                "min_sales_14": 12,
                "min_consecutive_sales_days": 4,
                "max_coverage_days": 18,
            },
            user["id"],
        )
        settings = db.get_settings(self.db_path)
        self.assertEqual(settings["schedule_weekdays"], [1, 3, 6])
        self.assertEqual(settings["schedule_time"], "09:15")
        self.assertEqual(settings["target_days"], 60)
        self.assertEqual(
            (
                settings["min_sales_7"],
                settings["min_sales_14"],
                settings["min_consecutive_sales_days"],
                settings["max_coverage_days"],
            ),
            (6, 12, 4, 18),
        )

    def test_excel_template_import_and_plan_export(self):
        template = data_template_bytes()
        workbook = load_workbook(io.BytesIO(template), data_only=True)
        self.assertEqual(set(workbook.sheetnames), {"SKU资料", "销售数据", "库存及在途"})
        user = db.authenticate(self.db_path, "merch", "demo123")
        counts = import_data_workbook(self.db_path, io.BytesIO(template), user["id"])
        self.assertEqual(counts, {"sku": 1, "sales": 1, "inventory": 1})
        plan = db.get_plan(self.db_path, 1)
        export = plan_workbook_bytes(plan, db.get_plan_items(self.db_path, 1))
        exported = load_workbook(io.BytesIO(export), data_only=True)
        self.assertEqual(exported.sheetnames, ["补货明细", "计算口径"])
        detail_headers = [cell.value for cell in exported["补货明细"][1]]
        self.assertIn("命中条件", detail_headers)
        self.assertIn("连续有销量天数", detail_headers)
        calculation_rows = dict(exported["计算口径"].iter_rows(min_row=2, values_only=True))
        self.assertIn("条件1 或 条件2", calculation_rows["筛选关系"])

    def test_role_specific_web_pages_and_login(self):
        cookie = self.login("merch")
        dashboard = self.request("/dashboard", cookie=cookie)
        self.assertEqual(dashboard["status"], "200 OK")
        body = dashboard["body"].decode()
        self.assertIn("销售库存监控", body)
        self.assertIn("本期补货货号数", body)
        self.assertIn("修改频率", body)
        settings = self.request("/settings", cookie=cookie)
        self.assertIn("周二", settings["body"].decode())
        followup_cookie = self.login("followup")
        forbidden = self.request("/settings", cookie=followup_cookie)
        self.assertEqual(forbidden["status"], "403 Forbidden")

    def test_vipshop_signature_matches_official_verification_tool(self):
        params = {
            "service": "vipapis.address.AddressService",
            "method": "getFullAddress",
            "version": "1.0.0",
            "timestamp": "1406851200",
            "format": "json",
            "appKey": "yourappkey",
            "accessToken": "youraccesstoken",
        }
        body = '{"area_code":"0","is_show_gat":"SHOW_GAT","is_bind":false}'
        self.assertEqual(calculate_sign(params, body, "yourappsecret"), "25F7915F9DD6666FAD8D412349A00ED6")

    def test_vipshop_authorization_url_encodes_callback_and_state(self):
        url = build_authorization_url(
            "app-key-1",
            "https://sienna.tiger8.com.cn/oauth/vipshop/callback",
            "state-value",
        )
        parsed = urlsplit(url)
        self.assertEqual(
            (parsed.scheme, parsed.netloc, parsed.path),
            ("https", "auth.vip.com", "/oauth2/authorize"),
        )
        query = parse_qs(parsed.query)
        self.assertEqual(query["client_id"], ["app-key-1"])
        self.assertEqual(query["response_type"], ["code"])
        self.assertEqual(
            query["redirect_uri"],
            ["https://sienna.tiger8.com.cn/oauth/vipshop/callback"],
        )
        self.assertEqual(query["state"], ["state-value"])

    def test_vipshop_token_exchange_posts_secret_and_validates_token(self):
        with patch(
            "replenishment_center.vipshop._oauth_post",
            side_effect=[
                {
                    "access_token": "access-token",
                    "refresh_token": "refresh-token",
                    "expires_in": 7776000,
                    "refresh_expires_time": 1790000000000,
                    "open_id": "OPEN-1",
                },
                {"access_token": "access-token", "open_id": "OPEN-1", "expires_in": 7776000},
            ],
        ) as oauth_post:
            result = exchange_authorization_code(
                app_key="app-key",
                app_secret="app-secret",
                code="single-use-code",
                redirect_uri="https://sienna.tiger8.com.cn/oauth/vipshop/callback",
                request_client_ip="203.0.113.5",
            )
        token_call = oauth_post.call_args_list[0]
        self.assertEqual(token_call.args[0], "https://auth.vip.com/oauth2/token")
        self.assertEqual(token_call.args[1]["client_secret"], "app-secret")
        self.assertEqual(token_call.args[1]["grant_type"], "authorization_code")
        self.assertEqual(result["open_id"], "OPEN-1")
        self.assertEqual(oauth_post.call_args_list[1].args[1], {"access_token": "access-token"})

    def test_vipshop_oauth_start_and_public_callback_complete_once(self):
        user = db.authenticate(self.db_path, "merch", "demo123")
        bnx = db.get_store(self.db_path, "VIP-BNX")
        db.save_api_config(
            self.db_path,
            {
                "environment": "production",
                "app_key": "bnx-app-key",
                "app_secret": "bnx-app-secret",
                "expected_store_name": "BNX",
            },
            user["id"],
            bnx["id"],
        )
        cookie = self.login("merch")
        with patch.dict(
            "os.environ",
            {"REPLENISH_PUBLIC_BASE_URL": "https://sienna.tiger8.com.cn"},
            clear=False,
        ):
            start = self.request(
                "/oauth/vipshop/start?store=VIP-BNX", method="POST", cookie=cookie
            )
        self.assertEqual(start["status"], "302 Found")
        authorize_url = dict(start["headers"])["Location"]
        authorize_query = parse_qs(urlsplit(authorize_url).query)
        state = authorize_query["state"][0]
        self.assertEqual(
            authorize_query["redirect_uri"],
            ["https://sienna.tiger8.com.cn/oauth/vipshop/callback"],
        )
        with db.get_connection(self.db_path) as connection:
            raw_state = connection.execute(
                "SELECT * FROM vipshop_oauth_states WHERE store_id = ?", (bnx["id"],)
            ).fetchone()
        self.assertNotEqual(raw_state["state_hash"], state)
        token_payload = {
            "access_token": "oauth-access-token",
            "refresh_token": "oauth-refresh-token",
            "expires_in": 7776000,
            "refresh_expires_time": 1790000000000,
            "open_id": "BNX-OPEN-ID",
            "token_info": {
                "access_token": "oauth-access-token",
                "open_id": "BNX-OPEN-ID",
            },
        }
        verification = {
            "ok": True,
            "credentials_complete": True,
            "gateway": {},
            "message": "鉴权成功",
            "store": {"store_name": "BNX唯品会官方店"},
        }
        with patch(
            "replenishment_center.web.exchange_authorization_code",
            return_value=token_payload,
        ) as exchange, patch(
            "replenishment_center.web.test_configured_connection",
            return_value=verification,
        ):
            callback = self.request(
                f"/oauth/vipshop/callback?{urlencode({'code': 'auth-code', 'state': state})}"
            )
        self.assertEqual(callback["status"], "303 See Other")
        self.assertEqual(dict(callback["headers"])["Cache-Control"], "no-store")
        self.assertEqual(dict(callback["headers"])["Referrer-Policy"], "no-referrer")
        self.assertEqual(
            dict(callback["headers"])["Location"],
            "/oauth/vipshop/result?store=VIP-BNX&result=success",
        )
        self.assertEqual(exchange.call_args.kwargs["request_client_ip"], "127.0.0.1")
        config = db.get_api_config(self.db_path, bnx["id"], include_secrets=True)
        self.assertEqual(config["access_token"], "oauth-access-token")
        self.assertEqual(config["refresh_token"], "oauth-refresh-token")
        self.assertEqual(config["open_id"], "BNX-OPEN-ID")
        self.assertTrue(config["access_token_expires_at"])
        self.assertTrue(config["refresh_token_expires_at"])
        with db.get_connection(self.db_path) as connection:
            raw = connection.execute(
                "SELECT access_token_enc, refresh_token_enc FROM vipshop_store_api_config WHERE store_id = ?",
                (bnx["id"],),
            ).fetchone()
        self.assertNotIn("oauth-access-token", raw["access_token_enc"])
        self.assertNotIn("oauth-refresh-token", raw["refresh_token_enc"])
        replay = self.request(
            f"/oauth/vipshop/callback?{urlencode({'code': 'auth-code', 'state': state})}"
        )
        self.assertEqual(replay["status"], "303 See Other")
        self.assertEqual(
            dict(replay["headers"])["Location"],
            "/oauth/vipshop/result?result=invalid",
        )
        result_page = self.request(
            "/oauth/vipshop/result?store=VIP-BNX&result=success"
        )
        self.assertEqual(result_page["status"], "200 OK")
        self.assertEqual(dict(result_page["headers"])["Cache-Control"], "no-store")
        self.assertIn("授权与店铺校验成功", result_page["body"].decode())

    def test_vipshop_oauth_callback_never_falls_through_to_login(self):
        missing_state = self.request("/oauth/vipshop/callback?code=TEST")
        self.assertEqual(missing_state["status"], "303 See Other")
        self.assertEqual(
            dict(missing_state["headers"])["Location"],
            "/oauth/vipshop/result?result=invalid",
        )

    def test_vipshop_credentials_are_encrypted_and_masked(self):
        user = db.authenticate(self.db_path, "merch", "demo123")
        db.save_api_config(
            self.db_path,
            {
                "environment": "production",
                "app_key": "app-key-1234",
                "app_secret": "secret-value",
                "access_token": "token-value",
                "expected_store_name": "马天奴",
            },
            user["id"],
        )
        config = db.get_api_config(self.db_path, include_secrets=True)
        self.assertTrue(config["credentials_complete"])
        self.assertEqual(config["app_secret"], "secret-value")
        self.assertEqual(config["access_token"], "token-value")
        with db.get_connection(self.db_path) as connection:
            store = db.get_store(self.db_path, "VIP-MTN")
            raw = connection.execute(
                "SELECT app_secret_enc, access_token_enc FROM vipshop_store_api_config WHERE store_id = ?",
                (store["id"],),
            ).fetchone()
        self.assertNotIn("secret-value", raw["app_secret_enc"])
        self.assertNotIn("token-value", raw["access_token_enc"])
        self.assertTrue((self.db_path.with_suffix(".key")).exists())

    def test_vipshop_fake_single_store_sync_normalizes_sales_and_inventory(self):
        class FakeClient:
            config = VipshopConfig("production", "key", "secret", "token", "马天奴")

            def call(self, service, method, payload=None, version="1.0.0"):
                if method == "getStoreInfo":
                    return {"success": {"store_id": "STORE-1", "seller_id": "SELLER-1", "store_name": "马天奴官方店", "store_status": "2"}}
                if method == "getProducts":
                    return {"success": {"products": [{"spu_id": "SPU-1", "outer_spu_id": "REAL-001", "title": "测试针织衫"}], "has_next": False}}
                if method == "getProductById":
                    return {"success": {"spu_id": "SPU-1", "outer_spu_id": "REAL-001", "title": "测试针织衫", "skus": [{"sku_id": "SKU-1", "outer_sku_id": "BAR-1", "sale_props": {"颜色": "黑色", "尺码": "M"}}]}}
                if method == "getOrders":
                    return {"success": {"total": 2, "orders": [{"order_id": "ORDER-1", "status": "25", "created": "2026-07-23 09:00:00"}, {"order_id": "ORDER-2", "status": "97", "created": "2026-07-23 10:00:00"}]}}
                if method == "getOrderDetail":
                    return {"success": [{"order_id": "ORDER-1", "order_products": [{"sku_id": "SKU-1", "spu_id": "SPU-1", "outer_spu_id": "REAL-001", "outer_sku_id": "BAR-1", "num": "3", "title": "测试针织衫", "size": "M", "color": "黑色"}]}, {"order_id": "ORDER-2", "order_products": [{"sku_id": "SKU-1", "num": "8", "size": "M", "color": "黑色"}]}]}
                if method == "getSkuStock":
                    return {"success": {"sku_stocks": [{"sku_id": "SKU-1", "leaving_stock": 12, "cart_hold": 2, "order_hold": 1, "warehouse": "W1"}]}}
                raise AssertionError((service, method, payload))

        normalized = collect_store_data(FakeClient(), as_of=datetime(2026, 7, 23).date())
        self.assertEqual(normalized["order_count"], 2)
        self.assertEqual(normalized["sales"][0]["gross_units"], 3)
        counts = db.apply_vipshop_data(self.db_path, normalized)
        self.assertEqual(counts["sku"], 1)
        with db.get_connection(self.db_path) as connection:
            sku = connection.execute("SELECT * FROM skus WHERE external_sku_id = 'SKU-1'").fetchone()
            inventory = connection.execute("SELECT * FROM inventory_current WHERE sku_id = ?", (sku["id"],)).fetchone()
            sale = connection.execute("SELECT * FROM sales_daily WHERE sku_id = ? AND sale_date = '2026-07-23'", (sku["id"],)).fetchone()
            active_demo = connection.execute("SELECT COUNT(*) AS count FROM skus WHERE is_demo = 1 AND lifecycle = 'active'").fetchone()["count"]
        self.assertEqual((sku["style_code"], sku["color_name"], sku["size_name"]), ("REAL-001", "黑色", "M"))
        self.assertEqual((inventory["on_hand"], inventory["locked"], inventory["source"]), (15, 3, "vipshop_api"))
        self.assertEqual((sale["gross_units"], sale["source"]), (3, "vipshop_api"))
        self.assertEqual(active_demo, 0)

    def test_api_configuration_page_is_available_to_merchandise_only(self):
        cookie = self.login("merch")
        page = self.request("/data/api", cookie=cookie)
        self.assertEqual(page["status"], "200 OK")
        self.assertIn("唯品会 API 配置", page["body"].decode())
        followup_cookie = self.login("followup")
        forbidden = self.request("/data/api", cookie=followup_cookie)
        self.assertEqual(forbidden["status"], "403 Forbidden")

    def test_store_workspaces_are_seeded_and_isolated(self):
        stores = db.list_stores(self.db_path)
        self.assertEqual(
            [(store["store_code"], store["store_name"]) for store in stores],
            [
                ("VIP-MTN", "马天奴唯品会"),
                ("TMALL-MTN-FLAGSHIP", "马天奴天猫官方旗舰店"),
                ("VIP-BNX", "BNX唯品会"),
            ],
        )
        tmall = db.get_store(self.db_path, "TMALL-MTN-FLAGSHIP")
        self.assertEqual(db.list_plans(self.db_path, store_id=tmall["id"]), [])
        self.assertIsNone(db.dashboard_data(self.db_path, store_id=tmall["id"])["current"])

    def test_bnx_vipshop_credentials_are_encrypted_and_isolated(self):
        user = db.authenticate(self.db_path, "merch", "demo123")
        mtn = db.get_store(self.db_path, "VIP-MTN")
        bnx = db.get_store(self.db_path, "VIP-BNX")
        mtn_before = db.get_api_config(self.db_path, mtn["id"])
        db.save_api_config(
            self.db_path,
            {
                "environment": "production",
                "app_key": "bnx-app-key",
                "app_secret": "bnx-secret-value",
                "access_token": "bnx-token-value",
                "expected_store_name": "BNX",
            },
            user["id"],
            bnx["id"],
        )
        config = db.get_api_config(self.db_path, bnx["id"], include_secrets=True)
        self.assertTrue(config["credentials_complete"])
        self.assertEqual(config["app_secret"], "bnx-secret-value")
        self.assertEqual(config["access_token"], "bnx-token-value")
        self.assertEqual(config["store_code"], "VIP-BNX")
        self.assertEqual(
            db.get_api_config(self.db_path, mtn["id"])["app_key"],
            mtn_before["app_key"],
        )
        with db.get_connection(self.db_path) as connection:
            raw = connection.execute(
                "SELECT app_secret_enc, access_token_enc FROM vipshop_store_api_config WHERE store_id = ?",
                (bnx["id"],),
            ).fetchone()
        self.assertNotIn("bnx-secret-value", raw["app_secret_enc"])
        self.assertNotIn("bnx-token-value", raw["access_token_enc"])

    def test_bnx_missing_credentials_probes_vop_gateway_only_for_bnx(self):
        mtn = db.get_store(self.db_path, "VIP-MTN")
        bnx = db.get_store(self.db_path, "VIP-BNX")
        mtn_status = db.get_api_config(self.db_path, mtn["id"])["last_test_status"]
        gateway = {
            "ok": True,
            "gateway": "https://vop.vipapis.com/",
            "latency_ms": 18,
            "message": "唯品会 VOP 正式网关可达",
        }
        with patch("replenishment_center.vipshop.probe_gateway", return_value=gateway):
            result = test_vipshop_connection(self.db_path, bnx["id"])
        self.assertFalse(result["ok"])
        self.assertFalse(result["credentials_complete"])
        self.assertIn("正式网关可达", result["message"])
        self.assertEqual(
            db.get_api_config(self.db_path, bnx["id"])["last_test_status"],
            "credentials_missing",
        )
        self.assertEqual(
            db.get_api_config(self.db_path, mtn["id"])["last_test_status"],
            mtn_status,
        )

    def test_bnx_api_data_can_generate_an_isolated_plan(self):
        bnx = db.get_store(self.db_path, "VIP-BNX")
        mtn_plan = db.list_plans(self.db_path, store_id=db.get_store(self.db_path, "VIP-MTN")["id"])[0]
        as_of = datetime.now().date()
        normalized = {
            "store": {"store_id": "BNX-STORE", "seller_id": "BNX-SELLER", "store_name": "BNX唯品会官方店"},
            "skus": [{
                "external_sku_id": "BNX-SKU-1", "external_spu_id": "BNX-SPU-1",
                "outer_sku_id": "BNX-GOODS-1", "style_code": "BNX-STYLE-1",
                "style_name": "BNX测试款", "color_name": "黑色", "size_name": "M",
            }],
            "sales": [
                {
                    "external_sku_id": "BNX-SKU-1",
                    "sale_date": (as_of - timedelta(days=offset)).isoformat(),
                    "gross_units": 1,
                    "return_units": 0,
                }
                for offset in range(14)
            ],
            "inventory": [{
                "external_sku_id": "BNX-SKU-1", "on_hand": 0, "locked": 0,
                "defective": 0, "inbound": 0, "inbound_date": "",
            }],
            "sales_through_date": as_of.isoformat(),
            "inventory_snapshot_at": f"{as_of.isoformat()} 10:00:00",
            "order_count": 14,
            "product_count": 1,
        }
        counts = db.apply_vipshop_data(self.db_path, normalized, store_id=bnx["id"])
        self.assertEqual((counts["sku"], counts["sales"], counts["inventory"]), (1, 14, 1))
        plan_id = db.generate_plan(
            self.db_path, generation_type="manual_api", force=True, store_id=bnx["id"]
        )
        plan = db.get_plan(self.db_path, plan_id)
        self.assertTrue(plan["plan_no"].startswith("VIP-BNX-"))
        self.assertEqual(plan["store_id"], bnx["id"])
        self.assertTrue(db.get_plan_items(self.db_path, plan_id))
        self.assertEqual(db.get_plan(self.db_path, mtn_plan["id"])["status"], mtn_plan["status"])

    def test_bnx_workspace_and_api_page_are_isolated(self):
        cookie = self.login("merch")
        dashboard = self.request("/dashboard?store=VIP-BNX", cookie=cookie)
        body = dashboard["body"].decode()
        self.assertEqual(dashboard["status"], "200 OK")
        self.assertIn("BNX唯品会", body)
        self.assertIn("BNX 唯品会 API 单店联调", body)
        self.assertIn("马天奴天猫官方旗舰店", body)
        self.assertNotIn("修改频率", body)
        self.assertNotIn("浏览器报表采集", body)

        api_page = self.request("/data/api?store=VIP-BNX", cookie=cookie)
        api_body = api_page["body"].decode()
        self.assertEqual(api_page["status"], "200 OK")
        self.assertIn("BNX唯品会 API 配置", api_body)
        self.assertIn('value="VIP-BNX"', api_body)
        self.assertIn("AccessToken", api_body)

        plans = self.request("/plans?store=VIP-BNX", cookie=cookie)
        self.assertIn("暂无补货计划", plans["body"].decode())
        self.assertNotIn("VIP-MTN-", plans["body"].decode())

        followup_cookie = self.login("followup")
        forbidden = self.request("/data/api?store=VIP-BNX", cookie=followup_cookie)
        self.assertEqual(forbidden["status"], "403 Forbidden")

    def test_tmall_signature_matches_known_hmac_md5_vector(self):
        params = {
            "app_key": "12345678",
            "method": "taobao.shop.seller.get",
            "timestamp": "2026-07-28 10:00:00",
            "v": "2.0",
            "format": "json",
            "sign_method": "hmac",
            "session": "session-value",
        }
        self.assertEqual(
            calculate_tmall_sign(params, "secret-value"),
            "2ED1F260984F7CCE0915C0CFB7E9617C",
        )

    def test_tmall_credentials_are_encrypted_and_masked(self):
        user = db.authenticate(self.db_path, "merch", "demo123")
        store = db.get_store(self.db_path, "TMALL-MTN-FLAGSHIP")
        db.save_tmall_api_config(
            self.db_path,
            store["id"],
            {
                "environment": "production",
                "app_key": "tmall-app-key",
                "app_secret": "tmall-secret-value",
                "session_key": "tmall-session-value",
                "expected_store_name": "马天奴天猫官方旗舰店",
            },
            user["id"],
        )
        config = db.get_tmall_api_config(self.db_path, store["id"])
        self.assertTrue(config["credentials_complete"])
        self.assertNotIn("app_secret", config)
        self.assertNotIn("session_key", config)
        self.assertTrue(config["app_key_masked"].endswith("-key"))
        with db.get_connection(self.db_path) as connection:
            raw = connection.execute(
                "SELECT app_secret_enc, session_key_enc FROM tmall_api_config WHERE store_id = ?",
                (store["id"],),
            ).fetchone()
        self.assertNotIn("tmall-secret-value", raw["app_secret_enc"])
        self.assertNotIn("tmall-session-value", raw["session_key_enc"])

    def test_tmall_missing_credentials_probes_gateway_without_authentication(self):
        store = db.get_store(self.db_path, "TMALL-MTN-FLAGSHIP")
        gateway = {
            "ok": True,
            "gateway": "https://eco.taobao.com/router/rest",
            "latency_ms": 25,
            "message": "淘宝开放平台官方网关可达",
        }
        with patch("replenishment_center.tmall.probe_gateway", return_value=gateway):
            result = test_tmall_connection(self.db_path, store["id"])
        self.assertFalse(result["ok"])
        self.assertFalse(result["credentials_complete"])
        self.assertIn("官方网关可达", result["message"])
        self.assertEqual(
            db.get_tmall_api_config(self.db_path, store["id"])["last_test_status"],
            "credentials_missing",
        )

    def test_tmall_workspace_pages_and_roles(self):
        cookie = self.login("merch")
        dashboard = self.request(
            "/dashboard?store=TMALL-MTN-FLAGSHIP", cookie=cookie
        )
        body = dashboard["body"].decode()
        self.assertEqual(dashboard["status"], "200 OK")
        self.assertIn("马天奴唯品会", body)
        self.assertIn("马天奴天猫官方旗舰店", body)
        self.assertIn("天猫 API 联调", body)
        self.assertNotIn("修改频率", body)

        api_page = self.request(
            "/data/tmall-api?store=TMALL-MTN-FLAGSHIP", cookie=cookie
        )
        self.assertEqual(api_page["status"], "200 OK")
        self.assertIn("天猫 API 配置与试连", api_page["body"].decode())
        self.assertIn("SessionKey", api_page["body"].decode())

        plans = self.request("/plans?store=TMALL-MTN-FLAGSHIP", cookie=cookie)
        plans_body = plans["body"].decode()
        self.assertIn("暂无补货计划", plans_body)
        self.assertNotIn("VIP-MTN-", plans_body)

        followup_cookie = self.login("followup")
        forbidden = self.request(
            "/data/tmall-api?store=TMALL-MTN-FLAGSHIP", cookie=followup_cookie
        )
        self.assertEqual(forbidden["status"], "403 Forbidden")

    def test_browser_capture_defaults_and_role_access(self):
        config = db.get_browser_capture_config(self.db_path)
        self.assertEqual(config["backend_url"], "https://vis.vip.com/")
        self.assertEqual(config["debug_port"], 9223)
        self.assertTrue(config["profile_dir"].endswith("vipshop_chrome_profile"))
        cookie = self.login("merch")
        page = self.request("/data/browser", cookie=cookie)
        self.assertEqual(page["status"], "200 OK")
        self.assertIn("唯品浏览器报表采集", page["body"].decode())
        followup_cookie = self.login("followup")
        forbidden = self.request("/data/browser", cookie=followup_cookie)
        self.assertEqual(forbidden["status"], "403 Forbidden")

    def test_browser_capture_accepts_only_official_report_hosts(self):
        user = db.authenticate(self.db_path, "merch", "demo123")
        db.save_browser_report_url(
            self.db_path,
            "sales",
            "https://compass.vip.com/frontend/index.html#/product/details",
            user["id"],
        )
        self.assertIn("compass.vip.com", db.get_browser_capture_config(self.db_path)["sales_report_url"])
        with self.assertRaisesRegex(ValueError, "官方报表"):
            db.save_browser_report_url(
                self.db_path,
                "inventory",
                "https://compass.vip.com.example.com/report",
                user["id"],
            )

    def test_browser_capture_archives_hashes_and_validates_xlsx(self):
        user = db.authenticate(self.db_path, "merch", "demo123")
        job_id = db.create_browser_capture_job(self.db_path, "sales", user["id"])
        download_dir = browser_paths(self.db_path)["download_dir"]
        download_dir.mkdir(parents=True, exist_ok=True)
        report = download_dir / "唯品销售明细.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["销售日期", "款号", "商家SKU编码", "颜色", "尺码", "销售数量"])
        sheet.append(["2026-07-22", "MTN260701", "SKU-001", "雾灰", "M", 3])
        workbook.save(report)

        self.assertEqual(db.process_browser_capture_jobs(self.db_path), [job_id])
        job = db.list_browser_capture_jobs(self.db_path)[0]
        archive = Path(job["archive_file"])
        self.assertEqual(job["status"], "ready_for_import")
        self.assertEqual(job["analysis"]["row_count"], 1)
        self.assertEqual(job["analysis"]["missing"], [])
        self.assertTrue(archive.exists())
        self.assertEqual(job["file_sha256"], sha256_file(archive))

    def test_browser_capture_matches_inventory_file_to_inventory_job(self):
        user = db.authenticate(self.db_path, "merch", "demo123")
        sales_job = db.create_browser_capture_job(self.db_path, "sales", user["id"])
        inventory_job = db.create_browser_capture_job(self.db_path, "inventory", user["id"])
        download_dir = browser_paths(self.db_path)["download_dir"]
        download_dir.mkdir(parents=True, exist_ok=True)
        report = download_dir / "唯品库存明细.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["款号", "商家SKU编码", "颜色", "尺码", "可售库存"])
        sheet.append(["MTN260701", "SKU-001", "雾灰", "M", 12])
        workbook.save(report)

        self.assertEqual(db.process_browser_capture_jobs(self.db_path), [inventory_job])
        jobs = {job["id"]: job for job in db.list_browser_capture_jobs(self.db_path)}
        self.assertEqual(jobs[inventory_job]["status"], "ready_for_import")
        self.assertEqual(jobs[sales_job]["status"], "waiting_download")

    def test_browser_capture_accepts_joint_report_without_color_for_both_jobs(self):
        user = db.authenticate(self.db_path, "merch", "demo123")
        sales_job = db.create_browser_capture_job(self.db_path, "sales", user["id"])
        inventory_job = db.create_browser_capture_job(self.db_path, "inventory", user["id"])
        download_dir = browser_paths(self.db_path)["download_dir"]
        download_dir.mkdir(parents=True, exist_ok=True)
        report = download_dir / "唯品商品明细-条码粒度.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["日期", "款号", "条码", "尺码名称", "销售量", "在售库存"])
        sheet.append(["2026-07-27", "MTN260701", "BAR-001", "M", 3, 12])
        workbook.save(report)

        self.assertEqual(db.process_browser_capture_jobs(self.db_path), [sales_job, inventory_job])
        jobs = {job["id"]: job for job in db.list_browser_capture_jobs(self.db_path)}
        self.assertEqual(jobs[sales_job]["status"], "ready_for_import")
        self.assertEqual(jobs[inventory_job]["status"], "ready_for_import")
        self.assertEqual(jobs[sales_job]["file_sha256"], jobs[inventory_job]["file_sha256"])

    def test_browser_reports_join_master_merge_barcodes_and_apply_real_data(self):
        report = Path(self.temp_dir.name) / "report.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "日期", "条码", "商品ID", "商品名称", "货号", "款号", "尺码名称",
            "品牌名称", "销售量", "在售库存",
        ])
        start = datetime(2026, 7, 9).date()
        for offset in range(14):
            current = start + timedelta(days=offset)
            sheet.append([current, "BAR-1", "VSKU-1", "测试连衣裙", "OUT-1", "REAL-001", "M", "马天奴MYTENO", 1, 12])
            sheet.append([current, "BAR-2", "VSKU-2", "测试连衣裙", "OUT-2", "REAL-001", "M", "马天奴MYTENO", 1, 8])
        workbook.save(report)

        master = Path(self.temp_dir.name) / "master.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "商品资料"
        sheet.append([
            "V_SPU", "V_SKU", "品牌名称", "款号", "货号", "条形码", "商品名称", "商品类目",
            "标准尺码", "自定义尺码", "标准颜色", "自定义颜色",
        ])
        sheet.append(["SPU-1", "VSKU-1", "马天奴MYTENO", "REAL-001", "OUT-1", "BAR-1", "测试连衣裙", "连衣裙", "M", "M", "红色", "酒红"])
        sheet.append(["SPU-1", "VSKU-2", "马天奴MYTENO", "REAL-001", "OUT-2", "BAR-2", "测试连衣裙", "连衣裙", "M", "M", "红色", "酒红"])
        workbook.save(master)

        normalized = normalize_browser_reports(report, master)
        self.assertEqual(normalized["stats"]["date_count"], 14)
        self.assertEqual(normalized["stats"]["matched_barcodes"], 2)
        self.assertEqual(normalized["stats"]["sku_count"], 1)
        self.assertEqual(normalized["stats"]["merged_barcodes"], 1)
        self.assertEqual(normalized["sales"][0]["gross_units"], 2)
        self.assertEqual(normalized["inventory"][0]["on_hand"], 20)
        self.assertEqual(normalized["skus"][0]["color_name"], "酒红")

        user = db.authenticate(self.db_path, "merch", "demo123")
        counts = db.apply_browser_report_data(self.db_path, normalized, user["id"])
        self.assertEqual((counts["sku"], counts["sales"], counts["inventory"]), (1, 14, 1))
        self.assertEqual(db.get_settings(self.db_path)["data_source_mode"], "browser")
        self.assertEqual(db.get_settings(self.db_path)["api_status"], "unconfigured")
        with db.get_connection(self.db_path) as connection:
            sku = connection.execute("SELECT * FROM skus WHERE external_sku_id = 'BAR-1'").fetchone()
            inventory = connection.execute("SELECT * FROM inventory_current WHERE sku_id = ?", (sku["id"],)).fetchone()
            sale = connection.execute(
                "SELECT * FROM sales_daily WHERE sku_id = ? AND sale_date = '2026-07-09'", (sku["id"],)
            ).fetchone()
        self.assertEqual((sku["style_code"], sku["color_name"], sku["size_name"]), ("REAL-001", "酒红", "M"))
        self.assertEqual((inventory["on_hand"], inventory["source"]), (20, "vipshop_browser"))
        self.assertEqual((sale["gross_units"], sale["source"]), (2, "vipshop_browser"))

        with db.get_connection(self.db_path) as connection:
            cursor = connection.execute(
                """
                INSERT INTO skus(
                    store_id, style_code, style_name, color_name, size_name, category, supplier,
                    lead_time_days, moq, pack_size, default_size_share, core_size, lifecycle,
                    external_sku_id, external_spu_id, outer_sku_id, is_demo
                ) VALUES (1, 'STALE-001', '旧款', '黑色', 'M', '', '', 14, 0, 1, 0, 1,
                    'active', 'STALE-BAR', '', 'STALE-OUT', 0)
                """
            )
            stale_id = int(cursor.lastrowid)
            connection.execute(
                """
                INSERT INTO sales_daily(sku_id, sale_date, gross_units, return_units, net_units, source)
                VALUES (?, '2026-07-22', 9, 0, 9, 'vipshop_browser')
                """,
                (stale_id,),
            )
        db.apply_browser_report_data(self.db_path, normalized, user["id"])
        with db.get_connection(self.db_path) as connection:
            stale = connection.execute("SELECT lifecycle FROM skus WHERE id = ?", (stale_id,)).fetchone()
            stale_sale = connection.execute(
                "SELECT id FROM sales_daily WHERE sku_id = ? AND sale_date = '2026-07-22'", (stale_id,)
            ).fetchone()
        self.assertEqual(stale["lifecycle"], "inactive")
        self.assertIsNone(stale_sale)

        previous_plan_id = db.list_plans(self.db_path)[0]["id"]
        plan_id = db.generate_plan(self.db_path, generation_type="browser_report", created_by=user["id"], force=True)
        plan = db.get_plan(self.db_path, plan_id)
        item = db.get_plan_items(self.db_path, plan_id)[0]
        self.assertEqual(db.get_plan(self.db_path, previous_plan_id)["status"], "superseded")
        self.assertEqual(plan["sales_through_date"], "2026-07-22")
        self.assertEqual((plan["min_sales_7"], plan["min_sales_14"], plan["max_coverage_days"]), (5, 10, 14))
        self.assertEqual((item["sales_7"], item["sales_14"]), (14, 28))
        self.assertTrue(all(row["suggested_qty"] > 0 for row in db.get_plan_items(self.db_path, plan_id)))


if __name__ == "__main__":
    unittest.main()
